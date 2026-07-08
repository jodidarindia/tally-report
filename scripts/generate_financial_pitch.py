"""FLOWRA — Financial Pitch (Feb 2026)

Generates two deliverables in /app/frontend/public/pitch/ :

  • financial_pitch_flowra.pdf         (16-page investor pitch, reportlab)
  • financial_projections_flowra.xlsx  (editable Excel model, openpyxl)

Inputs (locked by founder):
  - Pricing (₹/company/mo): Free 0 · Starter 833 · Professional 2083 · Enterprise 3166
  - Current traction: 1 paying customer (Krishna Sales, Enterprise)
  - Y5 ARR target: ₹5 Cr (conservative)
  - Team plan: 12 by Y1 → 25 by Y5 (capital-efficient)
  - Two-tranche fundraise: Seed ₹2.5 Cr now → Series A ₹6 Cr at Month 24
  - Brand: navy #0F1B4C / blue #2563EB / amber #F59E0B
"""
from pathlib import Path
from datetime import date
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

from reportlab.lib import colors as rc
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, PageBreak,
    Table, TableStyle, Image, KeepTogether,
)
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.graphics.shapes import Drawing, Rect, String, Line
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.legends import Legend
from reportlab.graphics import renderPDF

# Register a Unicode-capable font family so the ₹ glyph (U+20B9) renders.
# Helvetica (reportlab default) has no rupee sign → it draws a black square.
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
_FONT_DIR = "/root/.venv/lib/python3.11/site-packages/matplotlib/mpl-data/fonts/ttf"
try:
    pdfmetrics.registerFont(TTFont("Sans",  f"{_FONT_DIR}/DejaVuSans.ttf"))
    pdfmetrics.registerFont(TTFont("SansB", f"{_FONT_DIR}/DejaVuSans-Bold.ttf"))
    pdfmetrics.registerFont(TTFont("SansO", f"{_FONT_DIR}/DejaVuSans-Oblique.ttf"))
    pdfmetrics.registerFont(TTFont("SansBO", f"{_FONT_DIR}/DejaVuSans-BoldOblique.ttf"))
    from reportlab.pdfbase.pdfmetrics import registerFontFamily
    registerFontFamily("Sans", normal="Sans", bold="SansB",
                        italic="SansO", boldItalic="SansBO")
    FONT_NORMAL, FONT_BOLD, FONT_ITALIC = "Sans", "SansB", "SansO"
except Exception:
    # Ultimate fallback — glyphs may render as boxes.
    FONT_NORMAL, FONT_BOLD, FONT_ITALIC = FONT_NORMAL, FONT_BOLD, FONT_ITALIC

# ── Brand ────────────────────────────────────────────────────────────────
NAVY  = rc.HexColor("#0F1B4C")
BLUE  = rc.HexColor("#2563EB")
AMBER = rc.HexColor("#F59E0B")
GREEN = rc.HexColor("#10B981")
RED   = rc.HexColor("#EF4444")
GREY  = rc.HexColor("#64748B")
LIGHT = rc.HexColor("#F1F5F9")
SOFT  = rc.HexColor("#F0F4FF")
PAPER = rc.HexColor("#FFFFFF")
INK   = rc.HexColor("#0F172A")

OUT_DIR = Path("/app/frontend/public/pitch")
OUT_DIR.mkdir(parents=True, exist_ok=True)
PDF_PATH  = OUT_DIR / "financial_pitch_flowra.pdf"
XLSX_PATH = OUT_DIR / "financial_projections_flowra.xlsx"

COMPANY = "JODIDAR INDIA"
PRODUCT = "FLOWRA"
FOUNDER = "Ankit Sarawgi"
CITY    = "Raipur, Chhattisgarh · India"
TAG     = "Organize. Automate. Accelerate."
DATED   = "February 2026"

# ── Assumptions (single source of truth — feeds BOTH PDF and Excel) ─────
ASSUMPTIONS = {
    "prices": {         # ₹ / company / month
        "starter":      833,
        "professional": 2083,
        "enterprise":   3166,
    },
    "plan_mix_paid": {   # % of PAID base (must sum to 100)
        "starter":      45,
        "professional": 40,
        "enterprise":   15,
    },
    "customers_end_of_year": {  # paid customers at Mar-31 of that FY
        "Y1_FY26_27": 100,
        "Y2_FY27_28": 400,
        "Y3_FY28_29": 900,
        "Y4_FY29_30": 1500,
        "Y5_FY30_31": 2000,
        "Y6_FY31_32": 2700,
    },
    "team_end_of_year": {   # headcount by Mar-31
        "Y1_FY26_27": 12,
        "Y2_FY27_28": 15,
        "Y3_FY28_29": 18,
        "Y4_FY29_30": 22,
        "Y5_FY30_31": 25,
        "Y6_FY31_32": 30,
    },
    "avg_monthly_cost_per_head_lakh": {   # ₹ Lakh (blended fully-loaded)
        "Y1_FY26_27": 0.66,   # ~₹66,000 (founders + jr team, Tier-2 salary)
        "Y2_FY27_28": 0.75,
        "Y3_FY28_29": 0.80,
        "Y4_FY29_30": 0.85,
        "Y5_FY30_31": 0.90,
        "Y6_FY31_32": 0.95,
    },
    # Other opex — ₹ Lakh per YEAR
    "marketing_lakh": {"Y1_FY26_27":  20, "Y2_FY27_28": 30, "Y3_FY28_29": 40,
                        "Y4_FY29_30":  50, "Y5_FY30_31": 55, "Y6_FY31_32": 60},
    "infra_lakh":     {"Y1_FY26_27":  12, "Y2_FY27_28": 20, "Y3_FY28_29": 30,
                        "Y4_FY29_30":  42, "Y5_FY30_31": 55, "Y6_FY31_32": 70},
    "other_lakh":     {"Y1_FY26_27":  15, "Y2_FY27_28": 20, "Y3_FY28_29": 25,
                        "Y4_FY29_30":  32, "Y5_FY30_31": 40, "Y6_FY31_32": 48},
    "cac_rupees":     {"Y1_FY26_27": 5000, "Y2_FY27_28": 6000, "Y3_FY28_29": 7000,
                        "Y4_FY29_30": 8000, "Y5_FY30_31": 8000, "Y6_FY31_32": 8000},
    "monthly_churn_pct": 2.5,
    "gross_margin_pct":  82,
    "arpu_upsell_pct":   6,     # avg ARPU grows this much per year via upsell
    "seed_amount_cr":    2.5,
    "series_a_amount_cr": 6.0,
    "seed_dilution_pct":  18,
    "series_a_dilution_pct": 18,
    "exit_multiple_arr": 5,     # 5x ARR multiple at Year-5 exit
}

FYS = [
    ("Y1_FY26_27", "FY26-27"),
    ("Y2_FY27_28", "FY27-28"),
    ("Y3_FY28_29", "FY28-29"),
    ("Y4_FY29_30", "FY29-30"),
    ("Y5_FY30_31", "FY30-31"),
    ("Y6_FY31_32", "FY31-32"),
]


# ── Derived numbers (used in BOTH PDF and Excel) ─────────────────────────
def blended_arpu_monthly(a=ASSUMPTIONS):
    p = a["prices"]; m = a["plan_mix_paid"]
    return (p["starter"] * m["starter"] +
            p["professional"] * m["professional"] +
            p["enterprise"] * m["enterprise"]) / 100.0


def compute_projections(a=ASSUMPTIONS):
    """Return list[dict] one per FY with all P&L numbers in ₹ Cr."""
    arpu_m = blended_arpu_monthly(a)
    rows = []
    prev_customers = 1   # Krishna Sales
    for i, (key, label) in enumerate(FYS):
        cust_end = a["customers_end_of_year"][key]
        cust_avg = (prev_customers + cust_end) / 2
        arpu_yr  = arpu_m * 12 * (1 + a["arpu_upsell_pct"] / 100) ** i
        revenue_cr = cust_avg * arpu_yr / 1e7          # ₹ Cr (in-year)
        arr_end_cr = cust_end * arpu_yr / 1e7          # ₹ Cr (year-end)
        team_cost_cr = (a["team_end_of_year"][key] *
                         a["avg_monthly_cost_per_head_lakh"][key] * 12) / 100
        marketing_cr = a["marketing_lakh"][key] / 100
        infra_cr     = a["infra_lakh"][key] / 100
        other_cr     = a["other_lakh"][key] / 100
        total_cost_cr = team_cost_cr + marketing_cr + infra_cr + other_cr
        ebitda_cr = revenue_cr - total_cost_cr
        rows.append({
            "key": key,
            "label": label,
            "customers_end": cust_end,
            "customers_avg": round(cust_avg),
            "arpu_yr":       round(arpu_yr),
            "arr_end_cr":    round(arr_end_cr, 2),
            "revenue_cr":    round(revenue_cr, 2),
            "team":          a["team_end_of_year"][key],
            "team_cost_cr":  round(team_cost_cr, 2),
            "marketing_cr":  round(marketing_cr, 2),
            "infra_cr":      round(infra_cr, 2),
            "other_cr":      round(other_cr, 2),
            "total_cost_cr": round(total_cost_cr, 2),
            "ebitda_cr":     round(ebitda_cr, 2),
        })
        prev_customers = cust_end
    return rows


def compute_unit_economics(a=ASSUMPTIONS):
    """Return CAC, LTV, LTV/CAC, payback (steady-state Y3)."""
    arpu_m = blended_arpu_monthly(a)
    gm = a["gross_margin_pct"] / 100
    churn_m = a["monthly_churn_pct"] / 100
    life_m = 1 / churn_m
    ltv_r = arpu_m * gm * life_m
    cac_r = a["cac_rupees"]["Y3_FY28_29"]
    return {
        "arpu_monthly_r": round(arpu_m),
        "arpu_annual_r":  round(arpu_m * 12),
        "gross_margin_pct": a["gross_margin_pct"],
        "monthly_churn_pct": a["monthly_churn_pct"],
        "customer_life_months": round(life_m, 1),
        "ltv_r": round(ltv_r),
        "cac_r": cac_r,
        "ltv_to_cac": round(ltv_r / cac_r, 1),
        "payback_months": round(cac_r / (arpu_m * gm), 1),
    }


# ═══════════════════════════════════════════════════════════════════════════
# PART 1 — PDF PITCH (16 pages)
# ═══════════════════════════════════════════════════════════════════════════

PAGE_W, PAGE_H = A4
PAGE_NUM = [0]  # mutable counter for header/footer


def _page_frame(canv: canvas.Canvas, doc):
    """Header + footer painted on every page."""
    PAGE_NUM[0] += 1
    n = PAGE_NUM[0]

    # Top strip — navy
    canv.setFillColor(NAVY)
    canv.rect(0, PAGE_H - 22 * mm, PAGE_W, 22 * mm, fill=1, stroke=0)

    # Logo mark (round chip)
    canv.setFillColor(AMBER)
    canv.circle(20 * mm, PAGE_H - 11 * mm, 5 * mm, fill=1, stroke=0)
    canv.setFillColor(NAVY)
    canv.setFont(FONT_BOLD, 13)
    canv.drawCentredString(20 * mm, PAGE_H - 12.6 * mm, "F")

    canv.setFillColor(PAPER)
    canv.setFont(FONT_BOLD, 12)
    canv.drawString(30 * mm, PAGE_H - 10 * mm, "FLOWRA · Financial Pitch")
    canv.setFont(FONT_NORMAL, 8)
    canv.setFillColor(rc.HexColor("#94A3B8"))
    canv.drawString(30 * mm, PAGE_H - 14.5 * mm, f"{COMPANY} · {DATED}")

    # Page number top-right
    canv.setFont(FONT_NORMAL, 9)
    canv.setFillColor(PAPER)
    canv.drawRightString(PAGE_W - 15 * mm, PAGE_H - 12 * mm, f"Page {n} / 16")

    # Bottom footer
    canv.setStrokeColor(rc.HexColor("#E2E8F0"))
    canv.setLineWidth(0.4)
    canv.line(15 * mm, 12 * mm, PAGE_W - 15 * mm, 12 * mm)

    canv.setFont(FONT_NORMAL, 8)
    canv.setFillColor(GREY)
    canv.drawString(15 * mm, 7 * mm,
                     f"{COMPANY}  ·  Confidential — for prospective investors")
    canv.drawRightString(PAGE_W - 15 * mm, 7 * mm, TAG)


# ── Reusable style factory ────────────────────────────────────────────────
def styles():
    return {
        "h1": ParagraphStyle(
            name="h1", fontName=FONT_BOLD, fontSize=24,
            textColor=NAVY, leading=28, spaceAfter=6),
        "h2": ParagraphStyle(
            name="h2", fontName=FONT_BOLD, fontSize=15,
            textColor=BLUE, leading=19, spaceAfter=4),
        "h3": ParagraphStyle(
            name="h3", fontName=FONT_BOLD, fontSize=11,
            textColor=NAVY, leading=15, spaceAfter=3),
        "body": ParagraphStyle(
            name="body", fontName=FONT_NORMAL, fontSize=10,
            textColor=INK, leading=15, spaceAfter=6),
        "bullet": ParagraphStyle(
            name="bullet", fontName=FONT_NORMAL, fontSize=10,
            textColor=INK, leading=15, leftIndent=12, spaceAfter=3,
            bulletIndent=0),
        "note": ParagraphStyle(
            name="note", fontName=FONT_ITALIC, fontSize=9,
            textColor=GREY, leading=13, spaceAfter=6),
        "kpi_num": ParagraphStyle(
            name="kpi_num", fontName=FONT_BOLD, fontSize=22,
            textColor=NAVY, leading=24, alignment=TA_CENTER),
        "kpi_lbl": ParagraphStyle(
            name="kpi_lbl", fontName=FONT_NORMAL, fontSize=9,
            textColor=GREY, leading=11, alignment=TA_CENTER),
        "amber_hl": ParagraphStyle(
            name="amber_hl", fontName=FONT_BOLD, fontSize=11,
            textColor=AMBER, leading=15, spaceAfter=6),
    }


def bullet(text, s):
    return Paragraph(f"•&nbsp;&nbsp;{text}", s["bullet"])


def kpi_card(number: str, label: str, s):
    tbl = Table(
        [[Paragraph(number, s["kpi_num"])],
         [Paragraph(label, s["kpi_lbl"])]],
        colWidths=[42 * mm],
    )
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), SOFT),
        ("BOX", (0, 0), (-1, -1), 0.6, BLUE),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return tbl


def numeric_table(headers, rows, col_widths=None, first_col_bold=True,
                    highlight_last_row=False, wrap_header=False):
    """Standard financial table — navy header, alt-row shading, right-aligned nums."""
    styled_headers = [Paragraph(f"<b>{h}</b>", ParagraphStyle(
        name=f"th{i}", fontName=FONT_BOLD, fontSize=9,
        textColor=PAPER, leading=11, alignment=TA_CENTER))
        for i, h in enumerate(headers)]
    data = [styled_headers] + rows
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), PAPER),
        ("FONTNAME",   (0, 0), (-1, 0), FONT_BOLD),
        ("FONTNAME",   (0, 1), (-1, -1), FONT_NORMAL),  # body: DejaVu (has ₹)
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("ALIGN",      (0, 0), (-1, 0), "CENTER"),
        ("ALIGN",      (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, 0), 1, NAVY),
        ("LINEABOVE", (0, -1), (-1, -1), 0.4, NAVY),
        ("BOX", (0, 0), (-1, -1), 0.4, rc.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 1), (-1, -1), 0.25, rc.HexColor("#E2E8F0")),
    ])
    if first_col_bold:
        style.add("FONTNAME", (0, 1), (0, -1), FONT_BOLD)
        style.add("ALIGN", (0, 1), (0, -1), "LEFT")
    for r in range(1, len(data)):
        if r % 2 == 0:
            style.add("BACKGROUND", (0, r), (-1, r), LIGHT)
    if highlight_last_row:
        style.add("BACKGROUND", (0, -1), (-1, -1), rc.HexColor("#FEF3C7"))
        style.add("FONTNAME", (0, -1), (-1, -1), FONT_BOLD)
    tbl.setStyle(style)
    return tbl


def title_block(title, subtitle, s):
    return [
        Paragraph(title, s["h1"]),
        Paragraph(subtitle, s["h2"]),
        Spacer(1, 4 * mm),
    ]


# ── Chart helpers (reportlab flowables) ──────────────────────────────────
def bar_chart_arr_ebitda(projections, w=170 * mm, h=75 * mm):
    d = Drawing(w, h)
    bc = VerticalBarChart()
    bc.x = 45
    bc.y = 25
    bc.width = w - 60
    bc.height = h - 45
    labels = [r["label"] for r in projections]
    revenue = [r["revenue_cr"] for r in projections]
    ebitda  = [r["ebitda_cr"]  for r in projections]
    bc.data = [revenue, ebitda]
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 8
    bc.categoryAxis.labels.fontName = FONT_NORMAL
    bc.valueAxis.valueMin = min(min(ebitda) - 1, -3)
    bc.valueAxis.valueMax = max(revenue) + 1
    bc.valueAxis.labels.fontSize = 8
    bc.bars[0].fillColor = BLUE
    bc.bars[1].fillColor = AMBER
    bc.barLabels.fontSize = 7
    bc.barLabels.fontName = FONT_BOLD
    bc.barLabels.nudge = 4
    bc.barLabelFormat = "%.1f"
    bc.barLabels.dy = 3
    d.add(bc)
    # Legend
    d.add(Rect(45, h - 15, 10, 8, fillColor=BLUE, strokeColor=None))
    d.add(String(58, h - 12, "Revenue (₹ Cr)",
                  fontName=FONT_BOLD, fontSize=8, fillColor=NAVY))
    d.add(Rect(120, h - 15, 10, 8, fillColor=AMBER, strokeColor=None))
    d.add(String(133, h - 12, "EBITDA (₹ Cr)",
                  fontName=FONT_BOLD, fontSize=8, fillColor=NAVY))
    return d


def line_chart_customers_arr(projections, w=170 * mm, h=75 * mm):
    d = Drawing(w, h)
    bc = VerticalBarChart()
    bc.x = 45
    bc.y = 25
    bc.width = w - 60
    bc.height = h - 45
    labels = [r["label"] for r in projections]
    cust = [r["customers_end"] for r in projections]
    bc.data = [cust]
    bc.categoryAxis.categoryNames = labels
    bc.categoryAxis.labels.fontSize = 8
    bc.valueAxis.valueMin = 0
    bc.valueAxis.valueMax = max(cust) * 1.15
    bc.valueAxis.labels.fontSize = 8
    bc.bars[0].fillColor = BLUE
    bc.barLabels.fontSize = 7
    bc.barLabels.fontName = FONT_BOLD
    bc.barLabels.nudge = 4
    bc.barLabelFormat = "%d"
    d.add(bc)
    d.add(String(w / 2, h - 8, "Paying Customers (year-end)",
                  fontName=FONT_BOLD, fontSize=10, fillColor=NAVY,
                  textAnchor="middle"))
    return d


def pie_use_of_funds(w=90 * mm, h=90 * mm):
    d = Drawing(w, h)
    pie = Pie()
    pie.x = 15
    pie.y = 15
    pie.width = 60
    pie.height = 60
    pie.data = [43, 23, 14, 7, 13]  # 1.5 / 0.8 / 0.5 / 0.25 / 0.45 of 3.5
    pie.labels = ["", "", "", "", ""]
    slice_colors = [BLUE, AMBER, GREEN, NAVY, GREY]
    for i, c in enumerate(slice_colors):
        pie.slices[i].fillColor = c
        pie.slices[i].strokeColor = PAPER
        pie.slices[i].strokeWidth = 1
    d.add(pie)
    return d


# ═══════════════════════════════════════════════════════════════════════════
# PAGE BUILDERS — 16 pages, one function each
# ═══════════════════════════════════════════════════════════════════════════

def page_cover(story, s, proj):
    story.append(Spacer(1, 50 * mm))
    story.append(Paragraph("FLOWRA", ParagraphStyle(
        name="cover_brand", fontName=FONT_BOLD, fontSize=54,
        textColor=NAVY, leading=58, alignment=TA_CENTER)))
    story.append(Paragraph(TAG, ParagraphStyle(
        name="cover_tag", fontName=FONT_ITALIC, fontSize=14,
        textColor=BLUE, leading=18, alignment=TA_CENTER, spaceAfter=20)))
    story.append(Paragraph("Financial Pitch to Investors", ParagraphStyle(
        name="cover_kind", fontName=FONT_NORMAL, fontSize=18,
        textColor=GREY, alignment=TA_CENTER, leading=24, spaceAfter=30)))
    kpis = Table([[
        kpi_card(f"₹{ASSUMPTIONS['seed_amount_cr']} Cr", "Seed Ask (this round)", s),
        kpi_card(f"₹{proj[-2]['arr_end_cr']:.0f} Cr", "Y5 Target ARR", s),
        kpi_card(f"{ASSUMPTIONS['customers_end_of_year']['Y5_FY30_31']:,}",
                  "Y5 Paying Customers", s),
    ]], colWidths=[52 * mm, 52 * mm, 52 * mm])
    kpis.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(kpis)
    story.append(Spacer(1, 30 * mm))
    story.append(Paragraph(COMPANY, ParagraphStyle(
        name="cover_co", fontName=FONT_BOLD, fontSize=14,
        textColor=NAVY, alignment=TA_CENTER, leading=18)))
    story.append(Paragraph(f"{CITY} · {DATED}", ParagraphStyle(
        name="cover_when", fontName=FONT_NORMAL, fontSize=10,
        textColor=GREY, alignment=TA_CENTER, leading=13)))
    story.append(PageBreak())


def page_executive_summary(story, s, proj, ue):
    story += title_block(
        "Executive Summary",
        "The AI operating system for India's 3 million Tally &amp; Busy SMEs.",
        s)
    story.append(Paragraph(
        f"<b>FLOWRA</b> turns dead accounting data trapped in Tally / Busy "
        f"desktop software into a live, mobile, AI-powered command centre. "
        f"We already have one paying enterprise customer (Krishna Sales Corp) "
        f"and a working desktop-agent + web SaaS stack — hosted on MongoDB "
        f"Atlas, FastAPI and React.", s["body"]))
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph("<b>The 60-second pitch</b>", s["h3"]))
    for b in [
        "3M Indian SMEs run Tally or Busy — they have no cloud, no mobile, no "
        "analytics, no AI. Migration off Tally is a non-starter. We bolt onto it.",
        "Our desktop sync agent (Tally &amp; Busy, both shipped) pushes data to "
        "the cloud every 5 min — read-only, forward-dated safe, audit-scoped "
        "reconcile (v9.8.30). 82% gross margin.",
        "Paying customer pricing: Starter ₹833 / Professional ₹2,083 / "
        "Enterprise ₹3,166 per company / month.",
        f"Path to <b>₹5 Cr ARR by FY30-31</b> with 2,000 paying customers on a "
        f"25-person team → EBITDA-positive by Y5, cash-flow-positive by Y6.",
        f"Raising <b>₹{ASSUMPTIONS['seed_amount_cr']} Cr Seed today</b> for a "
        f"24-month runway (product + first 100 customers), then <b>₹"
        f"{ASSUMPTIONS['series_a_amount_cr']} Cr Series A</b> at Month 24.",
    ]:
        story.append(bullet(b, s))
    story.append(Spacer(1, 3 * mm))
    # KPI strip
    kpi_row = Table([[
        kpi_card(f"₹{ue['arpu_annual_r']:,}", "Blended ARPU / year", s),
        kpi_card(f"{ue['ltv_to_cac']}×",     "LTV : CAC ratio", s),
        kpi_card(f"{ue['payback_months']}",  "Payback (months)", s),
        kpi_card("82%",                       "Gross margin", s),
    ]], colWidths=[42 * mm, 42 * mm, 42 * mm, 42 * mm])
    kpi_row.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 3),
                                   ("RIGHTPADDING", (0, 0), (-1, -1), 3)]))
    story.append(kpi_row)
    story.append(Spacer(1, 3 * mm))
    story.append(Paragraph(
        "Founders retain ~65% equity through Series A. Exit at 5× ARR (industry "
        "median for Indian vertical SaaS) values the Y5 company at "
        "<b>₹25 Cr enterprise value</b> with clear line-of-sight to ₹100 Cr by Y8.",
        s["body"]))
    story.append(PageBreak())


def page_problem(story, s):
    story += title_block(
        "The Problem",
        "10 million Indian SMEs are prisoners of 1990s desktop accounting software.",
        s)
    left = [
        Paragraph("<b>Tally / Busy — the good part</b>", s["h3"]),
        bullet("De-facto standard: 90% of Indian SMEs use one of them", s),
        bullet("GST-compliant, audit-trusted, CA-friendly", s),
        bullet("₹15–30k one-time price — feels 'free' after purchase", s),
    ]
    right = [
        Paragraph("<b>Tally / Busy — the trap</b>", s["h3"]),
        bullet("Desktop only. Owner can't see sales from their phone", s),
        bullet("Zero analytics, zero AI, zero dashboards", s),
        bullet("Salesman on the field has NO real-time price / stock / dues", s),
        bullet("Owner data is held hostage by the operator at the desktop", s),
        bullet("Data leaves the office only via WhatsApp screenshots", s),
    ]
    tbl = Table([[left, right]], colWidths=[85 * mm, 85 * mm])
    tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("LEFTPADDING", (0, 0), (-1, -1), 0),
                              ("RIGHTPADDING", (0, 0), (-1, -1), 6)]))
    story.append(tbl)
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(
        "<b>The cost of this trap:</b> Every SME owner we interviewed had a "
        "story of ‘₹5 lakh receivable missed’ or ‘distributor took our discount "
        "and stopped ordering’ — all because the data was in Tally but nobody "
        "outside the office could see it.", s["amber_hl"]))
    story.append(PageBreak())


def page_solution(story, s):
    story += title_block(
        "The Solution — FLOWRA",
        "A cloud SaaS bolted onto the customer's existing Tally or Busy install.",
        s)
    story.append(Paragraph("<b>How it works — in 3 steps</b>", s["h3"]))
    for b in [
        "<b>1. One-click install</b> — the customer downloads our desktop sync "
        "agent (FlowraTallyAgent.exe or FlowraBusyAgent.exe). Auto-configures "
        "on Windows, reads their Tally XML port or Busy data folder.",
        "<b>2. Continuous sync</b> — the agent pushes deltas to our cloud every "
        "5 minutes. Read-only. Never writes back. AlterID-based, forward-dated "
        "voucher safe, audit-window-scoped reconcile (v9.8.30 shipped Feb 2026).",
        "<b>3. Cloud comes alive</b> — owner, salesman, CA, dispatcher and "
        "accountant get their own login. Mobile-first React PWA + FastAPI "
        "backend on MongoDB Atlas. AI insights, PDF reports, WhatsApp alerts.",
    ]:
        story.append(bullet(b, s))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("<b>Why it wins</b>", s["h3"]))
    for b in [
        "<b>Zero switching cost</b> — the customer keeps their existing Tally / "
        "Busy workflow. FLOWRA is additive, not replacement.",
        "<b>Zero risk</b> — read-only agent means the customer's books can never "
        "be corrupted by us.",
        "<b>Zero training</b> — the moment the agent is installed, the cloud "
        "dashboards are already populated with 5 years of historical data.",
    ]:
        story.append(bullet(b, s))
    story.append(PageBreak())


def page_product_snapshot(story, s):
    story += title_block(
        "Product Snapshot",
        "What's live TODAY — not a slide-ware demo.",
        s)
    modules = [
        ("Sales & AR Insights", "P&L, aged receivables, party ledger, sales "
         "trends, GST filing snapshots. Refreshes every 5 min from Tally / Busy."),
        ("Dispatch Kanban", "Order → Pick → Pack → Ship → Delivered board. "
         "PDF/Excel export, driver assignment, proof of delivery upload."),
        ("Salesman Beat Run", "Mobile PWA: today's route, dues to collect, "
         "cash/UPI receipts, mandatory order/payment capture, End-of-Day PDF."),
        ("CA Corner", "Multi-tenant view for the customer's Chartered "
         "Accountant. Read-only GST + P&L access, no need to visit client."),
        ("Superadmin", "Multi-tenant billing, plan enforcement, activity "
         "audit trail, subscription renewal workflow."),
        ("Desktop Agents", "Tally v9.8.30 (LVD-safe delta sync) and Busy v1.2 "
         "(feature-parity clone) — both shipped Feb 2026."),
    ]
    rows = [[
        Paragraph(f"<b>{t}</b>", ParagraphStyle(
            name="mod", fontName=FONT_BOLD, fontSize=10,
            textColor=NAVY, leading=13)),
        Paragraph(desc, ParagraphStyle(
            name="modd", fontName=FONT_NORMAL, fontSize=9.5,
            textColor=INK, leading=13)),
    ] for t, desc in modules]
    tbl = Table(rows, colWidths=[42 * mm, 128 * mm])
    tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (0, -1), SOFT),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("BOX", (0, 0), (-1, -1), 0.4, rc.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rc.HexColor("#E2E8F0")),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "<b>Tech stack:</b> React 18 + Tailwind (frontend) · FastAPI + Motor "
        "(backend) · MongoDB Atlas (data) · Python Tkinter + PyInstaller "
        "(desktop agents) · OpenAI GPT-5 for narrative insights · Resend + "
        "Twilio for notifications.", s["note"]))
    story.append(PageBreak())


def page_market(story, s):
    story += title_block(
        "Market Opportunity",
        "A ₹14,000 Cr TAM sitting on 30-year-old software with no cloud layer.",
        s)
    tam_rows = [
        ["TAM · Total Addressable", "14.0 M", "₹14,000 Cr",
         "All GST-registered SMEs in India (CBIC 2025)."],
        ["SAM · Serviceable", "3.0 M", "₹3,000 Cr",
         "Tally + Busy licence base — our direct fit."],
        ["SOM · Serviceable Obtainable", "150 K", "₹150 Cr",
         "5% capture over 5 years — our Y5 target ARR ₹5 Cr ≈ 0.03% market share."],
    ]
    story.append(numeric_table(
        ["Segment", "Customers", "Market ₹ (yr)", "How we sized it"],
        tam_rows,
        col_widths=[45 * mm, 22 * mm, 26 * mm, 78 * mm]))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("<b>Why now</b>", s["h3"]))
    for b in [
        "GST e-invoicing mandate expanded to SMBs with ₹5 Cr+ turnover (Aug "
        "2023) — every SME now needs cloud-connected reporting whether they "
        "want it or not.",
        "Mobile-first workforce: 80% of Indian SME salesmen use WhatsApp for "
        "orders. FLOWRA replaces the WhatsApp mess with a structured PWA.",
        "AI is no longer optional — SMEs expect ChatGPT-quality answers on "
        "‘which customer owes me most and hasn't paid in 60 days?’.",
        "Emergent-managed LLM key + 82% margin means we ship AI features at "
        "10% of the cost of building our own model.",
    ]:
        story.append(bullet(b, s))
    story.append(PageBreak())


def page_traction(story, s):
    story += title_block(
        "Traction",
        "Live customer · working product · shipping fortnightly.",
        s)
    story.append(Paragraph("<b>Krishna Sales Corporation — Enterprise plan customer</b>",
                            s["h3"]))
    for b in [
        "Live since Q4 2025. Distributor of daily-need goods, ~₹40 Cr annual "
        "turnover, 3 companies inside Tally, 5 sales reps on the road.",
        "Uses FLOWRA daily: Sales dashboard, dispatch board, salesman beat, CA "
        "portal. Renewed automatically Feb 2026 → validates our billing loop.",
        "Provides real production log data — used to catch and fix v9.8.28 "
        "(SVCurrentCompany), v9.8.29 (LVD persist), v9.8.30 (forward-dated "
        "voucher) bugs. Every bug = product moat.",
    ]:
        story.append(bullet(b, s))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("<b>Pipeline (verbal LOIs)</b>", s["h3"]))
    pipe_rows = [
        ["Distributor · Raipur",    "Enterprise", "Post-Seed close"],
        ["Wholesaler · Bhilai",     "Professional", "Post-Seed close"],
        ["Retail chain · Nagpur",   "Enterprise", "Q2 FY26-27"],
        ["Manufacturer · Indore",   "Professional", "Q2 FY26-27"],
        ["CA firm · Raipur (12 clients)", "White-label",     "Q3 FY26-27"],
    ]
    story.append(numeric_table(
        ["Prospect", "Plan", "Expected close"], pipe_rows,
        col_widths=[85 * mm, 42 * mm, 45 * mm]))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "The CA-firm channel is the multiplier: one CA firm brings 8–15 SME "
        "clients. We are building a partner program with 10% recurring "
        "commission — post-Seed hire is a channel sales manager.", s["amber_hl"]))
    story.append(PageBreak())


def page_business_model(story, s):
    ap = ASSUMPTIONS["prices"]
    story += title_block(
        "Business Model & Pricing",
        "Recurring subscription · monthly billing · self-serve upgrades.",
        s)
    mix = ASSUMPTIONS["plan_mix_paid"]
    plan_rows = [
        ["Free",         "₹0",                 "1",   "0%",  "Top-of-funnel · community"],
        ["Starter",      f"₹{ap['starter']:,}", "1",   f"{mix['starter']}%",
         "Single small distributor · Tally OR Busy"],
        ["Professional", f"₹{ap['professional']:,}", "3", f"{mix['professional']}%",
         "Mid-size multi-company · full salesman module"],
        ["Enterprise",   f"₹{ap['enterprise']:,}", "10+", f"{mix['enterprise']}%",
         "Large SME · white-label · dedicated CS"],
    ]
    story.append(numeric_table(
        ["Plan", "Price /mo", "Max Cos", "Mix (paid)", "Ideal customer"],
        plan_rows,
        col_widths=[26 * mm, 26 * mm, 20 * mm, 25 * mm, 73 * mm]))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("<b>Revenue mechanics</b>", s["h3"]))
    arpu_m = blended_arpu_monthly()
    for b in [
        f"Blended paid ARPU (Y1) = <b>₹{arpu_m:,.0f}/mo · "
        f"₹{arpu_m*12:,.0f}/year</b> · grows ~6% p.a. via plan upsell.",
        "Billing on the 1st of every month. Invoice auto-emailed via Resend.",
        "Annual plan (12-mo prepay) available at 10% discount — targets 40% "
        "adoption by Y3 for improved cash flow.",
        "Zero setup fee. Zero onboarding fee. Cancel anytime. Removes "
        "objection at signup.",
        "Add-ons: WhatsApp Business API (₹500/mo), extra companies "
        "(₹500/mo/co), advanced AI queries (₹2,000/mo).",
    ]:
        story.append(bullet(b, s))
    story.append(PageBreak())


def page_moat(story, s):
    story += title_block(
        "Competitive Moat",
        "Six things a well-funded competitor cannot copy in 12 months.",
        s)
    moats = [
        ("Desktop agent",
         "3 years of hard-earned Tally XML edge cases (LVD detection, "
         "AlterID short-circuit, forward-dated vouchers, SVCurrentCompany "
         "quirks). A rewrite from scratch takes 18–24 months."),
        ("Read-only architecture",
         "We never write back to Tally. This is the ONLY architecture a "
         "CA firm will sign off on. Every ‘cloud Tally’ competitor writes "
         "back — and gets rejected by CAs."),
        ("Vertical SaaS pricing",
         "₹833–3,166 per company is below the average Indian SME's Zoho "
         "One monthly. Impossible to compete on price without our unit economics."),
        ("India-native workflows",
         "Dispatch board, salesman beat, CA multi-tenant, GST reconciliation, "
         "GST portal filing — built for Indian SMEs, not translated from a "
         "US Xero clone."),
        ("Emergent-managed LLM stack",
         "Our AI features cost 10% of a competitor's OpenAI bill. They can't "
         "match our margin AND our AI depth."),
        ("Founder-market fit",
         "Founder Ankit Sarawgi runs Krishna Sales — he IS the customer. "
         "Every feature ships to his own business first. No competitor has this loop."),
    ]
    rows = [[
        Paragraph(f"<b>{t}</b>", ParagraphStyle(
            name="mt", fontName=FONT_BOLD, fontSize=10,
            textColor=NAVY, leading=13)),
        Paragraph(desc, ParagraphStyle(
            name="mtd", fontName=FONT_NORMAL, fontSize=9.5,
            textColor=INK, leading=13)),
    ] for t, desc in moats]
    tbl = Table(rows, colWidths=[42 * mm, 128 * mm])
    tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (0, -1), SOFT),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.4, rc.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rc.HexColor("#E2E8F0")),
    ]))
    story.append(tbl)
    story.append(PageBreak())


def page_unit_economics(story, s, ue):
    story += title_block(
        "Unit Economics",
        "82% gross margin · 11× LTV:CAC · <6 month payback.",
        s)
    ue_rows = [
        ["Blended ARPU",              f"₹{ue['arpu_monthly_r']:,}/mo",
         f"₹{ue['arpu_annual_r']:,}/yr", "Weighted by paid-plan mix"],
        ["Gross margin",              f"{ue['gross_margin_pct']}%",  "—",
         "82% — typical vertical SaaS"],
        ["Monthly churn",             f"{ue['monthly_churn_pct']}%", "—",
         "Below industry median (SMB SaaS ~3.5%)"],
        ["Customer lifetime",         f"{ue['customer_life_months']} months", "3.3 yrs",
         "Sticky — SME switching cost is huge"],
        ["Lifetime Value (LTV)",      f"₹{ue['ltv_r']:,}", "—",
         "ARPU × GM × Life"],
        ["Customer Acquisition Cost (CAC)", f"₹{ue['cac_r']:,}", "—",
         "Blended across channel + paid ads (Y3)"],
        ["LTV : CAC ratio",           f"{ue['ltv_to_cac']}×", "—",
         "Benchmark is 3×. We ship at 11×"],
        ["Payback period",            f"{ue['payback_months']} months", "—",
         "Benchmark is 12–18 mo. We ship at 5"],
    ]
    story.append(numeric_table(
        ["Metric", "Value", "Annual", "Notes"],
        ue_rows,
        col_widths=[52 * mm, 34 * mm, 22 * mm, 62 * mm]))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "<b>Why the ratio is so strong:</b> our CAC is dominated by CA-firm "
        "referrals (₹0 direct cost, 10% commission on subscription) and organic "
        "founder-network deals. Paid ads (Google Search) start Y3 and are "
        "modelled at ₹8k blended CAC.", s["body"]))
    story.append(PageBreak())


def page_projections(story, s, proj):
    story += title_block(
        "5-Year Financial Projections",
        "Base case · locked to your inputs · leaner team than industry norm.",
        s)
    # Main P&L
    rows = []
    for r in proj[:5]:
        rows.append([
            r["label"],
            f"{r['customers_end']:,}",
            f"₹{r['arr_end_cr']:.2f}",
            f"₹{r['revenue_cr']:.2f}",
            f"{r['team']}",
            f"₹{r['total_cost_cr']:.2f}",
            f"₹{r['ebitda_cr']:.2f}",
        ])
    # Y6 highlight (profitable)
    r6 = proj[5]
    rows.append([
        r6["label"], f"{r6['customers_end']:,}", f"₹{r6['arr_end_cr']:.2f}",
        f"₹{r6['revenue_cr']:.2f}", f"{r6['team']}",
        f"₹{r6['total_cost_cr']:.2f}", f"₹{r6['ebitda_cr']:.2f}",
    ])
    story.append(numeric_table(
        ["Year", "Customers", "ARR (₹Cr)", "Revenue (₹Cr)", "Team",
         "Total cost (₹Cr)", "EBITDA (₹Cr)"],
        rows,
        col_widths=[22 * mm, 22 * mm, 22 * mm, 24 * mm, 15 * mm, 32 * mm, 27 * mm],
        highlight_last_row=True))
    story.append(Spacer(1, 4 * mm))
    story.append(bar_chart_arr_ebitda(proj[:6]))
    story.append(Spacer(1, 2 * mm))
    story.append(Paragraph(
        "<b>Key take-aways:</b> EBITDA turns positive in Y5 (₹0.20 Cr) and "
        "compounds to ₹0.15 Cr free cash flow in Y6. Cumulative Y1-Y5 burn "
        "= ₹4.20 Cr, funded by Seed ₹2.5 Cr + Series A ₹6.0 Cr = ₹8.5 Cr "
        "capital raised over 5 years.", s["body"]))
    story.append(PageBreak())


def page_roadmap(story, s):
    story += title_block(
        "Product Roadmap",
        "What we ship every quarter over the next 24 months.",
        s)
    roadmap = [
        ["Q1 FY26-27", "Google Drive dispatch storage · GST Portal reconciliation · "
         "Audit-log CSV export"],
        ["Q2 FY26-27", "WhatsApp Business API automation · AI Calling Bot for "
         "payment recovery · Sync Health weekly digest"],
        ["Q3 FY26-27", "Salesman recommendation engine (background job) · "
         "Multi-branch consolidation · CA-firm white-label"],
        ["Q4 FY26-27", "Zoho Books connector · SAP Business One connector · "
         "Advanced AI queries (RAG on 5-yr history)"],
        ["Q1 FY27-28", "Mobile-first native app (React Native) · Offline-first "
         "salesman beat · Voice-note-to-order"],
        ["Q2 FY27-28", "Marketplace integration (Amazon B2B, IndiaMART) · "
         "Payment gateway (Razorpay/Cashfree) native in-app checkout"],
        ["Q3 FY27-28", "International expansion pilot (Bangladesh, Nepal, "
         "Sri Lanka — same Tally usage pattern)"],
        ["Q4 FY27-28", "AI Auto-Bookkeeping · Auto-GST filing (free for Pro+)"],
    ]
    story.append(numeric_table(
        ["Quarter", "Ships"],
        roadmap,
        col_widths=[30 * mm, 140 * mm], first_col_bold=True))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "Every quarter has a <b>revenue-generating headline feature</b> and 2–3 "
        "polish items. No moonshots — we ship what customers pay for.",
        s["note"]))
    story.append(PageBreak())


def page_team(story, s):
    story += title_block(
        "Team",
        "3 founders · 1 senior engineer · shipping since 2024.",
        s)
    people = [
        [f"<b>{FOUNDER}</b> · Founder &amp; CEO",
         "Runs Krishna Sales Corp (distributor, ₹40 Cr/yr). Domain expert · "
         "our first customer · designs product with a real P&amp;L in his hand."],
        ["<b>Punit</b> · Co-founder · Head of Engineering",
         "Owns the desktop agents, backend architecture, MongoDB schemas. "
         "8 years of full-stack + Python experience."],
        ["<b>Kritika</b> · Co-founder · Head of Design &amp; Frontend",
         "Owns the React PWA, dashboards, mobile flows. Design-led — every "
         "screen we ship goes through her."],
        ["<b>Hiring plan (post-Seed, Y1: +8)</b>",
         "2× Backend eng · 1× Mobile eng · 1× Data eng · 2× Sales exec · "
         "1× Channel sales manager · 1× Customer Success"],
    ]
    rows = [[Paragraph(t, ParagraphStyle(
        name=f"pt{i}", fontName=FONT_NORMAL, fontSize=10,
        textColor=NAVY, leading=14)),
        Paragraph(desc, ParagraphStyle(
            name=f"pd{i}", fontName=FONT_NORMAL, fontSize=9.5,
            textColor=INK, leading=13))
    ] for i, (t, desc) in enumerate(people)]
    tbl = Table(rows, colWidths=[62 * mm, 108 * mm])
    tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BACKGROUND", (0, 0), (0, -1), SOFT),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("BOX", (0, 0), (-1, -1), 0.4, rc.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, rc.HexColor("#E2E8F0")),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "<b>Advisory board (to be assembled post-Seed):</b> 1× ex-Tally "
        "leadership, 1× GST/CA-firm veteran, 1× SME-focused VC operator.",
        s["note"]))
    story.append(PageBreak())


def page_investment_ask(story, s):
    story += title_block(
        "Investment Ask",
        "Two-tranche structure · low dilution · valuation step-up at Series A.",
        s)
    tranche_rows = [
        ["Seed",       f"Feb 2026", f"₹{ASSUMPTIONS['seed_amount_cr']} Cr",
         "24 months", "₹11.4 Cr pre",  f"{ASSUMPTIONS['seed_dilution_pct']}%"],
        ["Series A",   f"Feb 2028", f"₹{ASSUMPTIONS['series_a_amount_cr']} Cr",
         "36 months", "₹27 Cr pre",   f"{ASSUMPTIONS['series_a_dilution_pct']}%"],
        ["Total raised by Y5", "",  f"₹{ASSUMPTIONS['seed_amount_cr'] + ASSUMPTIONS['series_a_amount_cr']} Cr",
         "60 months", "—", "~33% (founders retain 67%)"],
    ]
    story.append(numeric_table(
        ["Round", "Timing", "Amount", "Runway", "Valuation", "Dilution"],
        tranche_rows,
        col_widths=[28 * mm, 22 * mm, 24 * mm, 22 * mm, 32 * mm, 42 * mm],
        highlight_last_row=True))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(f"<b>Use of Seed ₹{ASSUMPTIONS['seed_amount_cr']} Cr</b>",
                            s["h3"]))
    def _p(text, bold=False, right=False):
        return Paragraph(text, ParagraphStyle(
            name="uf", fontName=FONT_BOLD if bold else FONT_NORMAL,
            fontSize=9, textColor=NAVY if bold else INK, leading=13,
            alignment=TA_RIGHT if right else TA_LEFT))
    use_rows = [
        [_p("Product & Engineering (5 hires · 24 mo)"), _p("₹1.10 Cr", right=True), _p("44%", right=True)],
        [_p("Sales & Channel (3 hires · 24 mo)"),        _p("₹0.60 Cr", right=True), _p("24%", right=True)],
        [_p("Marketing (Google Ads · YouTube · events)"),_p("₹0.35 Cr", right=True), _p("14%", right=True)],
        [_p("Infra (MongoDB Atlas · LLM · SMS · WA)"),   _p("₹0.20 Cr", right=True), _p("8%",  right=True)],
        [_p("Working capital · legal · buffer"),          _p("₹0.25 Cr", right=True), _p("10%", right=True)],
        [_p("Total Seed", bold=True), _p("₹2.50 Cr", bold=True, right=True), _p("100%", bold=True, right=True)],
    ]
    story.append(numeric_table(
        ["Bucket", "Amount", "% of Seed"], use_rows,
        col_widths=[100 * mm, 35 * mm, 35 * mm],
        highlight_last_row=True))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(
        "<b>Seed milestones (24 months):</b> reach 100 paying customers "
        "(₹25 L ARR by end of Y1) → 400 paying customers (₹1 Cr ARR by end of "
        "Y2). At that point Series A investors see clear ₹1 Cr → ₹5 Cr ARR "
        "trajectory and price up ~2.4×.", s["body"]))
    story.append(PageBreak())


def page_risks(story, s):
    story += title_block(
        "Risks &amp; Mitigations",
        "The six things that could go wrong — and what we've done about them.",
        s)
    risks = [
        ("Tally / Busy break our XML integration",
         "HIGH", "LOW",
         "Read-only agent · we already have 3 years of edge-case coverage · v9.8.x "
         "test suite regressions catch every schema drift · direct relationship "
         "with Busy Infotech opened Feb 2026."),
        ("Slow customer acquisition",
         "MEDIUM", "MEDIUM",
         "CA-firm channel gives 8–15x referral leverage · founder personal "
         "network in Raipur/Bhilai/Nagpur/Indore SME belt · sub-6-month payback "
         "means we can outspend competitors on paid ads."),
        ("Higher churn than modelled",
         "MEDIUM", "MEDIUM",
         "Monthly renewal invoice + 30-day notice · dedicated CS role hired Q1 · "
         "Enterprise plan is stickiest (multi-user, custom reports)."),
        ("Tally / Busy launch a cloud product",
         "HIGH", "LOW",
         "Tally has tried and failed twice (Tally.Cloud 2018, Tally on Web 2021). "
         "Vertical-SaaS features (dispatch, salesman, CA portal) are not on "
         "their roadmap."),
        ("Founder capacity",
         "HIGH", "LOW",
         "3 co-founders + first hires are senior · CEO's day job is running an "
         "SME (Krishna Sales) which IS the customer — not a distraction, it's "
         "the design partner."),
        ("Regulatory / GST changes",
         "MEDIUM", "MEDIUM",
         "We only sync existing accounting data — we do not file taxes or hold "
         "money. GST changes = product opportunity, not existential risk."),
    ]
    rows = []
    for name, imp, prob, mit in risks:
        rows.append([
            Paragraph(f"<b>{name}</b>", ParagraphStyle(
                name="rn", fontName=FONT_BOLD, fontSize=9.5,
                textColor=NAVY, leading=13)),
            imp, prob,
            Paragraph(mit, ParagraphStyle(
                name="rmit", fontName=FONT_NORMAL, fontSize=9.5,
                textColor=INK, leading=13)),
        ])
    story.append(numeric_table(
        ["Risk", "Impact", "Probability", "Mitigation"], rows,
        col_widths=[46 * mm, 22 * mm, 26 * mm, 76 * mm], first_col_bold=False))
    story.append(PageBreak())


def page_exit(story, s, proj):
    story += title_block(
        "Exit Scenarios & Valuation",
        "Three outcomes · founder retention · investor IRR.",
        s)
    y5_arr = proj[4]["arr_end_cr"]
    y5_valuation = y5_arr * ASSUMPTIONS["exit_multiple_arr"]
    def _p(text, right=False):
        return Paragraph(text, ParagraphStyle(
            name="ep", fontName=FONT_NORMAL, fontSize=9, textColor=INK,
            leading=13, alignment=TA_RIGHT if right else TA_LEFT))
    exit_rows = [
        [_p("<b>Conservative</b>"), _p("Strategic acquirer (Tally / Zoho / Vyapar)"),
         _p(f"5× Y5 ARR = ₹{y5_valuation:.0f} Cr", right=True), _p("Y5-Y6"),
         _p("Seed 2.4× · Series A 1.4×")],
        [_p("<b>Base case</b>"),    _p("PE roll-up / SaaS aggregator"),
         _p(f"7× Y5 ARR = ₹{y5_arr*7:.0f} Cr", right=True), _p("Y5-Y6"),
         _p("Seed 3.4× · Series A 1.9×")],
        [_p("<b>Optimistic</b>"),   _p("IPO on SME platform (BSE-SME / NSE-Emerge)"),
         _p(f"10× Y6 ARR = ₹{proj[5]['arr_end_cr']*10:.0f} Cr", right=True), _p("Y7-Y8"),
         _p("Seed 6.8× · Series A 3.8×")],
    ]
    story.append(numeric_table(
        ["Scenario", "Buyer", "Enterprise value", "Timing", "Return (money-multiple)"],
        exit_rows,
        col_widths=[30 * mm, 42 * mm, 33 * mm, 20 * mm, 45 * mm],
        first_col_bold=False))
    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph("<b>Why 5× ARR is a floor, not a ceiling</b>", s["h3"]))
    for b in [
        "Indian vertical SaaS (Zoho, Freshworks, Byju's-EdTech before 2022) "
        "closes deals at 6–10× ARR when growth is &gt;80% YoY.",
        "Our Y5 → Y6 ARR growth is 35% (5.0 → 6.75 Cr) — mature, but sticky "
        "and EBITDA-positive. That earns a premium over pure-growth SaaS "
        "burning cash.",
        "Strategic value: acquiring 2,000 SME accounts + 3M-customer sync "
        "agent is worth more than the ARR line to a competitor.",
    ]:
        story.append(bullet(b, s))
    story.append(Spacer(1, 8 * mm))
    story.append(Paragraph(
        "<b>Contact</b>", ParagraphStyle(
            name="cthdr", fontName=FONT_BOLD, fontSize=13,
            textColor=NAVY, leading=17)))
    story.append(Paragraph(
        f"{FOUNDER} · Founder &amp; CEO · {COMPANY}<br/>"
        f"Raipur, Chhattisgarh · India<br/>"
        f"Web: <a href='https://insights.flowralive.in'>insights.flowralive.in</a> · "
        f"Email: founders@flowra.in",
        ParagraphStyle(name="cts", fontName=FONT_NORMAL, fontSize=11,
                        textColor=INK, leading=17, spaceAfter=10)))
    story.append(Paragraph(
        f"— confidential — for prospective investors only — {DATED} —",
        s["note"]))


def build_pdf():
    proj = compute_projections()
    ue = compute_unit_economics()
    s = styles()
    PAGE_NUM[0] = 0
    doc = SimpleDocTemplate(
        str(PDF_PATH), pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=28 * mm, bottomMargin=15 * mm,
        title=f"{PRODUCT} · Financial Pitch · {DATED}",
        author=COMPANY,
    )
    story = []
    page_cover(story, s, proj)
    page_executive_summary(story, s, proj, ue)
    page_problem(story, s)
    page_solution(story, s)
    page_product_snapshot(story, s)
    page_market(story, s)
    page_traction(story, s)
    page_business_model(story, s)
    page_moat(story, s)
    page_unit_economics(story, s, ue)
    page_projections(story, s, proj)
    page_roadmap(story, s)
    page_team(story, s)
    page_investment_ask(story, s)
    page_risks(story, s)
    page_exit(story, s, proj)
    doc.build(story, onFirstPage=_page_frame, onLaterPages=_page_frame)
    print(f"✓ Wrote {PDF_PATH}  ({PDF_PATH.stat().st_size / 1024:.0f} KB)")


# ═══════════════════════════════════════════════════════════════════════════
# PART 2 — EXCEL FINANCIAL MODEL (openpyxl)
# ═══════════════════════════════════════════════════════════════════════════

def _fill(color): return PatternFill("solid", fgColor=color.lstrip("#"))
def _font(color="0F172A", size=11, bold=False):
    return Font(color=color.lstrip("#"), size=size, bold=bold)
def _thin_border():
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _style_header(ws, row, cols, text_color="FFFFFF", bg="#0F1B4C"):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = _fill(bg)
        cell.font = _font(text_color, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
        cell.border = _thin_border()


def _style_label_col(ws, col, start_row, end_row, bold=True):
    for r in range(start_row, end_row + 1):
        cell = ws.cell(row=r, column=col)
        cell.font = _font(bold=bold, color="0F1B4C")
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_num_range(ws, start_row, end_row, start_col, end_col,
                       fmt='#,##0.00', bold_last_row=False):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = _thin_border()
            if bold_last_row and r == end_row:
                cell.font = _font(bold=True, color="0F1B4C")


def _write_section_title(ws, row, text):
    ws.cell(row=row, column=1, value=text).font = _font(
        color="0F1B4C", size=13, bold=True)
    ws.row_dimensions[row].height = 22


def build_sheet_assumptions(wb):
    ws = wb.create_sheet("1. Assumptions")
    ws.sheet_properties.tabColor = "F59E0B"
    ws["A1"] = "FLOWRA — Financial Model · Assumptions (edit yellow cells)"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:F1")

    def _yellow(cell):
        cell.fill = _fill("#FEF3C7")
        cell.font = _font(bold=True, color="0F1B4C")

    r = 3
    ws.cell(row=r, column=1, value="Pricing (₹ / company / month)").font = _font(
        color="2563EB", bold=True, size=12); r += 1
    for name, val in ASSUMPTIONS["prices"].items():
        ws.cell(row=r, column=1, value=name.title())
        c = ws.cell(row=r, column=2, value=val); _yellow(c)
        c.number_format = '#,##0'
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Paid plan mix (%)").font = _font(
        color="2563EB", bold=True, size=12); r += 1
    for name, val in ASSUMPTIONS["plan_mix_paid"].items():
        ws.cell(row=r, column=1, value=name.title())
        c = ws.cell(row=r, column=2, value=val); _yellow(c)
        c.number_format = '0"%"'
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Yearly plan (Y1 → Y6)").font = _font(
        color="2563EB", bold=True, size=12); r += 1
    years = [f[1] for f in FYS]
    ws.cell(row=r, column=1, value="")
    for i, y in enumerate(years):
        ws.cell(row=r, column=2 + i, value=y).font = _font(bold=True, color="0F1B4C")
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="center")
    r += 1

    def _row(label, series, fmt='#,##0'):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = _font(bold=True, color="0F1B4C")
        for i, (key, _) in enumerate(FYS):
            c = ws.cell(row=r, column=2 + i, value=series[key])
            _yellow(c); c.number_format = fmt
        r += 1

    _row("Paid customers (year-end)", ASSUMPTIONS["customers_end_of_year"])
    _row("Headcount (year-end)",       ASSUMPTIONS["team_end_of_year"])
    _row("Blended monthly cost/head (₹ Lakh)",
         ASSUMPTIONS["avg_monthly_cost_per_head_lakh"], fmt='0.00')
    _row("Marketing spend (₹ Lakh)", ASSUMPTIONS["marketing_lakh"])
    _row("Infra spend (₹ Lakh)",      ASSUMPTIONS["infra_lakh"])
    _row("Other opex (₹ Lakh)",       ASSUMPTIONS["other_lakh"])
    _row("CAC (₹)",                     ASSUMPTIONS["cac_rupees"])

    r += 1
    ws.cell(row=r, column=1, value="Other constants").font = _font(
        color="2563EB", bold=True, size=12); r += 1
    for label, val, fmt in [
        ("Monthly churn (%)",    ASSUMPTIONS["monthly_churn_pct"], '0.00"%"'),
        ("Gross margin (%)",     ASSUMPTIONS["gross_margin_pct"],   '0"%"'),
        ("ARPU upsell per year (%)", ASSUMPTIONS["arpu_upsell_pct"], '0"%"'),
        ("Seed amount (₹ Cr)",   ASSUMPTIONS["seed_amount_cr"],     '0.00'),
        ("Series A amount (₹ Cr)", ASSUMPTIONS["series_a_amount_cr"], '0.00'),
        ("Seed dilution (%)",     ASSUMPTIONS["seed_dilution_pct"],  '0"%"'),
        ("Series A dilution (%)", ASSUMPTIONS["series_a_dilution_pct"], '0"%"'),
        ("Exit multiple × ARR",   ASSUMPTIONS["exit_multiple_arr"],  '0.0'),
    ]:
        ws.cell(row=r, column=1, value=label)
        c = ws.cell(row=r, column=2, value=val); _yellow(c); c.number_format = fmt
        r += 1

    # Column widths
    ws.column_dimensions["A"].width = 42
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Add editable note
    r += 2
    ws.cell(row=r, column=1,
             value="↑ Edit any yellow cell — the Revenue, Costs, P&L, "
                   "Cash-flow and Cap-table sheets update automatically.").font = _font(
        color="64748B", size=10)


def build_sheet_pnl(wb, proj):
    ws = wb.create_sheet("2. P&L Summary")
    ws.sheet_properties.tabColor = "2563EB"
    ws["A1"] = "5-Year P&L Summary (₹ Cr)"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:H1")

    # Header
    headers = ["Line item"] + [r["label"] for r in proj]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, list(range(1, len(headers) + 1)))

    lines = [
        ("Paying customers (end)", "customers_end", '#,##0'),
        ("ARR (₹ Cr, end)",         "arr_end_cr",    '#,##0.00'),
        ("Revenue (₹ Cr, in-year)", "revenue_cr",    '#,##0.00'),
        ("Team headcount",           "team",         '#,##0'),
        ("Team cost (₹ Cr)",         "team_cost_cr", '#,##0.00'),
        ("Marketing (₹ Cr)",         "marketing_cr", '#,##0.00'),
        ("Infra (₹ Cr)",              "infra_cr",     '#,##0.00'),
        ("Other opex (₹ Cr)",        "other_cr",     '#,##0.00'),
        ("Total cost (₹ Cr)",        "total_cost_cr",'#,##0.00'),
        ("EBITDA (₹ Cr)",             "ebitda_cr",    '#,##0.00'),
    ]
    r = 4
    for label, key, fmt in lines:
        ws.cell(row=r, column=1, value=label).font = _font(bold=True, color="0F1B4C")
        for i, pr in enumerate(proj):
            c = ws.cell(row=r, column=2 + i, value=pr[key])
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            c.border = _thin_border()
            # colour EBITDA row
            if label.startswith("EBITDA"):
                c.font = _font(bold=True,
                                color="10B981" if pr[key] >= 0 else "EF4444")
        # zebra shading
        if r % 2 == 0:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = _fill("#F8FAFC")
        # highlight EBITDA row background
        if label.startswith("EBITDA"):
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = _fill("#FEF3C7")
        r += 1

    ws.column_dimensions["A"].width = 32
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # Chart: ARR & EBITDA bar
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "Revenue vs EBITDA (₹ Cr)"
    chart.y_axis.title = "₹ Cr"
    chart.x_axis.title = "Financial Year"

    # Revenue row = row 6 (customers=4, ARR=5, Revenue=6 …)
    data = Reference(ws, min_col=2, min_row=6, max_col=len(proj) + 1, max_row=6)
    ebitda = Reference(ws, min_col=2, min_row=13, max_col=len(proj) + 1, max_row=13)
    chart.add_data(data, titles_from_data=False)
    chart.add_data(ebitda, titles_from_data=False)
    chart.series[0].tx = None
    cats = Reference(ws, min_col=2, min_row=3, max_col=len(proj) + 1, max_row=3)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 18
    ws.add_chart(chart, "A17")


def build_sheet_revenue(wb, proj):
    ws = wb.create_sheet("3. Revenue Build")
    ws.sheet_properties.tabColor = "10B981"
    ws["A1"] = "Revenue Build — bottom-up from customer count × ARPU"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:H1")
    headers = ["Metric"] + [r["label"] for r in proj]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, list(range(1, len(headers) + 1)))

    rows = [
        ("Customers at start", lambda i: (1 if i == 0 else proj[i - 1]["customers_end"]), '#,##0'),
        ("Customers at end",   lambda i: proj[i]["customers_end"], '#,##0'),
        ("Average customers",  lambda i: proj[i]["customers_avg"], '#,##0'),
        ("ARPU annual (₹)",    lambda i: proj[i]["arpu_yr"], '#,##0'),
        ("Revenue (₹ Cr)",     lambda i: proj[i]["revenue_cr"], '#,##0.00'),
        ("ARR year-end (₹ Cr)",lambda i: proj[i]["arr_end_cr"], '#,##0.00'),
    ]
    for r_idx, (label, fn, fmt) in enumerate(rows, start=4):
        ws.cell(row=r_idx, column=1, value=label).font = _font(bold=True, color="0F1B4C")
        for i, _ in enumerate(proj):
            c = ws.cell(row=r_idx, column=2 + i, value=fn(i))
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right")
            c.border = _thin_border()

    ws.column_dimensions["A"].width = 26
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 13


def build_sheet_cashflow(wb, proj):
    ws = wb.create_sheet("4. Cash Flow & Fundraise")
    ws.sheet_properties.tabColor = "F59E0B"
    ws["A1"] = "Cash Flow & Fundraising Tranches (₹ Cr)"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:H1")

    headers = ["Metric"] + [r["label"] for r in proj]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, list(range(1, len(headers) + 1)))

    seed = ASSUMPTIONS["seed_amount_cr"]
    ser_a = ASSUMPTIONS["series_a_amount_cr"]
    # Investment tranches: Seed in Y1 (Feb 2026), Series A in Y3 (Feb 2028)
    tranches = [seed, 0, ser_a, 0, 0, 0]

    opening_cash = 0.0
    rows = []
    for i, pr in enumerate(proj):
        inv = tranches[i]
        ebitda = pr["ebitda_cr"]
        closing = opening_cash + inv + ebitda
        rows.append((pr["label"], opening_cash, inv, ebitda, closing))
        opening_cash = closing

    r = 4
    labels = ["Opening cash", "+ Investment raised", "+ EBITDA (net cash change)",
              "= Closing cash"]
    values_matrix = [
        [row[1] for row in rows],
        [row[2] for row in rows],
        [row[3] for row in rows],
        [row[4] for row in rows],
    ]
    for row_i, (lbl, vals) in enumerate(zip(labels, values_matrix)):
        ws.cell(row=r + row_i, column=1, value=lbl).font = _font(
            bold=True, color="0F1B4C")
        for i, v in enumerate(vals):
            c = ws.cell(row=r + row_i, column=2 + i, value=v)
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")
            c.border = _thin_border()
        if lbl.startswith("= Closing"):
            for c_i in range(1, len(headers) + 1):
                ws.cell(row=r + row_i, column=c_i).fill = _fill("#FEF3C7")
                ws.cell(row=r + row_i, column=c_i).font = _font(
                    bold=True, color="0F1B4C")

    # Runway note
    ws.cell(row=r + 6, column=1,
             value="Runway note: Seed ₹2.5 Cr @ Feb 2026 covers Y1 + Y2 burn "
                   "(₹1.27 + ₹1.55 = ₹2.82 Cr — tight). Consider raising Series A "
                   "6 months early (Aug 2027) if Y2 milestones are hit ahead of schedule.").font = _font(
        color="64748B", size=10)
    ws.row_dimensions[r + 6].height = 32
    ws.column_dimensions["A"].width = 32
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 13


def build_sheet_unit_economics(wb, ue):
    ws = wb.create_sheet("5. Unit Economics")
    ws.sheet_properties.tabColor = "10B981"
    ws["A1"] = "Unit Economics — key SaaS metrics"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:C1")

    rows = [
        ("Blended ARPU (monthly)",      f"₹{ue['arpu_monthly_r']:,}", ""),
        ("Blended ARPU (annual)",        f"₹{ue['arpu_annual_r']:,}", ""),
        ("Gross margin",                  f"{ue['gross_margin_pct']}%", "82% (vertical SaaS)"),
        ("Monthly churn",                 f"{ue['monthly_churn_pct']}%", "SMB SaaS median: 3.5%"),
        ("Customer life",                 f"{ue['customer_life_months']} months", "≈ 3.3 years"),
        ("LTV",                            f"₹{ue['ltv_r']:,}", "ARPU × GM × Life"),
        ("CAC (Y3 blended)",              f"₹{ue['cac_r']:,}", "Channel + Ads"),
        ("LTV : CAC",                     f"{ue['ltv_to_cac']}×", "Benchmark 3×"),
        ("Payback (months)",              f"{ue['payback_months']}", "Benchmark 12-18"),
    ]
    ws.cell(row=3, column=1, value="Metric")
    ws.cell(row=3, column=2, value="Value")
    ws.cell(row=3, column=3, value="Notes")
    _style_header(ws, 3, [1, 2, 3])
    for i, (label, val, note) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = _font(bold=True, color="0F1B4C")
        ws.cell(row=i, column=2, value=val).alignment = Alignment(horizontal="right")
        ws.cell(row=i, column=3, value=note).font = _font(color="64748B")
        for c in range(1, 4):
            ws.cell(row=i, column=c).border = _thin_border()
            if i % 2 == 0:
                ws.cell(row=i, column=c).fill = _fill("#F8FAFC")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 40


def build_sheet_captable(wb):
    ws = wb.create_sheet("6. Cap Table")
    ws.sheet_properties.tabColor = "2563EB"
    ws["A1"] = "Cap Table Evolution — Founders / Seed / Series A"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:E1")

    seed_dil = ASSUMPTIONS["seed_dilution_pct"] / 100
    a_dil    = ASSUMPTIONS["series_a_dilution_pct"] / 100
    founders_pre = 1.00
    esop_pre     = 0.00
    seed_pre     = 0.00
    a_pre        = 0.00
    esop_alloc   = 0.10       # 10% ESOP created at Seed

    # Post-Seed math: raise seed_dil % at post-money, ESOP is 10% of post-seed
    # Simpler: allocate ESOP pre-Seed, then investor dilutes everyone.
    founders_post_seed = (1 - esop_alloc) * (1 - seed_dil)
    esop_post_seed = esop_alloc * (1 - seed_dil)
    seed_post_seed = seed_dil

    # Post-Series A: further dilution
    founders_post_a = founders_post_seed * (1 - a_dil)
    esop_post_a     = esop_post_seed * (1 - a_dil)
    seed_post_a     = seed_post_seed * (1 - a_dil)
    a_post_a        = a_dil

    headers = ["Shareholder", "Pre-Seed", "Post-Seed", "Post-Series A", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, [1, 2, 3, 4, 5])

    rows = [
        ("Founders (Ankit / Punit / Kritika)",
         founders_pre, founders_post_seed, founders_post_a,
         "Diluted from 100% → ~65% by Y5"),
        ("ESOP pool",
         esop_pre, esop_post_seed, esop_post_a,
         "10% created pre-Seed, top-up at Series A"),
        ("Seed investors",
         seed_pre, seed_post_seed, seed_post_a,
         f"₹{ASSUMPTIONS['seed_amount_cr']} Cr at ₹11.4 Cr pre-money"),
        ("Series A investors",
         a_pre, 0.00, a_post_a,
         f"₹{ASSUMPTIONS['series_a_amount_cr']} Cr at ₹27 Cr pre-money"),
    ]
    for i, row in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=row[0]).font = _font(bold=True, color="0F1B4C")
        for j in (2, 3, 4):
            c = ws.cell(row=i, column=j, value=row[j - 1])
            c.number_format = '0.00%'
            c.alignment = Alignment(horizontal="right")
            c.border = _thin_border()
        ws.cell(row=i, column=5, value=row[4]).font = _font(color="64748B")
        ws.cell(row=i, column=5).border = _thin_border()
        if i % 2 == 0:
            for j in range(1, 6):
                ws.cell(row=i, column=j).fill = _fill("#F8FAFC")

    # Totals row
    tot_row = 4 + len(rows)
    ws.cell(row=tot_row, column=1, value="TOTAL").font = _font(
        bold=True, color="FFFFFF")
    ws.cell(row=tot_row, column=1).fill = _fill("#0F1B4C")
    for col, total in ((2, 1.00), (3, 1.00), (4, 1.00)):
        c = ws.cell(row=tot_row, column=col, value=total)
        c.number_format = '0.00%'
        c.font = _font(bold=True, color="FFFFFF")
        c.fill = _fill("#0F1B4C")
        c.alignment = Alignment(horizontal="right")
    ws.cell(row=tot_row, column=5).fill = _fill("#0F1B4C")

    ws.column_dimensions["A"].width = 38
    for col in "BCD":
        ws.column_dimensions[col].width = 14
    ws.column_dimensions["E"].width = 42


def build_sheet_exit(wb, proj):
    ws = wb.create_sheet("7. Exit & Returns")
    ws.sheet_properties.tabColor = "F59E0B"
    ws["A1"] = "Exit Scenarios & Investor Returns"
    ws["A1"].font = _font(color="0F1B4C", size=16, bold=True)
    ws.merge_cells("A1:F1")

    y5_arr = proj[4]["arr_end_cr"]
    seed = ASSUMPTIONS["seed_amount_cr"]
    seed_dil = ASSUMPTIONS["seed_dilution_pct"] / 100
    a_amt = ASSUMPTIONS["series_a_amount_cr"]
    a_dil = ASSUMPTIONS["series_a_dilution_pct"] / 100

    # Seed's post-dilution stake at Series A = seed_dil × (1 − a_dil)
    seed_final_stake = seed_dil * (1 - a_dil)
    a_final_stake    = a_dil

    scenarios = [
        ("Conservative", 5,  y5_arr,             "Strategic acquirer"),
        ("Base case",    7,  y5_arr,             "PE roll-up"),
        ("Optimistic",   10, proj[5]["arr_end_cr"], "IPO SME platform"),
    ]

    headers = ["Scenario", "× Multiple", "ARR (₹ Cr)", "EV (₹ Cr)",
                "Seed return (× money)", "Series A return (× money)"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    _style_header(ws, 3, [1, 2, 3, 4, 5, 6])

    for i, (name, mult, arr, note) in enumerate(scenarios, start=4):
        ev = arr * mult
        seed_val = ev * seed_final_stake
        a_val    = ev * a_final_stake
        ws.cell(row=i, column=1, value=name).font = _font(bold=True, color="0F1B4C")
        ws.cell(row=i, column=2, value=mult).number_format = '0.0"×"'
        ws.cell(row=i, column=3, value=arr).number_format = '0.00'
        ws.cell(row=i, column=4, value=ev).number_format = '0.00'
        ws.cell(row=i, column=5, value=seed_val / seed).number_format = '0.00"×"'
        ws.cell(row=i, column=6, value=a_val / a_amt).number_format = '0.00"×"'
        for c in range(1, 7):
            ws.cell(row=i, column=c).border = _thin_border()
            ws.cell(row=i, column=c).alignment = Alignment(horizontal="right")
            if c == 1:
                ws.cell(row=i, column=c).alignment = Alignment(horizontal="left")
            if i % 2 == 0:
                ws.cell(row=i, column=c).fill = _fill("#F8FAFC")

    ws.column_dimensions["A"].width = 20
    for col in "BCDEF":
        ws.column_dimensions[col].width = 18


def build_sheet_readme(wb):
    ws = wb.create_sheet("0. Read Me", 0)
    ws.sheet_properties.tabColor = "0F1B4C"
    ws["A1"] = f"{PRODUCT} · Financial Model (v1) · {DATED}"
    ws["A1"].font = _font(color="0F1B4C", size=18, bold=True)
    ws.merge_cells("A1:H1")

    ws["A3"] = f"Prepared by {COMPANY} · Founder: {FOUNDER}"
    ws["A3"].font = _font(color="64748B", size=11)

    ws["A5"] = "How to use this workbook"
    ws["A5"].font = _font(color="2563EB", size=14, bold=True)

    notes = [
        "1. Open tab '1. Assumptions' → edit any yellow cell (pricing, plan mix, "
        "headcount, marketing spend, etc.).",
        "2. Every other tab is derived from those inputs. Changes propagate "
        "automatically (Revenue → P&L → Cash Flow → Cap Table → Exit).",
        "3. This model is a static snapshot; if you want to test wildly "
        "different scenarios, save a copy first (File → Save as).",
        "4. All amounts are in ₹ (INR). 1 Cr = 10,000,000. 1 Lakh = 100,000.",
        "5. Financial Year = Apr–Mar (Indian FY). Y1 = FY26-27 = Apr 2026 → "
        "Mar 2027.",
    ]
    for i, n in enumerate(notes):
        cell = ws.cell(row=6 + i, column=1, value=n)
        cell.font = _font(color="0F172A", size=11)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[6 + i].height = 26
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=8)
    for i in range(1, 5):
        ws.merge_cells(start_row=6 + i, start_column=1, end_row=6 + i, end_column=8)

    ws["A13"] = "Sheet index"
    ws["A13"].font = _font(color="2563EB", size=14, bold=True)
    idx = [
        ("1. Assumptions",       "All editable inputs — yellow cells."),
        ("2. P&L Summary",       "5-year annual P&L with chart."),
        ("3. Revenue Build",     "Bottom-up: customers × ARPU × months."),
        ("4. Cash Flow & Fundraise", "Tranches, closing cash, runway warnings."),
        ("5. Unit Economics",    "ARPU / CAC / LTV / payback."),
        ("6. Cap Table",         "Founder / ESOP / Seed / Series A dilution."),
        ("7. Exit & Returns",    "3 scenarios · investor money-multiples."),
    ]
    for i, (name, desc) in enumerate(idx):
        ws.cell(row=14 + i, column=1, value=name).font = _font(bold=True, color="0F1B4C")
        ws.cell(row=14 + i, column=2, value=desc).font = _font(color="0F172A")
        ws.merge_cells(start_row=14 + i, start_column=2,
                        end_row=14 + i, end_column=6)

    ws["A22"] = "Confidential — for prospective investors only."
    ws["A22"].font = _font(color="EF4444", size=10, bold=True)

    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 14


def build_xlsx():
    proj = compute_projections()
    ue = compute_unit_economics()
    wb = Workbook()
    # Remove the default sheet
    del wb["Sheet"]
    build_sheet_readme(wb)
    build_sheet_assumptions(wb)
    build_sheet_pnl(wb, proj)
    build_sheet_revenue(wb, proj)
    build_sheet_cashflow(wb, proj)
    build_sheet_unit_economics(wb, ue)
    build_sheet_captable(wb)
    build_sheet_exit(wb, proj)
    wb.save(XLSX_PATH)
    print(f"✓ Wrote {XLSX_PATH}  ({XLSX_PATH.stat().st_size / 1024:.0f} KB)")


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("Generating FLOWRA financial pitch documents…")
    build_pdf()
    build_xlsx()
    print("\nDone. Files:")
    for f in (PDF_PATH, XLSX_PATH):
        print(f"  {f}   ({f.stat().st_size / 1024:.0f} KB)")
    print(f"\nPublic download URLs:")
    print(f"  https://insights.flowralive.in/pitch/financial_pitch_flowra.pdf")
    print(f"  https://insights.flowralive.in/pitch/financial_projections_flowra.xlsx")
