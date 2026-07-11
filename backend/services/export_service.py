import csv
import io
import os
from typing import List, Dict, Any
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)

LOGO_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'flowra-logo.png')
BRAND_COLOR = '#2563EB'
BRAND_COLOR_DARK = '#0F1B4C'
BG_STRIPE = '#F0F4FF'


class ExportService:
    """Service for exporting reports in various formats (PDF, Excel, CSV)"""

    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "report.csv",
                      company_name: str = "") -> io.BytesIO:
        # iter-111: previously used io.TextIOWrapper(BytesIO()) — that pattern is
        # unreliable across Python versions because the wrapper is GC'd on
        # return and may close its underlying BytesIO before FastAPI streams it,
        # producing an empty download. The deterministic fix is to build the
        # CSV as a text string first, then return its UTF-8-encoded bytes.
        text_buf = io.StringIO(newline='')
        if not data:
            return io.BytesIO(b"")
        # iter-121: prepend the useradmin company name as a banner row so
        # CSV opens in Excel with the same context as PDF/xlsx exports.
        if company_name:
            text_buf.write(f"{company_name}\n\n")
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(text_buf, fieldnames=fieldnames)
        writer.writeheader()
        for row in data:
            # csv.DictWriter expects scalars — coerce dict / list values to str.
            safe_row = {k: (v if isinstance(v, (str, int, float, bool)) or v is None else str(v))
                        for k, v in row.items()}
            writer.writerow(safe_row)
        encoded = text_buf.getvalue().encode('utf-8-sig')  # BOM helps Excel auto-detect UTF-8
        return io.BytesIO(encoded)

    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], report_type: str = "Report",
                        company_name: str = "") -> io.BytesIO:
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = report_type

        if not data:
            wb.save(output)
            output.seek(0)
            return output

        # FLOWRA blue header
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        stripe_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")

        headers = list(data[0].keys())

        # Row 1: Company banner (spans all columns) — iter-121, replaces
        # the previous "Anonymous"/blank title with the useradmin's synced
        # company name pulled from db.sync_status.
        banner_row = 1
        header_row = 1
        first_data_row = 2
        if company_name:
            title_cell = ws.cell(row=1, column=1, value=company_name)
            title_cell.font = Font(bold=True, color="0F1B4C", size=14)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            if len(headers) > 1:
                ws.merge_cells(start_row=1, start_column=1,
                               end_row=1, end_column=len(headers))
            subtitle_cell = ws.cell(row=2, column=1, value=f"{report_type} Report")
            subtitle_cell.font = Font(italic=True, color="64748B", size=10)
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            if len(headers) > 1:
                ws.merge_cells(start_row=2, start_column=1,
                               end_row=2, end_column=len(headers))
            ws.row_dimensions[1].height = 22
            ws.row_dimensions[2].height = 16
            header_row = 3
            first_data_row = 4

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_offset, row_data in enumerate(data):
            row_idx = first_data_row + row_offset
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                # iter-111: openpyxl rejects list/dict cell values with
                # "Cannot convert [] to Excel". Coerce non-scalars to a
                # readable string so the export never bombs on aliases /
                # nested objects.
                if value is None:
                    safe_value = ""
                elif isinstance(value, (str, int, float, bool)):
                    safe_value = value
                elif isinstance(value, (list, tuple)):
                    safe_value = ", ".join(str(v) for v in value)
                else:
                    safe_value = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=safe_value)
                if isinstance(safe_value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                # Alternating stripe (based on relative data row index)
                if (row_offset % 2) == 1:
                    cell.fill = stripe_fill

        # iter-121: `ws.columns` returns MergedCell objects for cells inside
        # the banner merge range, and MergedCell has no `.column_letter`.
        # Iterate by column index using `get_column_letter` instead.
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 10
            for row_idx in range(header_row, first_data_row + len(data)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], report_type: str = "Report",
                      title: str = "FLOWRA Report", company_name: str = "") -> io.BytesIO:
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()

        # Add logo if available
        if os.path.exists(LOGO_PATH):
            try:
                logo = Image(LOGO_PATH, width=1.8*inch, height=0.6*inch)
                logo.hAlign = 'LEFT'
                elements.append(logo)
                elements.append(Spacer(1, 0.15*inch))
            except Exception as e:
                logger.warning(f"Could not add logo to PDF: {e}")

        # iter-121: heading is now the ACTUAL synced company name (falls
        # back to the passed-in `title` — previously all PDFs showed
        # "Anonymous" / hardcoded "FLOWRA Report" regardless of tenant).
        header_text = (company_name or title or "FLOWRA Report").strip()
        title_style = ParagraphStyle(
            'FlowraTitle', parent=styles['Title'],
            textColor=colors.HexColor(BRAND_COLOR_DARK),
            fontSize=18, spaceAfter=6, alignment=1
        )
        elements.append(Paragraph(f"<b>{header_text}</b>", title_style))

        # Subtitle
        sub_style = ParagraphStyle(
            'FlowraSub', parent=styles['Heading2'],
            textColor=colors.HexColor(BRAND_COLOR),
            fontSize=12, spaceAfter=12, alignment=1
        )
        elements.append(Paragraph(f"{report_type} Report", sub_style))
        elements.append(Spacer(1, 0.15*inch))

        if not data:
            elements.append(Paragraph("No data available", styles['Normal']))
            doc.build(elements)
            output.seek(0)
            return output

        # Table
        headers = list(data[0].keys())
        table_data = [headers]
        for row in data:
            table_data.append([str(row.get(h, "")) for h in headers])

        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(BRAND_COLOR)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor(BG_STRIPE)])
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

        # Footer
        footer_style = ParagraphStyle(
            'FlowraFooter', parent=styles['Normal'],
            textColor=colors.HexColor('#94A3B8'),
            fontSize=7, alignment=1
        )
        elements.append(Paragraph("Generated by FLOWRA | Organize. Automate. Accelerate.", footer_style))

        doc.build(elements)
        output.seek(0)
        return output
