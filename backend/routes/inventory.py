from fastapi import APIRouter, Request
from typing import Optional
import math
from datetime import datetime
import logging

from db import db
from models import (
    InventoryItem, SalesVoucher, APIResponse,
    PurchaseOrder, PurchaseOrderItem
)
from utils import safe_num, filter_vouchers_by_fy, fy_to_date_range
from services.purchase_order_ai import PurchaseOrderAI
from services.tenant_context import get_tenant_context
from services.auth_service import get_current_user
from services.audit_service import log_audit, get_client_ip
from routes.branch_ledgers import get_branch_parties

logger = logging.getLogger(__name__)
router = APIRouter()


def _build_query(ctx, company_id=None, extra=None):
    """Build a tenant-filtered query dict."""
    q = {}
    if ctx and ctx.get("tenant_id"):
        q["tenant_id"] = ctx["tenant_id"]
    cid = company_id or (ctx.get("company_id") if ctx else None)
    if cid:
        q["company_id"] = cid
    if extra:
        q.update(extra)
    return q


async def _get_branch_set(request, ctx):
    """Return set of branch party names if X-Exclude-Branches header is set, else empty set."""
    if request.headers.get("X-Exclude-Branches", "").lower() != "true":
        return set()
    bp = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
    return set(bp) if bp else set()


async def _get_purchase_branch_set(ctx):
    """Detect branch-like parties in purchase vouchers (non-sundry-creditor).
    These are internal transfers, not actual procurement from external suppliers."""
    from services.id_mapping_service import get_company_name
    import re
    if not ctx:
        return set()
    tenant_id = ctx.get("tenant_id", "")
    company_id = ctx.get("company_id", "")
    company_name = await get_company_name(tenant_id, company_id)
    if not company_name:
        return set()
    name_clean = re.sub(r'\b(private|limited|pvt|ltd|llp|inc|corp)\b', '', company_name, flags=re.IGNORECASE).strip()
    tokens = [w.lower() for w in name_clean.split() if len(w) > 3]
    if not tokens:
        return set()
    parties = await db.purchase_vouchers.distinct(
        "party_name",
        {"tenant_id": tenant_id, "company_id": company_id}
    )
    branch_parties = set()
    for party in parties:
        if not party:
            continue
        party_lower = party.lower()
        matching = sum(1 for t in tokens if t in party_lower)
        if matching >= 2:
            branch_parties.add(party)
    return branch_parties


def _filter_branch_vouchers(vouchers, branch_set):
    """Filter out vouchers whose party_name is in branch_set."""
    if not branch_set:
        return vouchers
    return [v for v in vouchers if v.get("party_name") not in branch_set]


@router.patch("/inventory/items/{item_id}/abc")
async def set_abc_category(request: Request, item_id: str):
    """Set A/B/C/D classification for a single inventory item.
    Body: { "abc_category": "A" | "B" | "C" | "D" | "" }  (empty unsets)
    """
    try:
        ctx = await get_tenant_context(request)
        body = await request.json()
        cat = (body.get("abc_category") or "").strip().upper()
        if cat and cat not in ("A", "B", "C", "D"):
            return APIResponse(success=False, error="abc_category must be A, B, C, or D")
        q = _build_query(ctx, body.get("company_id"))
        update = {"$set": {"abc_category": cat}} if cat else {"$unset": {"abc_category": ""}}
        result = await db.inventory_items.update_one({**q, "item_id": item_id}, update)
        if result.matched_count == 0:
            # Try by item_name fallback (item_id sometimes is the name)
            result = await db.inventory_items.update_one({**q, "item_name": item_id}, update)
        return APIResponse(success=True, data={"matched": result.matched_count, "modified": result.modified_count})
    except Exception as e:
        return APIResponse(success=False, error=str(e))


@router.post("/inventory/abc/auto-assign")
async def auto_assign_abc(request: Request):
    """Bulk-assign A/B/C/D using Pareto / 80-15-4-1 rule on FY revenue:
       A = top 80% of revenue, B = next 15%, C = next 4%, D = remainder.
    Body: { "fy": "2026-27", "company_id": "..." }
    """
    try:
        ctx = await get_tenant_context(request)
        body = await request.json()
        fy = body.get("fy") or ""
        q = _build_query(ctx, body.get("company_id"))

        # Pull sales vouchers and tally per item_name (FY-scoped if provided)
        sales = await db.sales_vouchers.find(q, {"_id": 0, "voucher_date": 1, "items": 1}).to_list(50000)
        if fy:
            from utils import filter_vouchers_by_fy
            sales = filter_vouchers_by_fy(sales, fy)
        from collections import defaultdict
        rev_by_item = defaultdict(float)
        for v in sales:
            for vi in v.get("items", []) or []:
                iname = (vi.get("item") or vi.get("item_name") or "").strip().lower()
                if not iname:
                    continue
                rev_by_item[iname] += abs(safe_num(vi.get("amount") or vi.get("value") or 0))

        total_rev = sum(rev_by_item.values())
        if total_rev <= 0:
            return APIResponse(success=False, error="No sales revenue found for the selected FY")

        # Sort items by revenue descending, then assign A/B/C/D by cumulative %
        sorted_items = sorted(rev_by_item.items(), key=lambda x: -x[1])
        cum = 0.0
        item_to_abc = {}
        for iname, rev in sorted_items:
            cum += rev
            pct = cum / total_rev * 100
            if pct <= 80:
                item_to_abc[iname] = "A"
            elif pct <= 95:
                item_to_abc[iname] = "B"
            elif pct <= 99:
                item_to_abc[iname] = "C"
            else:
                item_to_abc[iname] = "D"

        # Apply to inventory_items
        all_inv = await db.inventory_items.find(q, {"_id": 0, "item_id": 1, "item_name": 1}).to_list(50000)
        modified = 0
        for it in all_inv:
            iname = (it.get("item_name") or "").strip().lower()
            if iname in item_to_abc:
                await db.inventory_items.update_one(
                    {**q, "item_id": it.get("item_id")},
                    {"$set": {"abc_category": item_to_abc[iname]}},
                )
                modified += 1
            else:
                # Items with zero revenue → D
                await db.inventory_items.update_one(
                    {**q, "item_id": it.get("item_id")},
                    {"$set": {"abc_category": "D"}},
                )

        # Counts
        counts = {"A": 0, "B": 0, "C": 0, "D": 0}
        for v in item_to_abc.values():
            counts[v] += 1
        counts["D"] += len(all_inv) - len(item_to_abc)

        return APIResponse(success=True, data={"counts": counts, "modified": modified, "total_items": len(all_inv)})
    except Exception as e:
        logger.error(f"ABC auto-assign error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/category-sales")
async def category_sales_drill(request: Request, abc: str, fy: Optional[str] = None, company_id: Optional[str] = None):
    """For an ABC category, return:
      - items in that category with FY qty + revenue
      - per-item top customers (qty + revenue)
    Used by the Inventory Analytics "Category Sales" tab.

    Honors the 'X-Exclude-Branches: true' header — branch transfers are
    excluded from both totals and customer breakdowns.
    """
    try:
        from collections import defaultdict
        from utils import filter_vouchers_by_fy
        from routes.branch_ledgers import get_branch_parties
        ctx = await get_tenant_context(request)
        cat = (abc or "").upper().strip()
        if cat not in ("A", "B", "C", "D"):
            return APIResponse(success=False, error="abc must be A, B, C, or D")
        q = _build_query(ctx, company_id)

        # Branch exclusion (driven by global navbar toggle via X-Exclude-Branches header)
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        branch_set = set()
        if exclude_branches:
            tid = (ctx or {}).get("tenant_id") or ""
            cid = (ctx or {}).get("company_id") or company_id or ""
            bp = await get_branch_parties(tid, cid)
            branch_set = {p.lower().strip() for p in bp}

        items = await db.inventory_items.find({**q, "abc_category": cat},
                                               {"_id": 0, "item_id": 1, "item_name": 1,
                                                "part_number": 1, "stock_group": 1,
                                                "quantity": 1, "price": 1, "standard_price": 1}).to_list(5000)
        if not items:
            return APIResponse(success=True, data={"abc": cat, "items": [], "summary": {"items": 0, "revenue": 0, "qty": 0}})

        sales = await db.sales_vouchers.find(q, {"_id": 0, "voucher_date": 1, "party_name": 1, "items": 1}).to_list(50000)
        if fy:
            sales = filter_vouchers_by_fy(sales, fy)

        # Aggregate per-item: total qty/revenue + per-customer breakdown + frequency
        per_item = {}
        for it in items:
            per_item[(it.get("item_name") or "").lower().strip()] = {
                "item_name": it.get("item_name", ""),
                "part_number": it.get("part_number", ""),
                "stock_group": it.get("stock_group", ""),
                "current_stock": safe_num(it.get("quantity")),
                "standard_price": safe_num(it.get("standard_price")),
                "total_qty": 0.0,
                "total_revenue": 0.0,
                "order_count": 0,
                "customers": defaultdict(lambda: {"qty": 0.0, "revenue": 0.0, "count": 0}),
            }

        for v in sales:
            party = (v.get("party_name") or "").strip()
            # Skip branch transfers entirely when toggle is on
            if branch_set and party.lower() in branch_set:
                continue
            for vi in v.get("items", []) or []:
                iname = (vi.get("item") or vi.get("item_name") or "").strip().lower()
                if iname not in per_item:
                    continue
                qty = abs(safe_num(vi.get("quantity")))
                rev = abs(safe_num(vi.get("amount") or vi.get("value") or 0))
                row = per_item[iname]
                row["total_qty"] += qty
                row["total_revenue"] += rev
                row["order_count"] += 1
                if party:
                    cb = row["customers"][party]
                    cb["qty"] += qty
                    cb["revenue"] += rev
                    cb["count"] += 1

        # Finalize: convert customers dict to sorted list, aggregate totals
        result_items = []
        sum_rev = 0.0
        sum_qty = 0.0
        for row in per_item.values():
            top_customers = sorted(
                [{"customer_name": k, "qty": round(v["qty"], 2),
                  "revenue": round(v["revenue"], 2), "count": v["count"]}
                 for k, v in row["customers"].items()],
                key=lambda x: -x["revenue"],
            )[:10]
            sum_rev += row["total_revenue"]
            sum_qty += row["total_qty"]
            result_items.append({
                "item_name": row["item_name"],
                "part_number": row["part_number"],
                "stock_group": row["stock_group"],
                "current_stock": row["current_stock"],
                "standard_price": row["standard_price"],
                "total_qty": round(row["total_qty"], 2),
                "total_revenue": round(row["total_revenue"], 2),
                "order_count": row["order_count"],
                "top_customers": top_customers,
            })
        result_items.sort(key=lambda x: -x["total_revenue"])

        return APIResponse(success=True, data={
            "abc": cat, "fy": fy, "items": result_items,
            "summary": {"items": len(result_items),
                        "revenue": round(sum_rev, 2),
                        "qty": round(sum_qty, 2)},
        })
    except Exception as e:
        import traceback
        logger.error(f"Category sales error: {e}\n{traceback.format_exc()}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/items")
async def get_inventory_items(request: Request, category: Optional[str] = None, stock_group: Optional[str] = None, min_quantity: Optional[float] = None, company_id: Optional[str] = None, fy: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        extra = {}
        if category and category != 'all':
            extra["category"] = category
        if stock_group and stock_group != 'all':
            extra["stock_group"] = stock_group
        if min_quantity is not None:
            extra["quantity"] = {"$gte": min_quantity}

        query = _build_query(ctx, company_id, extra)
        items = await db.inventory_items.find(query, {"_id": 0}).to_list(5000)

        # If FY is specified, compute closing stock for that FY from vouchers
        if fy:
            base_q = _build_query(ctx, company_id)
            sales_v = await db.sales_vouchers.find(base_q, {"_id": 0}).to_list(50000)
            purchase_v = await db.purchase_vouchers.find(base_q, {"_id": 0}).to_list(50000)

            # Apply branch filter
            branch_set = await _get_branch_set(request, ctx)
            sales_v = _filter_branch_vouchers(sales_v, branch_set)
            purchase_v = _filter_branch_vouchers(purchase_v, branch_set)

            # FY end date
            fy_end_year = int(fy.split('-')[0]) + 1 if '-' in fy else 2027
            fy_end_date = f"{fy_end_year}-03-31"

            # For the selected FY, closing = current_stock + sales_after_fy - purchases_after_fy
            from collections import defaultdict
            post_fy_sold = defaultdict(float)
            post_fy_purchased = defaultdict(float)
            for v in sales_v:
                vdate = v.get("voucher_date", v.get("date", ""))
                if vdate > fy_end_date:
                    for vi in v.get("items", []):
                        iname = (vi.get("item", "") or "").strip().lower()
                        post_fy_sold[iname] += abs(safe_num(vi.get("quantity", 0)))
            for v in purchase_v:
                vdate = v.get("voucher_date", v.get("date", ""))
                if vdate > fy_end_date:
                    for vi in v.get("items", []):
                        iname = (vi.get("item", "") or "").strip().lower()
                        post_fy_purchased[iname] += abs(safe_num(vi.get("quantity", 0)))

            for item in items:
                iname = (item.get("item_name") or "").strip().lower()
                current_qty = safe_num(item.get("quantity"))
                # Closing of FY = current_qty + post_fy_sales - post_fy_purchases
                fy_closing = current_qty + post_fy_sold.get(iname, 0) - post_fy_purchased.get(iname, 0)
                item["quantity"] = round(fy_closing, 2)
                # Recalc value
                price = safe_num(item.get("price"))
                item["closing_value"] = round(fy_closing * price, 2)

        base_q = _build_query(ctx, company_id)
        all_items = await db.inventory_items.find(base_q, {"_id": 0, "stock_group": 1}).to_list(5000)
        stock_groups = sorted(list(set(item.get("stock_group", "General") for item in all_items if item.get("stock_group"))))

        return APIResponse(
            success=True,
            data={"items": items, "count": len(items), "stock_groups": stock_groups}
        )
    except Exception as e:
        logger.error(f"Error fetching inventory: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/summary")
async def get_inventory_summary(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)

        if not items:
            return APIResponse(
                success=True,
                data={"total_items": 0, "total_value": 0, "low_stock_items": 0, "categories": []}
            )

        total_items = len(items)

        # Compute total value: prefer closing_value, then qty*price
        total_value = 0.0
        for item in items:
            cv = safe_num(item.get("closing_value"))
            if cv > 0:
                total_value += cv
            else:
                total_value += safe_num(item.get("quantity")) * safe_num(item.get("price"))

        # Low stock: use movement-based analysis
        # Items with qty=0 but that had sales activity are out-of-stock (genuinely low)
        # Items with qty=0 and no sales data: skip (master data only)
        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        # Apply branch filter if header set
        branch_set = await _get_branch_set(request, ctx)
        all_vouchers = _filter_branch_vouchers(all_vouchers, branch_set)
        fy_vouchers = filter_vouchers_by_fy(all_vouchers, fy) if fy else all_vouchers

        # Build set of items that had sales (i.e. actively traded)
        active_items = set()
        for v in fy_vouchers:
            for vi in v.get("items", []):
                iname = vi.get("item", "").strip()
                if iname:
                    active_items.add(iname.lower())

        low_stock_items = 0
        for item in items:
            qty = safe_num(item.get("quantity"))
            name = (item.get("item_name") or "").lower()
            reorder = safe_num(item.get("reorder_level"))
            if qty > 0 and reorder > 0 and qty < reorder:
                low_stock_items += 1
            elif qty == 0 and name in active_items:
                # Out of stock but actively sold — flag as low stock
                low_stock_items += 1

        categories = list(set(item.get("category") for item in items if item.get("category")))

        fy_sales_value = sum(safe_num(v.get("total_amount")) for v in fy_vouchers)

        return APIResponse(
            success=True,
            data={
                "total_items": total_items,
                "total_value": round(total_value, 2),
                "low_stock_items": low_stock_items,
                "categories": categories,
                "fy_sales_value": round(fy_sales_value, 2)
            }
        )
    except Exception as e:
        logger.error(f"Error getting inventory summary: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/inventory/generate-purchase-order")
async def generate_purchase_order(request: Request, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        inventory_items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        sales_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)

        po_ai = PurchaseOrderAI()
        result = await po_ai.generate_purchase_order(inventory_items, sales_vouchers)

        if result.get("success"):
            po_number = f"PO-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            po_data = result.get("purchase_order", {})

            po_items = []
            for item in po_data.get("urgent_items", po_data.get("items", [])):
                try:
                    po_items.append(PurchaseOrderItem(**item))
                except Exception:
                    po_items.append(PurchaseOrderItem(
                        item_name=str(item.get("item_name", item.get("name", "Unknown"))),
                        current_stock=safe_num(item.get("current_stock")),
                        recommended_quantity=safe_num(item.get("recommended_quantity", item.get("quantity", 0))),
                        priority=str(item.get("priority", "medium")),
                        reason=str(item.get("reason", "")),
                        estimated_cost=safe_num(item.get("estimated_cost", item.get("cost", 0)))
                    ))

            purchase_order = PurchaseOrder(
                po_number=po_number,
                items=po_items,
                total_items=len(po_items),
                total_cost=po_data.get("total_estimated_cost", sum(i.estimated_cost for i in po_items)),
                ai_analysis=po_data.get("analysis", ""),
                status="draft"
            )

            doc = purchase_order.model_dump()
            doc['created_at'] = doc['created_at'].isoformat()
            await db.purchase_orders.insert_one(doc)

            return APIResponse(
                success=True,
                data=po_data,
                message=f"Purchase order {po_number} generated"
            )
        else:
            return APIResponse(success=False, error=result.get("error"))
    except Exception as e:
        logger.error(f"Error generating purchase order: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/purchase-orders")
async def get_purchase_orders(request: Request, status: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        extra = {}
        if status:
            extra["status"] = status
        query = _build_query(ctx, company_id, extra)
        pos = await db.purchase_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
        return APIResponse(success=True, data={"purchase_orders": pos, "count": len(pos)})
    except Exception as e:
        logger.error(f"Error fetching purchase orders: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/sales-frequency")
async def get_sales_frequency(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)
        all_vouchers = _filter_branch_vouchers(all_vouchers, branch_set)
        sales_vouchers = filter_vouchers_by_fy(all_vouchers, fy)

        if start_date or end_date:
            filtered_vouchers = []
            for v in sales_vouchers:
                v_date = v.get("voucher_date", "")
                if start_date and v_date < start_date:
                    continue
                if end_date and v_date > end_date:
                    continue
                filtered_vouchers.append(v)
            sales_vouchers = filtered_vouchers

        item_stats = {}
        for voucher in sales_vouchers:
            party = voucher.get("party_name", "Unknown")
            for item in voucher.get("items", []):
                item_name = item.get("item", "")
                qty = safe_num(item.get("quantity"))

                if item_name not in item_stats:
                    item_stats[item_name] = {
                        "item_name": item_name,
                        "total_quantity_sold": 0,
                        "transaction_count": 0,
                        "unique_customers": set(),
                        "total_revenue": 0
                    }

                item_stats[item_name]["total_quantity_sold"] += qty
                item_stats[item_name]["transaction_count"] += 1
                item_stats[item_name]["unique_customers"].add(party)
                item_stats[item_name]["total_revenue"] += qty * safe_num(item.get("rate"))

        frequency_data = []
        for item_name, stats in item_stats.items():
            frequency_data.append({
                "item_name": item_name,
                "total_quantity_sold": stats["total_quantity_sold"],
                "transaction_count": stats["transaction_count"],
                "unique_customers": len(stats["unique_customers"]),
                "total_revenue": round(stats["total_revenue"], 2),
                "avg_quantity_per_transaction": round(
                    stats["total_quantity_sold"] / stats["transaction_count"], 1
                ) if stats["transaction_count"] > 0 else 0,
                "customer_list": list(stats["unique_customers"])
            })

        frequency_data.sort(key=lambda x: x["transaction_count"], reverse=True)

        return APIResponse(success=True, data={"frequency": frequency_data, "total_items": len(frequency_data)})
    except Exception as e:
        logger.error(f"Error getting sales frequency: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/movement-analysis")
async def get_inventory_movement(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        # Guard: if the requested FY is entirely BEFORE the earliest synced voucher,
        # there's no data to display. Returning today's master quantities would
        # incorrectly suggest stock levels for an un-synced period.
        if fy:
            try:
                fy_start_str, fy_end_str = fy_to_date_range(fy)
            except Exception:
                fy_start_str = fy_end_str = None
            if fy_start_str and fy_end_str:
                earliest = await db.sales_vouchers.find(q, {"_id": 0, "voucher_date": 1}).sort("voucher_date", 1).limit(1).to_list(1)
                earliest_purch = await db.purchase_vouchers.find(q, {"_id": 0, "voucher_date": 1}).sort("voucher_date", 1).limit(1).to_list(1)
                earliest_voucher = None
                for d in (earliest, earliest_purch):
                    if d and d[0].get("voucher_date"):
                        ev = d[0]["voucher_date"]
                        if earliest_voucher is None or ev < earliest_voucher:
                            earliest_voucher = ev
                if earliest_voucher and earliest_voucher > fy_end_str:
                    # Requested FY ends BEFORE the earliest synced voucher → no data
                    return {
                        "items": [],
                        "summary": {
                            "total_items": 0,
                            "total_opening_stock": 0,
                            "total_inward": 0,
                            "total_sales_qty": 0,
                            "total_closing_stock": 0,
                            "total_revenue": 0,
                            "fast_moving_count": 0,
                            "slow_moving_count": 0,
                            "dead_stock_count": 0,
                            "fy_days": 0,
                        },
                        "notices": [
                            f"FY {fy} was not synced from Tally (earliest synced voucher: "
                            f"{earliest_voucher}). Movement Analysis is empty for this FY."
                        ],
                        "fy_synced": False,
                        "earliest_voucher_date": earliest_voucher,
                    }

        inventory_items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        raw_sales_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)

        # Try to get purchase data (if synced)
        raw_purchase_vouchers = await db.purchase_vouchers.find(q, {"_id": 0}).to_list(10000)

        # Detect branch-like parties in purchases (non-sundry-creditor)
        purchase_branch_set = await _get_purchase_branch_set(ctx)
        sundry_creditor_purchases = _filter_branch_vouchers(raw_purchase_vouchers, purchase_branch_set)

        # ALL sales and ALL purchases for opening stock (must use full data for balance)
        all_sales_fy = filter_vouchers_by_fy(raw_sales_vouchers, fy)
        all_purchases_fy = filter_vouchers_by_fy(raw_purchase_vouchers, fy) if raw_purchase_vouchers else []

        # Compute unfiltered item totals for opening stock
        all_item_sales_qty = {}
        for voucher in all_sales_fy:
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                if item_name:
                    key = item_name.lower()
                    all_item_sales_qty[key] = all_item_sales_qty.get(key, 0) + qty

        # ALL purchases (including branch) for opening stock calculation
        all_item_purchase_qty = {}
        for voucher in all_purchases_fy:
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                if item_name:
                    key = item_name.lower()
                    all_item_purchase_qty[key] = all_item_purchase_qty.get(key, 0) + qty

        # Sundry creditor purchases only — for Inward display column
        sc_purchases_fy = filter_vouchers_by_fy(sundry_creditor_purchases, fy) if sundry_creditor_purchases else []

        # FILTERED sales for display (branch toggle affects sales only)
        filtered_sales = _filter_branch_vouchers(raw_sales_vouchers, branch_set)
        sales_vouchers = filter_vouchers_by_fy(filtered_sales, fy)
        # Inward display = sundry creditor purchases only
        purchase_vouchers = sc_purchases_fy

        # Calculate FY duration in days for rate calculations
        from datetime import date as date_type
        if fy:
            fy_start_str, fy_end_str = fy_to_date_range(fy)
            try:
                fy_start_parts = fy_start_str.split('-')
                fy_end_parts = fy_end_str.split('-')
                fy_start = date_type(int(fy_start_parts[0]), int(fy_start_parts[1]), int(fy_start_parts[2]))
                fy_end = date_type(int(fy_end_parts[0]), int(fy_end_parts[1]), int(fy_end_parts[2]))
                today = date_type.today()
                # If FY is still running, use today as end date
                effective_end = min(fy_end, today)
                fy_days = max((effective_end - fy_start).days, 1)
            except (ValueError, TypeError):
                fy_days = 365
        else:
            fy_days = 365

        # Aggregate item-wise sales: qty, revenue, txn count, first/last date
        item_sales = {}
        for voucher in sales_vouchers:
            vdate = voucher.get("voucher_date", "")
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                amount = safe_num(item.get("amount"))
                if item_name:
                    key = item_name.lower()
                    if key not in item_sales:
                        item_sales[key] = {"qty": 0, "revenue": 0, "txns": 0, "first_date": vdate, "last_date": vdate}
                    item_sales[key]["qty"] += qty
                    item_sales[key]["revenue"] += amount
                    item_sales[key]["txns"] += 1
                    if vdate < item_sales[key]["first_date"]:
                        item_sales[key]["first_date"] = vdate
                    if vdate > item_sales[key]["last_date"]:
                        item_sales[key]["last_date"] = vdate

        # Aggregate item-wise purchases (inward)
        item_purchases = {}
        for voucher in purchase_vouchers:
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                if item_name:
                    key = item_name.lower()
                    item_purchases[key] = item_purchases.get(key, 0) + qty

        movement_data = []
        seen_items = set()
        for item in inventory_items:
            item_name = item.get("item_name", "")
            closing_stock = safe_num(item.get("quantity"))
            cost_price = safe_num(item.get("price", item.get("rate", 0)))
            key = item_name.lower()
            seen_items.add(key)

            sales_info = item_sales.get(key, {"qty": 0, "revenue": 0, "txns": 0})
            sales_qty = sales_info["qty"]
            purchase_qty = item_purchases.get(key, 0)

            # Opening stock from UNFILTERED data: Closing + AllSales - AllPurchases
            total_sales_qty = all_item_sales_qty.get(key, 0)
            total_purchase_qty = all_item_purchase_qty.get(key, 0)
            opening_stock = max(closing_stock + total_sales_qty - total_purchase_qty, 0)

            # Inward from filtered purchases
            inward = purchase_qty

            # Movement Rate = (Sales / (Opening + Inward)) * 100
            # Represents what % of total available stock was sold
            available_stock = opening_stock + inward
            if available_stock > 0:
                movement_rate = round((sales_qty / available_stock) * 100, 1)
            elif sales_qty > 0:
                movement_rate = 100.0
            else:
                movement_rate = 0.0

            # Days to sell remaining stock at current rate
            daily_sales = sales_qty / fy_days if fy_days > 0 else 0
            if closing_stock > 0 and daily_sales > 0:
                days_to_sell = round(closing_stock / daily_sales, 1)
            elif closing_stock > 0 and sales_qty == 0:
                days_to_sell = 999  # Stock exists but no sales
            else:
                days_to_sell = 0  # No stock remaining

            # Monthly average sales
            fy_months = max(fy_days / 30, 1)
            monthly_avg = round(sales_qty / fy_months, 1) if sales_qty > 0 else 0

            # Classification based on sales frequency and movement
            if sales_qty == 0:
                classification = "non-moving"
            elif sales_info["txns"] >= fy_months * 2:  # Sells twice+ per month
                classification = "fast-moving"
            elif sales_info["txns"] >= fy_months * 0.5:  # Sells at least every 2 months
                classification = "moderate"
            elif sales_info["txns"] > 0:
                classification = "slow-moving"
            else:
                classification = "non-moving"

            movement_data.append({
                "item_name": item_name,
                "part_number": item.get("part_number", ""),
                "category": item.get("category", item.get("stock_group", "General")),
                "opening_stock": round(opening_stock, 1),
                "inward": round(inward, 1),
                "sales": round(sales_qty, 1),
                "closing_stock": round(closing_stock, 1),
                "movement_rate": movement_rate,
                "days_to_sell": min(days_to_sell, 999),
                "monthly_avg_sales": monthly_avg,
                "transactions": sales_info.get("txns", 0),
                "revenue": round(sales_info.get("revenue", 0), 2),
                "cost_price": round(cost_price, 2),
                "classification": classification,
            })

        # Items sold but not in inventory master
        for item_key, info in item_sales.items():
            if item_key not in seen_items and info["qty"] > 0:
                purchase_qty = item_purchases.get(item_key, 0)
                fy_months = max(fy_days / 30, 1)
                monthly_avg = round(info["qty"] / fy_months, 1)
                classification = "fast-moving" if info["txns"] >= fy_months * 2 else "moderate" if info["txns"] >= fy_months * 0.5 else "slow-moving"
                movement_data.append({
                    "item_name": item_key.title(),
                    "category": "General",
                    "opening_stock": 0,
                    "inward": round(purchase_qty, 1),
                    "sales": round(info["qty"], 1),
                    "closing_stock": 0,
                    "movement_rate": 100.0,
                    "days_to_sell": 0,
                    "monthly_avg_sales": monthly_avg,
                    "transactions": info["txns"],
                    "revenue": round(info["revenue"], 2),
                    "classification": classification,
                })

        movement_data.sort(key=lambda x: x["transactions"], reverse=True)

        return APIResponse(
            success=True,
            data={
                "movements": movement_data,
                "summary": {
                    "fast_moving": len([m for m in movement_data if m["classification"] == "fast-moving"]),
                    "moderate": len([m for m in movement_data if m["classification"] == "moderate"]),
                    "slow_moving": len([m for m in movement_data if m["classification"] == "slow-moving"]),
                    "non_moving": len([m for m in movement_data if m["classification"] == "non-moving"]),
                },
                "fy_days": fy_days,
            }
        )
    except Exception as e:
        logger.error(f"Error analyzing inventory movement: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/inventory/pivot-data")
async def get_pivot_data(request: Request, group_by: str = "category", metric: str = "value", company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        inventory_items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)

        pivot_data = {}
        for item in inventory_items:
            group_key = item.get(group_by, "Uncategorized")
            if group_key not in pivot_data:
                pivot_data[group_key] = {
                    "group": group_key,
                    "total_items": 0,
                    "total_quantity": 0,
                    "total_value": 0,
                    "items": []
                }
            pivot_data[group_key]["total_items"] += 1
            pivot_data[group_key]["total_quantity"] += safe_num(item.get("quantity"))
            pivot_data[group_key]["total_value"] += safe_num(item.get("quantity")) * safe_num(item.get("price"))
            pivot_data[group_key]["items"].append(item)

        pivot_list = list(pivot_data.values())

        if metric == "value":
            pivot_list.sort(key=lambda x: x["total_value"], reverse=True)
        elif metric == "quantity":
            pivot_list.sort(key=lambda x: x["total_quantity"], reverse=True)
        else:
            pivot_list.sort(key=lambda x: x["total_items"], reverse=True)

        return APIResponse(
            success=True,
            data={"pivot_table": pivot_list, "group_by": group_by, "metric": metric}
        )
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== BELOW COST SALES ====================

@router.get("/inventory/below-cost-sales")
async def get_below_cost_sales(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Find items where sales price < purchase cost price (negative margin)."""
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        inventory_items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)
        all_vouchers = _filter_branch_vouchers(all_vouchers, branch_set)
        sales_vouchers = filter_vouchers_by_fy(all_vouchers, fy)

        # Also get purchase vouchers for cost price
        purchase_vouchers_raw = await db.purchase_vouchers.find(q, {"_id": 0}).to_list(10000)
        purchase_vouchers = filter_vouchers_by_fy(purchase_vouchers_raw, fy) if purchase_vouchers_raw else []

        # Build cost price map from inventory (Tally's rate/price field) and purchases
        cost_map = {}
        for item in inventory_items:
            name = item.get("item_name", "").lower()
            price = safe_num(item.get("price", item.get("rate", 0)))
            if price > 0:
                cost_map[name] = price

        # Override with purchase price if available (more accurate for FY)
        purchase_price_map = {}
        for pv in purchase_vouchers:
            for item in pv.get("items", []):
                iname = item.get("item", "").strip().lower()
                rate = safe_num(item.get("rate", 0))
                qty = safe_num(item.get("quantity", 0))
                if iname and rate > 0:
                    if iname not in purchase_price_map:
                        purchase_price_map[iname] = {"total_cost": 0, "total_qty": 0}
                    purchase_price_map[iname]["total_cost"] += abs(rate * qty)
                    purchase_price_map[iname]["total_qty"] += abs(qty)

        for iname, pdata in purchase_price_map.items():
            if pdata["total_qty"] > 0:
                cost_map[iname] = pdata["total_cost"] / pdata["total_qty"]

        # Build sales price map
        sales_price_map = {}
        for sv in sales_vouchers:
            for item in sv.get("items", []):
                iname = item.get("item", "").strip().lower()
                rate = safe_num(item.get("rate", 0))
                qty = safe_num(item.get("quantity", 0))
                amt = safe_num(item.get("amount", 0))
                if iname and (rate > 0 or (qty > 0 and amt != 0)):
                    if iname not in sales_price_map:
                        sales_price_map[iname] = {"total_revenue": 0, "total_qty": 0, "txns": 0}
                    sales_price_map[iname]["total_revenue"] += abs(amt) if amt else abs(rate * qty)
                    sales_price_map[iname]["total_qty"] += abs(qty)
                    sales_price_map[iname]["txns"] += 1

        below_cost_items = []
        for iname, sdata in sales_price_map.items():
            cost = cost_map.get(iname, 0)
            if cost <= 0 or sdata["total_qty"] <= 0:
                continue

            avg_selling_price = sdata["total_revenue"] / sdata["total_qty"]
            margin = avg_selling_price - cost
            margin_pct = (margin / cost) * 100

            if margin < 0:
                # Find display name
                display_name = iname
                for inv in inventory_items:
                    if inv.get("item_name", "").lower() == iname:
                        display_name = inv["item_name"]
                        break

                below_cost_items.append({
                    "item_name": display_name,
                    "cost_price": round(cost, 2),
                    "avg_selling_price": round(avg_selling_price, 2),
                    "margin": round(margin, 2),
                    "margin_pct": round(margin_pct, 1),
                    "qty_sold": round(sdata["total_qty"], 1),
                    "total_revenue": round(sdata["total_revenue"], 2),
                    "total_loss": round(abs(margin) * sdata["total_qty"], 2),
                    "transactions": sdata["txns"],
                })

        below_cost_items.sort(key=lambda x: x["total_loss"], reverse=True)

        return APIResponse(
            success=True,
            data={
                "items": below_cost_items,
                "summary": {
                    "total_items": len(below_cost_items),
                    "total_loss": round(sum(i["total_loss"] for i in below_cost_items), 2),
                    "total_affected_revenue": round(sum(i["total_revenue"] for i in below_cost_items), 2),
                },
            }
        )
    except Exception as e:
        logger.error(f"Error analyzing below-cost sales: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== MOVEMENT ANALYSIS EXCEL EXPORT ====================

@router.get("/inventory/movement-export")
async def export_movement_analysis(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Export movement analysis to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        import io

        # Reuse the movement-analysis logic
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        inventory_items_raw = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        raw_sales = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)
        raw_purchases = await db.purchase_vouchers.find(q, {"_id": 0}).to_list(10000)

        # Filter purchases to sundry creditors only for inward display
        purchase_branch_set = await _get_purchase_branch_set(ctx)
        sc_purchases = _filter_branch_vouchers(raw_purchases, purchase_branch_set)

        # ALL sales + ALL purchases for opening stock (must balance)
        all_sales_fy = filter_vouchers_by_fy(raw_sales, fy)
        all_purchases_fy = filter_vouchers_by_fy(raw_purchases, fy) if raw_purchases else []
        all_item_sales_qty = {}
        for v in all_sales_fy:
            for item in v.get("items", []):
                n = item.get("item", "").strip()
                if n:
                    k = n.lower()
                    all_item_sales_qty[k] = all_item_sales_qty.get(k, 0) + safe_num(item.get("quantity"))
        all_item_purchase_qty = {}
        for v in all_purchases_fy:
            for item in v.get("items", []):
                n = item.get("item", "").strip()
                if n:
                    k = n.lower()
                    all_item_purchase_qty[k] = all_item_purchase_qty.get(k, 0) + safe_num(item.get("quantity"))

        # Filtered sales for display; inward uses sundry creditor purchases only
        sales_vouchers = filter_vouchers_by_fy(_filter_branch_vouchers(raw_sales, branch_set), fy)
        sc_purchases_fy = filter_vouchers_by_fy(sc_purchases, fy) if sc_purchases else []
        purchase_vouchers = sc_purchases_fy

        # Build item sales and purchases maps
        item_sales = {}
        for voucher in sales_vouchers:
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                if item_name:
                    key = item_name.lower()
                    item_sales[key] = item_sales.get(key, 0) + qty

        item_purchases = {}
        for voucher in purchase_vouchers:
            for item in voucher.get("items", []):
                item_name = item.get("item", "").strip()
                qty = safe_num(item.get("quantity"))
                if item_name:
                    key = item_name.lower()
                    item_purchases[key] = item_purchases.get(key, 0) + qty

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Movement Analysis FY {fy or 'All'}"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        ws.merge_cells('A1:J1')
        ws['A1'] = f"Movement Analysis | FY: {fy or 'All'}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([])

        headers = ["Item Name", "Category", "Opening Qty", "Inward (Purchases)", "Outward (Sales)", "Closing Qty", "Movement %", "Days to Sell", "Transactions", "Classification"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for inv_item in inventory_items_raw:
            name = inv_item.get("item_name", "")
            key = name.lower()
            closing = safe_num(inv_item.get("quantity"))
            s_qty = item_sales.get(key, 0)
            p_qty = item_purchases.get(key, 0)
            # Opening from unfiltered data
            total_s = all_item_sales_qty.get(key, 0)
            total_p = all_item_purchase_qty.get(key, 0)
            opening = max(closing + total_s - total_p, 0)
            available = opening + p_qty
            rate = round((s_qty / available * 100), 1) if available > 0 else (100.0 if s_qty > 0 else 0.0)
            daily = s_qty / 365 if s_qty > 0 else 0
            dts = round(closing / daily, 1) if closing > 0 and daily > 0 else (999 if closing > 0 else 0)
            cls_tag = "fast-moving" if s_qty > 0 and rate >= 50 else "moderate" if rate >= 20 else "slow-moving" if s_qty > 0 else "non-moving"

            ws.append([name, inv_item.get("category", ""), round(opening, 1), round(p_qty, 1), round(s_qty, 1), round(closing, 1), rate, dts if dts < 999 else "N/A", 0, cls_tag])

        for col_cells in ws.columns:
            valid_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not valid_cells:
                continue
            max_len = max((len(str(cell.value or "")) for cell in valid_cells), default=10)
            ws.column_dimensions[valid_cells[0].column_letter].width = min(max_len + 4, 35)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"movement_analysis_{fy or 'all'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Error exporting movement analysis: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== BELOW COST SALES EXCEL EXPORT ====================

@router.get("/inventory/below-cost-export")
async def export_below_cost_sales(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Export below-cost sales items to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from fastapi.responses import StreamingResponse
        import io

        # Get the below-cost data
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        inventory_items_list = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)
        all_vouchers = _filter_branch_vouchers(all_vouchers, branch_set)
        sales_vouchers = filter_vouchers_by_fy(all_vouchers, fy)
        purchase_vouchers_raw = await db.purchase_vouchers.find(q, {"_id": 0}).to_list(10000)
        purchase_vouchers = filter_vouchers_by_fy(purchase_vouchers_raw, fy) if purchase_vouchers_raw else []

        cost_map = {}
        for item in inventory_items_list:
            name = item.get("item_name", "").lower()
            price = safe_num(item.get("price", item.get("rate", 0)))
            if price > 0:
                cost_map[name] = price

        for pv in purchase_vouchers:
            for item in pv.get("items", []):
                iname = item.get("item", "").strip().lower()
                rate = safe_num(item.get("rate", 0))
                if iname and rate > 0:
                    cost_map[iname] = rate

        sales_map = {}
        for sv in sales_vouchers:
            for item in sv.get("items", []):
                iname = item.get("item", "").strip().lower()
                rate = safe_num(item.get("rate", 0))
                qty = safe_num(item.get("quantity", 0))
                amt = safe_num(item.get("amount", 0))
                if iname and qty > 0:
                    if iname not in sales_map:
                        sales_map[iname] = {"revenue": 0, "qty": 0}
                    sales_map[iname]["revenue"] += abs(amt) if amt else abs(rate * qty)
                    sales_map[iname]["qty"] += abs(qty)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Below Cost Sales FY {fy or 'All'}"

        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        ws.merge_cells('A1:H1')
        ws['A1'] = f"Below Cost Sales | FY: {fy or 'All'}"
        ws['A1'].font = Font(bold=True, size=12, color="EF4444")
        ws.append([])

        headers = ["Item Name", "Cost Price", "Avg Selling Price", "Margin", "Margin %", "Qty Sold", "Revenue", "Total Loss"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for iname, sdata in sorted(sales_map.items()):
            cost = cost_map.get(iname, 0)
            if cost <= 0 or sdata["qty"] <= 0:
                continue
            avg_sell = sdata["revenue"] / sdata["qty"]
            margin = avg_sell - cost
            if margin >= 0:
                continue
            ws.append([iname.title(), round(cost, 2), round(avg_sell, 2), round(margin, 2), round(margin / cost * 100, 1), round(sdata["qty"], 1), round(sdata["revenue"], 2), round(abs(margin) * sdata["qty"], 2)])

        for col_cells in ws.columns:
            valid_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not valid_cells:
                continue
            max_len = max((len(str(cell.value or "")) for cell in valid_cells), default=10)
            ws.column_dimensions[valid_cells[0].column_letter].width = min(max_len + 4, 35)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"below_cost_sales_{fy or 'all'}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"Error exporting below-cost sales: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== SALES FREQUENCY EXCEL/PDF EXPORT ====================

@router.get("/inventory/sales-frequency-export")
async def export_sales_frequency(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, fy: Optional[str] = None, format: str = "excel", company_id: Optional[str] = None):
    """Export sales frequency data to Excel or PDF."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
        import io

        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        branch_set = await _get_branch_set(request, ctx)
        all_vouchers = _filter_branch_vouchers(all_vouchers, branch_set)
        sales_vouchers = filter_vouchers_by_fy(all_vouchers, fy)

        if start_date or end_date:
            filtered = []
            for v in sales_vouchers:
                vd = v.get("voucher_date", "")
                if start_date and vd < start_date:
                    continue
                if end_date and vd > end_date:
                    continue
                filtered.append(v)
            sales_vouchers = filtered

        item_stats = {}
        for voucher in sales_vouchers:
            party = voucher.get("party_name", "Unknown")
            for item in voucher.get("items", []):
                item_name = item.get("item", "")
                qty = safe_num(item.get("quantity"))
                if item_name not in item_stats:
                    item_stats[item_name] = {"qty": 0, "txns": 0, "customers": set(), "revenue": 0}
                item_stats[item_name]["qty"] += qty
                item_stats[item_name]["txns"] += 1
                item_stats[item_name]["customers"].add(party)
                item_stats[item_name]["revenue"] += qty * safe_num(item.get("rate"))

        rows = sorted(item_stats.items(), key=lambda x: x[1]["txns"], reverse=True)

        if format == "pdf":
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet

            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
            styles = getSampleStyleSheet()
            elements = [Paragraph(f"Sales Frequency Report | FY: {fy or 'All'}", styles['Title']), Spacer(1, 12)]

            table_data = [["Item Name", "Transactions", "Total Qty", "Unique Customers", "Revenue", "Avg Qty/Txn"]]
            for name, s in rows:
                avg = round(s["qty"] / s["txns"], 1) if s["txns"] > 0 else 0
                table_data.append([name, s["txns"], round(s["qty"], 1), len(s["customers"]), f"Rs.{round(s['revenue'], 2):,.2f}", avg])

            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ]))
            elements.append(t)
            doc.build(elements)
            buf.seek(0)
            return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="sales_frequency_{fy or "all"}.pdf"'})

        # Default: Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Sales Frequency FY {fy or 'All'}"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Sales Frequency Report | FY: {fy or 'All'}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([])

        headers = ["Item Name", "Transactions", "Total Qty Sold", "Unique Customers", "Revenue", "Avg Qty/Txn"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for name, s in rows:
            avg = round(s["qty"] / s["txns"], 1) if s["txns"] > 0 else 0
            ws.append([name, s["txns"], round(s["qty"], 1), len(s["customers"]), round(s["revenue"], 2), avg])

        for col_cells in ws.columns:
            valid_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not valid_cells:
                continue
            max_len = max((len(str(cell.value or "")) for cell in valid_cells), default=10)
            ws.column_dimensions[valid_cells[0].column_letter].width = min(max_len + 4, 35)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="sales_frequency_{fy or "all"}.xlsx"'}
        )
    except Exception as e:
        logger.error(f"Error exporting sales frequency: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/inventory/set-reorder-level")
async def set_reorder_level(request: Request):
    """Set manual reorder level for an inventory item."""
    try:
        ctx = await get_tenant_context(request)
        body = await request.json()
        item_id = body.get("item_id", "")
        reorder_level = math.ceil(float(body.get("reorder_level", 0)))

        if not item_id:
            return APIResponse(success=False, error="Item ID required")

        q = _build_query(ctx, None, {"item_id": item_id})
        result = await db.inventory_items.update_one(q, {"$set": {"reorder_level": reorder_level}})
        if result.matched_count == 0:
            return APIResponse(success=False, error="Item not found")

        user = await get_current_user(request, db)
        await log_audit("reorder_level_set", user.get("username", "") if user else "",
                         tenant_id=ctx.get("tenant_id", "") if ctx else "",
                         company_id=ctx.get("company_id", "") if ctx else "",
                         target=item_id, details=f"Reorder level: {reorder_level}",
                         ip_address=get_client_ip(request))

        return APIResponse(success=True, message=f"Reorder level set to {reorder_level}")
    except Exception as e:
        logger.error(f"Error setting reorder level: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/inventory/auto-reorder-levels")
async def auto_set_reorder_levels(request: Request):
    """Auto-calculate reorder levels based on 2-month average sales consumption."""
    try:
        ctx = await get_tenant_context(request)
        body = await request.json()
        company_id = body.get("company_id")
        q = _build_query(ctx, company_id)

        items = await db.inventory_items.find(q, {"_id": 0}).to_list(10000)
        sales = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)

        if not sales:
            return APIResponse(success=False, error="No sales data to calculate reorder levels")

        # Calculate per-item monthly average sales qty
        from collections import defaultdict
        from datetime import datetime as dt
        item_sales_qty = defaultdict(float)
        dates = []
        for v in sales:
            vdate = v.get("voucher_date", v.get("date", ""))
            if vdate:
                try:
                    dates.append(dt.fromisoformat(vdate.replace("Z", "")))
                except Exception:
                    pass
            for vi in v.get("items", []):
                iname = (vi.get("item", "") or "").strip()
                qty = abs(safe_num(vi.get("quantity", 0)))
                if iname and qty > 0:
                    item_sales_qty[iname.lower()] += qty

        if not dates:
            return APIResponse(success=False, error="No valid sales dates found")

        min_date = min(dates)
        max_date = max(dates)
        months_span = max(1, (max_date - min_date).days / 30)

        updated = 0
        for item in items:
            iname = (item.get("item_name") or "").strip()
            total_sold = item_sales_qty.get(iname.lower(), 0)
            if total_sold > 0:
                monthly_avg = total_sold / months_span
                reorder_level = math.ceil(monthly_avg * 2)  # 2-month stock, rounded up
                item_q = _build_query(ctx, company_id, {"item_id": item["item_id"]})
                await db.inventory_items.update_one(item_q, {"$set": {"reorder_level": reorder_level}})
                updated += 1

        user = await get_current_user(request, db)
        await log_audit("auto_reorder_levels", user.get("username", "") if user else "",
                         tenant_id=ctx.get("tenant_id", "") if ctx else "", company_id=ctx.get("company_id", "") if ctx else "",
                         details=f"Updated {updated} items", ip_address=get_client_ip(request))

        return APIResponse(success=True, message=f"Reorder levels set for {updated} items (2-month stock)", data={"updated": updated})
    except Exception as e:
        logger.error(f"Error auto-setting reorder levels: {e}")
        return APIResponse(success=False, error=str(e))
