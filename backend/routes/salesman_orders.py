"""Salesman Order System — order placement, admin approval, dispatch integration."""
from fastapi import APIRouter, Request
from typing import Optional
from datetime import datetime, timezone
import logging
import uuid

from db import db
from models import APIResponse
from utils import safe_num, fy_to_date_range, get_current_fy, build_fuzzy_regex
from services.auth_service import get_current_user
from services.tenant_context import get_tenant_context

logger = logging.getLogger(__name__)
router = APIRouter()

ORDER_STATUSES = ["pending", "approved", "rejected", "billed", "hold"]


def _q(ctx, company_id=None):
    q = {}
    if ctx and ctx.get("tenant_id"):
        q["tenant_id"] = ctx["tenant_id"]
    cid = company_id or (ctx.get("company_id") if ctx else None)
    if cid:
        q["company_id"] = cid
    return q


# ═══════════════════════════════════════════════════════
# SALESMAN: MAPPED CUSTOMERS
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/my-customers")
async def get_my_customers(request: Request, company_id: Optional[str] = None):
    """Get customers mapped to the logged-in salesman."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        fy = get_current_fy()

        master = await db.salesman_master.find_one(
            {**q, "salesman_name": user.get("name", "")}, {"_id": 0})
        if not master:
            master = await db.salesman_master.find_one(
                {**q, "salesman_name": {"$regex": f"^{user.get('name', '')}$", "$options": "i"}}, {"_id": 0})

        if not master:
            return APIResponse(success=True, data={"customers": [], "message": "No salesman mapping found"})

        fy_customers = master.get("fy_customers", {}).get(fy, master.get("customers", []))
        customer_details = []
        for name in fy_customers:
            cust = await db.customers.find_one({**q, "customer_name": name}, {"_id": 0, "customer_name": 1, "phone": 1, "state": 1, "opening_balance": 1})
            customer_details.append({
                "customer_name": name,
                "phone": cust.get("phone", "") if cust else "",
                "state": cust.get("state", "") if cust else "",
            })

        return APIResponse(success=True, data={"customers": customer_details, "salesman_name": master.get("salesman_name")})
    except Exception as e:
        logger.error(f"Get my customers error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# PRODUCT CATALOG (from inventory)
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/catalog")
async def get_catalog(request: Request, search: Optional[str] = None, company_id: Optional[str] = None):
    """Get product catalog with real-time stock & Tally price.
    Search matches against item_name OR part_number (global search).
    Returns ALL items including zero-stock — salesman still needs the standard
    sale price to quote the customer.
    """
    try:
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        inv_q = {**q}
        if search:
            # Fuzzy match — ignores spaces & separators so "tvs 10" matches
            # "TVS-10", "TVS(10)", "TVS/10", etc.
            fuzzy = build_fuzzy_regex(search)
            if fuzzy:
                # v9.8.7 — alias matching: salesmen often know items by their
                # Tally LANGUAGENAME alias (customer's SKU, brand short-name, etc.)
                # Mongo's regex against an array field matches if ANY element matches.
                inv_q["$or"] = [
                    {"item_name": {"$regex": fuzzy, "$options": "i"}},
                    {"part_number": {"$regex": fuzzy, "$options": "i"}},
                    {"aliases": {"$regex": fuzzy, "$options": "i"}},
                ]
        items = await db.inventory_items.find(inv_q, {"_id": 0}).sort("item_name", 1).to_list(2000)

        # Last-sale-price fallback: when Tally master STANDARDPRICE is unset,
        # surface the most recent sale rate so salesmen can quote without
        # cleaning up STDPRICE for thousands of items in Tally master.
        # Tally master always wins when present.
        from routes.inventory import _last_sale_price_map
        try:
            lsp = await _last_sale_price_map(q)
        except Exception as e:
            logger.warning(f"catalog last-sale-price computation failed (non-fatal): {e}")
            lsp = {}

        catalog = []
        for it in items:
            std = safe_num(it.get("standard_price", 0))
            name = (it.get("item_name") or "").strip().lower()
            entry = lsp.get(name) or {}
            last_price = safe_num(entry.get("price", 0))
            last_date = entry.get("date", "")
            if std > 0:
                effective_price = std
                source = "tally_master"
            elif last_price > 0:
                effective_price = last_price
                source = "last_sale"
            else:
                effective_price = 0
                source = "unset"
            catalog.append({
                "item_name": it.get("item_name", ""),
                "item_id": it.get("item_id", ""),
                "part_number": it.get("part_number", "") or "",
                "aliases": it.get("aliases") or [],
                "stock_qty": safe_num(it.get("quantity", 0)),
                # `price` is the effective price the salesman quotes:
                # Tally master STANDARDPRICE if set, else last sale rate, else 0.
                # Cost (`it.price`) is NEVER used — that would torch margins.
                "price": effective_price,
                "standard_price": std,
                "last_sale_price": last_price,
                "last_sale_date": last_date,
                "sale_price_source": source,
                "unit": it.get("unit", ""),
                "stock_group": it.get("stock_group", ""),
            })
        return APIResponse(success=True, data={"items": catalog, "total": len(catalog)})
    except Exception as e:
        logger.error(f"Catalog error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# CUSTOMER ORDER SUGGESTIONS — repeat-order + cross-sell
# ═══════════════════════════════════════════════════════

def _last_n_months_cutoff(months: int) -> "datetime":
    """ISO date string for `months` months ago (UTC)."""
    from datetime import datetime, timezone, timedelta
    return datetime.now(timezone.utc) - timedelta(days=int(months) * 30)


def _voucher_party_norm(v) -> str:
    return (v.get("party_name") or "").strip().lower()


async def _build_inventory_lookup(q: dict) -> dict:
    """item_name (lowercased) → {price, stock_qty, unit, part_number, item_id, ...}.
    Joined once and reused by both suggestion endpoints."""
    from routes.inventory import _last_sale_price_map
    items = await db.inventory_items.find(q, {"_id": 0}).to_list(20000)
    try:
        lsp = await _last_sale_price_map(q)
    except Exception:
        lsp = {}
    out = {}
    for it in items:
        name = (it.get("item_name") or "").strip()
        if not name:
            continue
        std = safe_num(it.get("standard_price", 0))
        entry = lsp.get(name.lower()) or {}
        last_price = safe_num(entry.get("price", 0))
        effective = std if std > 0 else (last_price if last_price > 0 else 0)
        out[name.lower()] = {
            "item_name": name,
            "item_id": it.get("item_id", ""),
            "part_number": it.get("part_number", "") or "",
            "aliases": it.get("aliases") or [],
            "stock_qty": safe_num(it.get("quantity", 0)),
            "price": effective,
            "standard_price": std,
            "last_sale_price": last_price,
            "unit": it.get("unit", ""),
            "stock_group": it.get("stock_group", ""),
        }
    return out


@router.get("/salesman-orders/customer-history/{customer_name}")
async def customer_purchase_history(
    customer_name: str,
    request: Request,
    months: int = 10,
    company_id: Optional[str] = None,
):
    """Items this customer bought in the last `months` months.

    Aggregated per item: total qty, total revenue, # of distinct orders, last
    purchase date, average qty per order. Joined with current inventory so the
    salesman sees stock + standard price next to each suggestion. Sorted by
    most-recent purchase first (most relevant for repeat-order context).
    """
    try:
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        months = max(1, min(36, int(months or 10)))
        cutoff = _last_n_months_cutoff(months)
        cust_norm = (customer_name or "").strip().lower()
        if not cust_norm:
            return APIResponse(success=False, error="customer_name required")

        # Pull the customer's vouchers within the window (case-insensitive on party).
        # Mongo regex is anchored exact-match (after escape) — `^name$`.
        import re as _re
        vouchers = await db.sales_vouchers.find({
            **q,
            "party_name": {"$regex": f"^{_re.escape(customer_name.strip())}$",
                           "$options": "i"},
        }, {"_id": 0}).to_list(20000)

        # Filter by date — voucher_date is stored as ISO string OR datetime
        from datetime import datetime as _dt
        def _parse(v):
            d = v.get("voucher_date") or v.get("date")
            if isinstance(d, _dt):
                return d
            if isinstance(d, str):
                try:
                    return _dt.fromisoformat(d.replace("Z", "+00:00"))
                except Exception:
                    return None
            return None

        per_item: dict = {}
        for v in vouchers:
            vdate = _parse(v)
            if vdate is None:
                continue
            v_aware = vdate if vdate.tzinfo else vdate.replace(tzinfo=cutoff.tzinfo)
            if v_aware < cutoff:
                continue
            for it in (v.get("items") or []):
                # Tally syncs sometimes use 'item', sometimes 'item_name' —
                # accept both so suggestions work across both schema variants.
                iname = (it.get("item_name") or it.get("item") or "").strip()
                if not iname:
                    continue
                key = iname.lower()
                row = per_item.setdefault(key, {
                    "item_name": iname, "total_qty": 0.0,
                    "total_revenue": 0.0, "order_count": 0,
                    "last_date": None, "last_qty": 0, "last_price": 0,
                })
                qty = safe_num(it.get("quantity"))
                amt = safe_num(it.get("amount"))
                row["total_qty"] += qty
                row["total_revenue"] += amt
                row["order_count"] += 1
                if (row["last_date"] is None) or (vdate > row["last_date"]):
                    row["last_date"] = vdate
                    row["last_qty"] = qty
                    rate = safe_num(it.get("rate"))
                    row["last_price"] = rate if rate > 0 else (amt / qty if qty else 0)

        inv = await _build_inventory_lookup(q)
        out = []
        for key, row in per_item.items():
            inv_meta = inv.get(key, {})
            out.append({
                **row,
                "last_date": row["last_date"].isoformat() if row["last_date"] else None,
                "avg_qty_per_order": round(row["total_qty"] / row["order_count"], 2)
                                     if row["order_count"] else 0,
                "stock_qty": inv_meta.get("stock_qty", 0),
                "price": inv_meta.get("price", 0) or row.get("last_price", 0),
                "standard_price": inv_meta.get("standard_price", 0),
                "unit": inv_meta.get("unit", ""),
                "stock_group": inv_meta.get("stock_group", ""),
                "part_number": inv_meta.get("part_number", ""),
                "item_id": inv_meta.get("item_id", ""),
            })
        # Sort by recency (last purchased), break ties by total qty
        out.sort(key=lambda r: (r["last_date"] or "", r["total_qty"]), reverse=True)
        return APIResponse(success=True, data={
            "customer_name": customer_name,
            "months_window": months,
            "items": out,
            "total_items": len(out),
        })
    except Exception as e:
        logger.error(f"customer-history error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/salesman-orders/related-items/{customer_name}")
async def related_items_for_customer(
    customer_name: str,
    request: Request,
    months: int = 12,
    limit: int = 12,
    company_id: Optional[str] = None,
):
    """Cross-sell suggestions blending two signals (excluding items the
    customer has already bought):

      1. AFFINITY — items frequently co-purchased with the customer's past
         basket (basic market-basket co-occurrence across all vouchers).
      2. VELOCITY — fast-moving items by total quantity over the same window.

    Final score = 0.5 * affinity_norm + 0.5 * velocity_norm. Each suggestion
    carries a `signal` tag: 'affinity', 'fast_moving', or 'both' so the UI
    can label why it is being shown.
    """
    try:
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        months = max(1, min(36, int(months or 12)))
        limit = max(1, min(50, int(limit or 12)))
        cutoff = _last_n_months_cutoff(months)
        if not customer_name or not customer_name.strip():
            return APIResponse(success=False, error="customer_name required")

        # Single sweep over all vouchers in window — used for both signals.
        from datetime import datetime as _dt
        vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(80000)

        def _parse(v):
            d = v.get("voucher_date") or v.get("date")
            if isinstance(d, _dt):
                return d
            if isinstance(d, str):
                try:
                    return _dt.fromisoformat(d.replace("Z", "+00:00"))
                except Exception:
                    return None
            return None

        cust_norm = customer_name.strip().lower()
        # Items this customer has already bought (window-bound)
        bought: set = set()
        # Velocity counter: total qty per item across all vouchers
        velocity: dict = {}
        # Voucher-level item baskets — used for affinity scoring
        baskets: list = []

        for v in vouchers:
            vdate = _parse(v)
            if vdate is None:
                continue
            v_aware = vdate if vdate.tzinfo else vdate.replace(tzinfo=cutoff.tzinfo)
            if v_aware < cutoff:
                continue
            items_lc = []
            for it in (v.get("items") or []):
                iname = (it.get("item_name") or it.get("item") or "").strip()
                if not iname:
                    continue
                key = iname.lower()
                items_lc.append((key, iname, safe_num(it.get("quantity"))))
                vel = velocity.setdefault(key, {"item_name": iname, "qty": 0.0})
                vel["qty"] += safe_num(it.get("quantity"))
            if items_lc:
                baskets.append([k for k, _, _ in items_lc])
                if _voucher_party_norm(v) == cust_norm:
                    for k, _, _ in items_lc:
                        bought.add(k)

        # ── Affinity: count co-occurrences with `bought` items ───────────
        affinity: dict = {}
        for basket in baskets:
            basket_set = set(basket)
            overlap = basket_set & bought
            if not overlap:
                continue
            for other in basket_set - bought:
                affinity[other] = affinity.get(other, 0) + len(overlap)

        # ── Normalise & blend ────────────────────────────────────────────
        max_aff = max(affinity.values()) if affinity else 1
        max_vel = max((v["qty"] for v in velocity.values() if v["qty"]), default=1)

        candidates = set(velocity.keys()) - bought
        scored: list = []
        for key in candidates:
            aff = affinity.get(key, 0)
            vel_qty = velocity.get(key, {}).get("qty", 0)
            aff_n = aff / max_aff if max_aff else 0
            vel_n = vel_qty / max_vel if max_vel else 0
            score = 0.5 * aff_n + 0.5 * vel_n
            if score <= 0:
                continue
            if aff > 0 and vel_qty > 0 and aff_n > 0.15 and vel_n > 0.15:
                signal = "both"
            elif aff > 0:
                signal = "affinity"
            else:
                signal = "fast_moving"
            scored.append((key, score, aff, vel_qty, signal))

        scored.sort(key=lambda x: -x[1])
        scored = scored[:limit]

        inv = await _build_inventory_lookup(q)
        out = []
        for key, score, aff, vel_qty, signal in scored:
            meta = inv.get(key)
            if not meta:
                # Item present in vouchers but not in current inventory master —
                # skip rather than show an un-orderable suggestion.
                continue
            out.append({
                **meta,
                "co_occurrence": aff,
                "velocity_qty": round(vel_qty, 2),
                "score": round(score, 3),
                "signal": signal,  # 'affinity' | 'fast_moving' | 'both'
            })
        return APIResponse(success=True, data={
            "customer_name": customer_name,
            "months_window": months,
            "items": out,
            "total_items": len(out),
            "bought_count": len(bought),
        })
    except Exception as e:
        logger.error(f"related-items error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# SALESMAN DASHBOARD (own targets, achievement, customer breakdown)
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/my-stats")
async def get_my_stats(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Return logged-in salesman's target + achievement + customer breakdown
    for the requested FY (defaults to current FY).
    """
    try:
        from utils import get_current_fy, filter_vouchers_by_fy
        from datetime import date as _date
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        q = _q(ctx, company_id)
        target_fy = fy or get_current_fy()

        # Find salesman master record (case-insensitive)
        sname = user.get("name", "")
        master = await db.salesman_master.find_one({**q, "salesman_name": sname}, {"_id": 0})
        if not master:
            master = await db.salesman_master.find_one(
                {**q, "salesman_name": {"$regex": f"^{sname}$", "$options": "i"}}, {"_id": 0})
        if not master:
            return APIResponse(success=True, data={
                "salesman_name": sname, "fy": target_fy, "has_master": False,
                "monthly_target": 0, "quarterly_target": 0, "annual_target": 0,
                "expected_target": 0, "achieved_amount": 0, "achievement_percentage": 0,
                "customers": [], "items_sold": [],
            })

        # FY-scoped target + customer mapping
        fy_targets = master.get("fy_targets", {}).get(target_fy, {
            "monthly_target": master.get("monthly_target", 0),
            "quarterly_target": master.get("quarterly_target", 0),
        })
        monthly_target = safe_num(fy_targets.get("monthly_target"))
        quarterly_target = safe_num(fy_targets.get("quarterly_target"))
        annual_target = monthly_target * 12 if monthly_target else 0

        fy_customers = master.get("fy_customers", {}).get(target_fy, master.get("customers", []))
        fy_customers_lower = {c.lower().strip() for c in fy_customers}

        # Pull FY-scoped sales vouchers and filter to this salesman's customers
        all_v = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        v_fy = filter_vouchers_by_fy(all_v, target_fy)

        per_customer = {}  # name → {amount, count, items: {item: {qty, amount}}}
        per_item = {}      # item → {qty, revenue}
        total_achieved = 0.0

        for v in v_fy:
            party = (v.get("party_name") or "").strip()
            if party.lower() not in fy_customers_lower:
                continue
            amt = safe_num(v.get("total_amount"))
            total_achieved += amt
            row = per_customer.setdefault(party, {"customer_name": party, "amount": 0, "count": 0, "items": []})
            row["amount"] += amt
            row["count"] += 1
            for it in v.get("items", []) or []:
                iname = (it.get("item_name") or "").strip()
                if not iname:
                    continue
                qty = safe_num(it.get("quantity"))
                rev = safe_num(it.get("amount"))
                row["items"].append({"item_name": iname, "quantity": qty, "amount": rev})
                ag = per_item.setdefault(iname, {"item_name": iname, "quantity": 0, "revenue": 0})
                ag["quantity"] += qty
                ag["revenue"] += rev

        # Achievement % vs YTD-prorated target (for current FY) or annual (past FY)
        today = _date.today()
        cur_fy_year = today.year if today.month >= 4 else today.year - 1
        cur_fy = f"{cur_fy_year}-{str(cur_fy_year + 1)[-2:]}"
        if target_fy == cur_fy:
            months_elapsed = today.month - 3 if today.month >= 4 else today.month + 9
            months_elapsed = max(1, min(12, months_elapsed))
            expected_target = monthly_target * months_elapsed
        else:
            expected_target = annual_target

        achievement = 0
        if expected_target > 0:
            achievement = round(total_achieved / expected_target * 100, 1)

        return APIResponse(success=True, data={
            "salesman_name": master.get("salesman_name", sname),
            "fy": target_fy,
            "has_master": True,
            "monthly_target": monthly_target,
            "quarterly_target": quarterly_target if quarterly_target else monthly_target * 3,
            "annual_target": annual_target,
            "expected_target": round(expected_target, 2),
            "achieved_amount": round(total_achieved, 2),
            "achievement_percentage": achievement,
            "total_customers": len(per_customer),
            "customers": sorted(per_customer.values(), key=lambda x: -x["amount"]),
            "items_sold": sorted(per_item.values(), key=lambda x: -x["revenue"])[:50],
        })
    except Exception as e:
        logger.error(f"my-stats error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# ORDER CRUD
# ═══════════════════════════════════════════════════════

@router.post("/salesman-orders/orders")
async def create_order(request: Request):
    """Salesman creates a new order for a customer."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        body = await request.json()
        q = _q(ctx)

        customer_name = (body.get("customer_name") or "").strip()
        items = body.get("items", [])
        if not customer_name:
            return APIResponse(success=False, error="Customer name required")
        if not items or len(items) == 0:
            return APIResponse(success=False, error="At least one item required")

        order_items = []
        total = 0
        for it in items:
            qty = safe_num(it.get("quantity", 0))
            price = safe_num(it.get("price", 0))
            amt = round(qty * price, 2)
            total += amt
            order_items.append({
                "item_name": it.get("item_name", ""),
                "part_number": it.get("part_number", "") or "",
                "quantity": qty,
                "price": price,
                "amount": amt,
                "unit": it.get("unit", ""),
                "remark": (it.get("remark") or "").strip(),
            })

        now = datetime.now(timezone.utc).isoformat()
        order = {
            "order_id": f"SO-{uuid.uuid4().hex[:8].upper()}",
            "salesman": user.get("name", user.get("username", "")),
            "salesman_username": user.get("username", ""),
            "customer_name": customer_name,
            "items": order_items,
            "total_amount": round(total, 2),
            "status": "pending",
            "notes": (body.get("notes") or "").strip(),
            "admin_notes": "",
            "invoice_number": "",
            "created_at": now,
            "updated_at": now,
            "status_history": [{"status": "pending", "at": now, "by": user.get("username", "")}],
            **q
        }
        await db.salesman_orders.insert_one(order)
        return APIResponse(success=True, data={"order_id": order["order_id"]}, message="Order submitted for approval")
    except Exception as e:
        logger.error(f"Create order error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/salesman-orders/orders")
async def get_orders(request: Request, status: Optional[str] = None, search: Optional[str] = None,
                     date_from: Optional[str] = None, date_to: Optional[str] = None,
                     salesman: Optional[str] = None, company_id: Optional[str] = None,
                     page: int = 1, limit: int = 100):
    """Get orders. Salesman sees own, admin sees all."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)

        if user.get("role") == "salesman":
            q["salesman_username"] = user.get("username", "")

        if status:
            q["status"] = status
        if salesman:
            q["salesman"] = {"$regex": salesman, "$options": "i"}
        if search:
            fuzzy = build_fuzzy_regex(search)
            if fuzzy:
                q["$or"] = [
                    {"order_id": {"$regex": fuzzy, "$options": "i"}},
                    {"customer_name": {"$regex": fuzzy, "$options": "i"}},
                    {"invoice_number": {"$regex": fuzzy, "$options": "i"}},
                ]
        if date_from:
            q.setdefault("created_at", {})["$gte"] = f"{date_from}T00:00:00"
        if date_to:
            q.setdefault("created_at", {})["$lte"] = f"{date_to}T23:59:59"

        total = await db.salesman_orders.count_documents(q)
        skip = (page - 1) * limit
        orders = await db.salesman_orders.find(q, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

        return APIResponse(success=True, data={"orders": orders, "total": total, "page": page})
    except Exception as e:
        logger.error(f"Get orders error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/salesman-orders/orders/{order_id}")
async def get_order(order_id: str, request: Request):
    try:
        ctx = await get_tenant_context(request)
        order = await db.salesman_orders.find_one({**_q(ctx), "order_id": order_id}, {"_id": 0})
        if not order:
            return APIResponse(success=False, error="Order not found")
        return APIResponse(success=True, data=order)
    except Exception as e:
        return APIResponse(success=False, error=str(e))


@router.patch("/salesman-orders/orders/{order_id}")
async def update_order(order_id: str, request: Request):
    """Salesman edits order — only if status is 'pending'."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        body = await request.json()
        q = _q(ctx)

        order = await db.salesman_orders.find_one({**q, "order_id": order_id}, {"_id": 0})
        if not order:
            return APIResponse(success=False, error="Order not found")
        if order.get("status") != "pending":
            return APIResponse(success=False, error="Cannot edit order after approval. Only pending orders can be modified.")

        updates = {}
        if "items" in body:
            items = body["items"]
            total = 0
            order_items = []
            for it in items:
                qty = safe_num(it.get("quantity", 0))
                price = safe_num(it.get("price", 0))
                amt = round(qty * price, 2)
                total += amt
                order_items.append({
                    "item_name": it.get("item_name", ""),
                    "quantity": qty, "price": price, "amount": amt,
                    "unit": it.get("unit", ""), "remark": (it.get("remark") or "").strip(),
                })
            updates["items"] = order_items
            updates["total_amount"] = round(total, 2)
        if "notes" in body:
            updates["notes"] = body["notes"]
        if "customer_name" in body:
            updates["customer_name"] = body["customer_name"]

        if updates:
            updates["updated_at"] = datetime.now(timezone.utc).isoformat()
            await db.salesman_orders.update_one({**q, "order_id": order_id}, {"$set": updates})

        return APIResponse(success=True, message="Order updated")
    except Exception as e:
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# ADMIN: APPROVE / REJECT / HOLD / BILL
# ═══════════════════════════════════════════════════════

@router.patch("/salesman-orders/orders/{order_id}/status")
async def update_order_status(order_id: str, request: Request):
    """Admin changes order status: approve, reject, hold, billed."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        body = await request.json()
        new_status = (body.get("status") or "").lower()

        if new_status not in ORDER_STATUSES:
            return APIResponse(success=False, error=f"Invalid status")

        q = _q(ctx)
        order = await db.salesman_orders.find_one({**q, "order_id": order_id}, {"_id": 0})
        if not order:
            return APIResponse(success=False, error="Order not found")

        now = datetime.now(timezone.utc).isoformat()
        update_fields = {"status": new_status, "updated_at": now}

        if new_status == "rejected":
            reject_reason = (body.get("reject_reason") or "").strip()
            if not reject_reason:
                return APIResponse(success=False, error="Rejection reason is required")

        if new_status == "billed":
            invoice_number = (body.get("invoice_number") or "").strip()
            if not invoice_number:
                return APIResponse(success=False, error="Invoice number required for billing")
            update_fields["invoice_number"] = invoice_number

        if body.get("admin_notes"):
            update_fields["admin_notes"] = body["admin_notes"]

        entry = {"status": new_status, "at": now, "by": user.get("username", "")}
        if body.get("reject_reason"):
            entry["reason"] = body["reject_reason"]

        await db.salesman_orders.update_one({**q, "order_id": order_id},
            {"$set": update_fields, "$push": {"status_history": entry}})

        return APIResponse(success=True, message=f"Order {order_id} marked as {new_status}")
    except Exception as e:
        logger.error(f"Update order status error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# ADMIN: ORDER STATS
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/stats")
async def get_order_stats(request: Request, company_id: Optional[str] = None):
    """Get order counts by status for admin dashboard."""
    try:
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        pipeline = [
            {"$match": q},
            {"$group": {"_id": "$status", "count": {"$sum": 1}, "total": {"$sum": "$total_amount"}}}
        ]
        results = await db.salesman_orders.aggregate(pipeline).to_list(20)
        stats = {r["_id"]: {"count": r["count"], "total": round(r["total"], 2)} for r in results}
        return APIResponse(success=True, data={"stats": stats})
    except Exception as e:
        return APIResponse(success=False, error=str(e))



# ═══════════════════════════════════════════════════════
# PENDING BILLING — Approved orders not yet billed + item verification
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/pending-billing")
async def get_pending_billing(request: Request, company_id: Optional[str] = None):
    """Get approved orders pending billing, grouped by customer with item details.
    Also for billed orders, compare order items vs actual Tally invoice items."""
    try:
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)

        # Approved orders (pending billing)
        approved = await db.salesman_orders.find(
            {**q, "status": "approved"}, {"_id": 0}
        ).sort("created_at", 1).to_list(500)

        # Billed orders — for verification
        billed = await db.salesman_orders.find(
            {**q, "status": "billed", "invoice_number": {"$ne": ""}}, {"_id": 0}
        ).sort("created_at", -1).to_list(500)

        # Group approved by customer
        pending_by_customer = {}
        for o in approved:
            cust = o.get("customer_name", "Unknown")
            pending_by_customer.setdefault(cust, {"orders": [], "total_items": [], "total_amount": 0})
            pending_by_customer[cust]["orders"].append({
                "order_id": o.get("order_id"),
                "salesman": o.get("salesman", ""),
                "created_at": o.get("created_at", ""),
                "total_amount": o.get("total_amount", 0),
                "items": o.get("items", []),
            })
            pending_by_customer[cust]["total_amount"] += safe_num(o.get("total_amount", 0))
            for it in o.get("items", []):
                pending_by_customer[cust]["total_items"].append({
                    "item_name": it.get("item_name", ""),
                    "quantity": safe_num(it.get("quantity", 0)),
                    "price": safe_num(it.get("price", 0)),
                    "remark": it.get("remark", ""),
                    "order_id": o.get("order_id"),
                })

        pending_list = []
        for cust, data in sorted(pending_by_customer.items()):
            # Aggregate items by name
            item_agg = {}
            for it in data["total_items"]:
                name = it["item_name"]
                item_agg.setdefault(name, {"item_name": name, "total_qty": 0, "price": it["price"], "orders": []})
                item_agg[name]["total_qty"] += it["quantity"]
                item_agg[name]["orders"].append(it["order_id"])
            pending_list.append({
                "customer_name": cust,
                "order_count": len(data["orders"]),
                "total_amount": round(data["total_amount"], 2),
                "items": list(item_agg.values()),
                "orders": data["orders"],
            })

        # Billed verification — compare order items vs Tally invoice
        verified = []
        for o in billed[:50]:
            inv_num = o.get("invoice_number", "")
            if not inv_num:
                continue
            # Find matching Tally invoice
            tally_inv = await db.sales_vouchers.find_one(
                {**q, "$or": [{"voucher_id": inv_num}, {"reference_number": inv_num}]},
                {"_id": 0, "items": 1, "party_name": 1, "total_amount": 1, "voucher_id": 1}
            )
            order_items = {it.get("item_name", ""): safe_num(it.get("quantity", 0)) for it in o.get("items", [])}
            invoice_items = {}
            if tally_inv:
                for it in tally_inv.get("items", []):
                    name = it.get("item") or it.get("item_name") or ""
                    invoice_items[name] = safe_num(it.get("quantity", 0))

            # Compare
            all_items = set(list(order_items.keys()) + list(invoice_items.keys()))
            discrepancies = []
            for name in all_items:
                oq = order_items.get(name, 0)
                iq = invoice_items.get(name, 0)
                if abs(oq - iq) > 0.01:
                    discrepancies.append({"item_name": name, "ordered": oq, "billed": iq, "diff": round(iq - oq, 2)})

            verified.append({
                "order_id": o.get("order_id"),
                "invoice_number": inv_num,
                "customer_name": o.get("customer_name", ""),
                "tally_matched": tally_inv is not None,
                "match_status": "matched" if not discrepancies and tally_inv else ("discrepancy" if discrepancies else "not_synced"),
                "discrepancies": discrepancies,
            })

        return APIResponse(success=True, data={
            "pending": pending_list,
            "pending_count": len(approved),
            "verified": verified,
        })
    except Exception as e:
        logger.error(f"Pending billing error: {e}")
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# BEAT MANAGEMENT
# ═══════════════════════════════════════════════════════

@router.get("/salesman-orders/beats")
async def get_beats(request: Request, salesman: Optional[str] = None, company_id: Optional[str] = None):
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        q = _q(ctx, company_id)
        if user.get("role") == "salesman":
            q["salesman"] = user.get("name", "")
        elif salesman:
            q["salesman"] = salesman
        beats = await db.salesman_beats.find(q, {"_id": 0}).to_list(500)
        return APIResponse(success=True, data={"beats": beats})
    except Exception as e:
        return APIResponse(success=False, error=str(e))


@router.post("/salesman-orders/beats")
async def save_beats(request: Request):
    """Admin saves beat plan for a salesman."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        body = await request.json()
        q = _q(ctx)
        salesman = body.get("salesman", "")
        beats = body.get("beats", [])

        await db.salesman_beats.delete_many({**q, "salesman": salesman})
        now = datetime.now(timezone.utc).isoformat()
        for b in beats:
            await db.salesman_beats.insert_one({
                "beat_id": f"BT-{uuid.uuid4().hex[:6].upper()}",
                "salesman": salesman,
                "customer_name": b.get("customer_name", ""),
                "day_of_week": b.get("day_of_week", ""),
                "frequency": b.get("frequency", "weekly"),
                "created_at": now,
                **q
            })

        return APIResponse(success=True, message=f"{len(beats)} beats saved")
    except Exception as e:
        return APIResponse(success=False, error=str(e))


@router.post("/salesman-orders/beats/{beat_id}/visit")
async def mark_visited(beat_id: str, request: Request):
    """Salesman marks a beat customer as visited."""
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        now = datetime.now(timezone.utc).isoformat()
        await db.salesman_beats.update_one(
            {**_q(ctx), "beat_id": beat_id},
            {"$push": {"visits": {"at": now, "by": user.get("username", "")}}}
        )
        return APIResponse(success=True, message="Visit recorded")
    except Exception as e:
        return APIResponse(success=False, error=str(e))


# ═══════════════════════════════════════════════════════
# BEAT RUN TODAY — daily field-coverage tracking
# ═══════════════════════════════════════════════════════

DAY_NAMES = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
DAY_NAMES_FULL = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
# Acceptable variants for each weekday — we match beat plan day_of_week against
# either short or full names so the agent's existing data still works.
DAY_VARIANTS = {DAY_NAMES[i]: {DAY_NAMES[i], DAY_NAMES_FULL[i], DAY_NAMES[i].lower(), DAY_NAMES_FULL[i].lower()} for i in range(7)}


def _ist_today() -> str:
    """Current date in IST (YYYY-MM-DD) — used as the run_date key."""
    from datetime import datetime, timedelta
    return (datetime.utcnow() + timedelta(hours=5, minutes=30)).strftime("%Y-%m-%d")


def _ist_dow(date_str: str) -> str:
    """Day-of-week label for an IST date string."""
    from datetime import datetime
    return DAY_NAMES[datetime.strptime(date_str, "%Y-%m-%d").weekday()]


def _is_locked(run_date: str) -> bool:
    """A run is locked once the calendar day has passed (IST)."""
    return run_date < _ist_today()


async def _resolve_salesman_name(user, ctx) -> str:
    """Resolve current user → salesman master name (case-insensitive)."""
    if user.get("role") == "salesman":
        sname = user.get("name", "")
        master = await db.salesman_master.find_one(
            {**_q(ctx), "salesman_name": {"$regex": f"^{sname}$", "$options": "i"}},
            {"_id": 0, "salesman_name": 1},
        )
        return (master or {}).get("salesman_name", sname)
    return user.get("name", "")


@router.get("/salesman-orders/beat-run/today")
async def get_beat_run_today(request: Request, salesman: Optional[str] = None, run_date: Optional[str] = None, company_id: Optional[str] = None):
    """Return today's (or a specific date's) beat run for the logged-in salesman
    (or, for admin, for the requested `salesman`). If no run exists yet, build
    one from the salesman's beat plan filtered by today's day-of-week.

    Response shape:
      { salesman, run_date, day_of_week, locked,
        planned: [{customer_name, beat_id, frequency, visited_at, notes}],
        unplanned: [{visit_id, customer_name, details, added_at}] }
    """
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        q = _q(ctx, company_id)

        rd = run_date or _ist_today()
        # Validate date format (YYYY-MM-DD)
        try:
            from datetime import datetime as _dt
            _dt.strptime(rd, "%Y-%m-%d")
        except Exception:
            return APIResponse(success=False, error="run_date must be YYYY-MM-DD")

        # Determine target salesman
        if user.get("role") in ("admin", "super_admin"):
            target = salesman or await _resolve_salesman_name(user, ctx)
        else:
            target = await _resolve_salesman_name(user, ctx)
        if not target:
            return APIResponse(success=False, error="No salesman context")

        dow = _ist_dow(rd)
        locked = _is_locked(rd)

        # Find existing run
        run = await db.beat_runs.find_one({**q, "salesman": target, "run_date": rd}, {"_id": 0})

        if not run:
            # Build from beat plan (read-only construction — written on first check-in).
            # Match either short ('Mon') or full ('Monday') day_of_week values.
            day_match = list(DAY_VARIANTS.get(dow, {dow}))
            plan = await db.salesman_beats.find(
                {**q, "salesman": target, "day_of_week": {"$in": day_match}}, {"_id": 0}
            ).to_list(200)
            run = {
                "salesman": target, "run_date": rd, "day_of_week": dow,
                "planned": [{
                    "customer_name": b.get("customer_name", ""),
                    "beat_id": b.get("beat_id", ""),
                    "frequency": b.get("frequency", "weekly"),
                    "visited_at": None, "notes": "",
                } for b in plan],
                "unplanned": [],
                "created_at": None,
            }

        return APIResponse(success=True, data={
            "salesman": run["salesman"], "run_date": rd, "day_of_week": dow,
            "locked": locked,
            "planned": run.get("planned", []),
            "unplanned": run.get("unplanned", []),
            "created_at": run.get("created_at"),
        })
    except Exception as e:
        logger.error(f"beat-run/today error: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/salesman-orders/beat-run/check-in")
async def beat_run_check_in(request: Request):
    """Toggle visited status for a planned customer in TODAY's run only.
    Body: { customer_name: str, visited: bool, notes?: str, company_id?: str }
    """
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        body = await request.json()
        customer_name = (body.get("customer_name") or "").strip()
        visited = bool(body.get("visited"))
        notes = (body.get("notes") or "").strip()
        if not customer_name:
            return APIResponse(success=False, error="customer_name is required")

        q = _q(ctx, body.get("company_id"))
        rd = _ist_today()  # check-ins ALWAYS apply to today (server-enforced)
        target = await _resolve_salesman_name(user, ctx)
        # Admin can check in on behalf of a salesman (rare — for field auditing)
        if user.get("role") in ("admin", "super_admin") and body.get("salesman"):
            target = body.get("salesman")
        dow = _ist_dow(rd)
        now_iso = datetime.now(timezone.utc).isoformat()

        # Ensure run document exists (auto-create from plan if missing)
        existing = await db.beat_runs.find_one({**q, "salesman": target, "run_date": rd}, {"_id": 0})
        if not existing:
            day_match = list(DAY_VARIANTS.get(dow, {dow}))
            plan = await db.salesman_beats.find(
                {**q, "salesman": target, "day_of_week": {"$in": day_match}}, {"_id": 0}
            ).to_list(200)
            existing = {
                **q, "salesman": target, "run_date": rd, "day_of_week": dow,
                "planned": [{
                    "customer_name": b.get("customer_name", ""),
                    "beat_id": b.get("beat_id", ""),
                    "frequency": b.get("frequency", "weekly"),
                    "visited_at": None, "notes": "",
                } for b in plan],
                "unplanned": [],
                "created_at": now_iso,
                "updated_at": now_iso,
            }
            await db.beat_runs.insert_one({**existing})

        # Update the matching planned entry, or add it if not present (e.g., off-day visit)
        planned = existing.get("planned", [])
        found = False
        for p in planned:
            if (p.get("customer_name") or "").strip().lower() == customer_name.lower():
                p["visited_at"] = now_iso if visited else None
                if notes:
                    p["notes"] = notes
                found = True
                break
        if not found:
            planned.append({
                "customer_name": customer_name, "beat_id": "",
                "frequency": "ad-hoc",
                "visited_at": now_iso if visited else None,
                "notes": notes,
            })

        await db.beat_runs.update_one(
            {**q, "salesman": target, "run_date": rd},
            {"$set": {"planned": planned, "updated_at": now_iso}},
        )
        return APIResponse(success=True, message="Check-in saved")
    except Exception as e:
        logger.error(f"beat-run/check-in error: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/salesman-orders/beat-run/add-unplanned")
async def beat_run_add_unplanned(request: Request):
    """Add an unplanned visit (new prospect) to TODAY's run.
    Body: { customer_name: str, details?: str, company_id?: str }
    Marked with `is_new: true`. No CRM impact until the customer appears in Tally.
    """
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        body = await request.json()
        cname = (body.get("customer_name") or "").strip()
        details = (body.get("details") or "").strip()
        if not cname:
            return APIResponse(success=False, error="customer_name is required")

        q = _q(ctx, body.get("company_id"))
        rd = _ist_today()  # always today — past dates locked
        target = await _resolve_salesman_name(user, ctx)
        dow = _ist_dow(rd)
        now_iso = datetime.now(timezone.utc).isoformat()
        new_visit = {
            "visit_id": f"UV-{uuid.uuid4().hex[:8].upper()}",
            "customer_name": cname,
            "details": details,
            "is_new": True,
            "added_at": now_iso,
        }

        existing = await db.beat_runs.find_one({**q, "salesman": target, "run_date": rd}, {"_id": 0})
        if not existing:
            day_match = list(DAY_VARIANTS.get(dow, {dow}))
            plan = await db.salesman_beats.find(
                {**q, "salesman": target, "day_of_week": {"$in": day_match}}, {"_id": 0}
            ).to_list(200)
            await db.beat_runs.insert_one({
                **q, "salesman": target, "run_date": rd, "day_of_week": dow,
                "planned": [{
                    "customer_name": b.get("customer_name", ""),
                    "beat_id": b.get("beat_id", ""),
                    "frequency": b.get("frequency", "weekly"),
                    "visited_at": None, "notes": "",
                } for b in plan],
                "unplanned": [new_visit],
                "created_at": now_iso, "updated_at": now_iso,
            })
        else:
            await db.beat_runs.update_one(
                {**q, "salesman": target, "run_date": rd},
                {"$push": {"unplanned": new_visit}, "$set": {"updated_at": now_iso}},
            )
        return APIResponse(success=True, data={"visit": new_visit}, message="Unplanned visit added")
    except Exception as e:
        logger.error(f"beat-run/add-unplanned error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/salesman-orders/beat-run/history")
async def beat_run_history(request: Request,
                            salesman: Optional[str] = None,
                            from_date: Optional[str] = None,
                            to_date: Optional[str] = None,
                            company_id: Optional[str] = None,
                            limit: int = 60):
    """Return past beat runs.
    - Salesman: only their own runs (server-enforced).
    - Admin/super_admin: any salesman (or all if `salesman` not given).
    Optional date filters use YYYY-MM-DD (IST). Default: last 60 runs.
    """
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        q = _q(ctx, company_id)

        if user.get("role") not in ("admin", "super_admin"):
            q["salesman"] = await _resolve_salesman_name(user, ctx)
        elif salesman:
            q["salesman"] = salesman

        if from_date or to_date:
            date_q = {}
            if from_date:
                date_q["$gte"] = from_date
            if to_date:
                date_q["$lte"] = to_date
            q["run_date"] = date_q

        runs = await db.beat_runs.find(q, {"_id": 0}).sort("run_date", -1).limit(min(limit, 365)).to_list(min(limit, 365))
        # Compute summary stats
        for r in runs:
            r["locked"] = _is_locked(r.get("run_date", ""))
            planned = r.get("planned", [])
            r["planned_count"] = len(planned)
            r["visited_count"] = sum(1 for p in planned if p.get("visited_at"))
            r["unplanned_count"] = len(r.get("unplanned", []))
        return APIResponse(success=True, data={"runs": runs, "count": len(runs)})
    except Exception as e:
        logger.error(f"beat-run/history error: {e}")
        return APIResponse(success=False, error=str(e))



# ═══════════════════════════════════════════════════════
# Beat Run — Monthly Report (admin only)
# ═══════════════════════════════════════════════════════

def _month_window(month: str):
    """'2026-05' → ('2026-05-01', '2026-06-01'). Raises ValueError on bad input."""
    from datetime import datetime as _dt
    y, m = month.split("-")
    y, m = int(y), int(m)
    if not (1 <= m <= 12):
        raise ValueError("month out of range")
    start = f"{y:04d}-{m:02d}-01"
    if m == 12:
        end = f"{y+1:04d}-01-01"
    else:
        end = f"{y:04d}-{m+1:02d}-01"
    # Touch _dt to validate
    _dt.strptime(start, "%Y-%m-%d")
    return start, end


def _prev_month(month: str) -> str:
    y, m = month.split("-")
    y, m = int(y), int(m)
    if m == 1:
        return f"{y-1:04d}-12"
    return f"{y:04d}-{m-1:02d}"


def _summarize_runs(runs: list) -> dict:
    """Aggregate a list of beat_runs docs → totals (planned, visited, unplanned, coverage_pct)."""
    planned = sum(len(r.get("planned", []) or []) for r in runs)
    visited = sum(
        sum(1 for p in (r.get("planned") or []) if p.get("visited_at"))
        for r in runs
    )
    unplanned = sum(len(r.get("unplanned", []) or []) for r in runs)
    coverage_pct = round((visited / planned * 100), 1) if planned else 0.0
    return {
        "planned": planned,
        "visited": visited,
        "unplanned": unplanned,
        "coverage_pct": coverage_pct,
        "run_days": len(runs),
    }


@router.get("/salesman-orders/beat-run/monthly-report")
async def beat_run_monthly_report(
    request: Request,
    month: Optional[str] = None,
    salesman: Optional[str] = None,
    company_id: Optional[str] = None,
    trend_months: int = 6,
):
    """Monthly aggregated report (admin/super_admin only).
    Returns totals, per-salesman, per-customer visit-frequency, daily breakdown, and
    a last-N-month coverage trend for the sparkline.

    Query: month=YYYY-MM (default current IST month), salesman (optional filter),
    trend_months (1..24, default 6).
    """
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        if user.get("role") not in ("admin", "super_admin"):
            return APIResponse(success=False, error="Admin access required")

        month = month or _ist_today()[:7]
        try:
            start, end = _month_window(month)
        except (ValueError, IndexError):
            return APIResponse(success=False, error="month must be YYYY-MM")

        trend_months = max(1, min(int(trend_months or 6), 24))

        q = _q(ctx, company_id)
        q["run_date"] = {"$gte": start, "$lt": end}
        if salesman:
            q["salesman"] = salesman

        runs = await db.beat_runs.find(q, {"_id": 0}).to_list(20000)

        # ── Top-level summary
        summary = _summarize_runs(runs)
        summary["salesmen_count"] = len({r.get("salesman", "") for r in runs if r.get("salesman")})
        summary["month"] = month

        # ── Per-salesman roll-up
        per_sm_map: dict = {}
        for r in runs:
            name = r.get("salesman", "") or "—"
            per_sm_map.setdefault(name, []).append(r)
        per_salesman = []
        for name, srunsx in per_sm_map.items():
            row = {"salesman": name, **_summarize_runs(srunsx)}
            per_salesman.append(row)
        per_salesman.sort(key=lambda x: (-x["coverage_pct"], -x["visited"]))

        # ── Per-customer visit frequency (visited customers + unplanned visits)
        per_cust: dict = {}
        for r in runs:
            sm = r.get("salesman", "") or "—"
            for p in (r.get("planned") or []):
                if not p.get("visited_at"):
                    continue
                name = (p.get("customer_name") or "").strip()
                if not name:
                    continue
                e = per_cust.setdefault(name, {
                    "customer_name": name, "visit_count": 0,
                    "last_visit_date": "", "salesmen": set(), "unplanned": False,
                })
                e["visit_count"] += 1
                e["salesmen"].add(sm)
                vd = (p.get("visited_at") or "")[:10]
                if vd > e["last_visit_date"]:
                    e["last_visit_date"] = vd
            for u in (r.get("unplanned") or []):
                name = (u.get("customer_name") or "").strip()
                if not name:
                    continue
                e = per_cust.setdefault(name, {
                    "customer_name": name, "visit_count": 0,
                    "last_visit_date": "", "salesmen": set(), "unplanned": True,
                })
                e["visit_count"] += 1
                e["unplanned"] = True
                e["salesmen"].add(sm)
                ad = (u.get("added_at") or "")[:10]
                if ad > e["last_visit_date"]:
                    e["last_visit_date"] = ad
        per_customer = [
            {**v, "salesmen": sorted(v["salesmen"])}
            for v in per_cust.values()
        ]
        per_customer.sort(key=lambda x: (-x["visit_count"], x["customer_name"].lower()))

        # ── Daily breakdown
        daily_map: dict = {}
        for r in runs:
            d = r.get("run_date", "")
            if not d:
                continue
            daily_map.setdefault(d, []).append(r)
        daily_breakdown = []
        for d in sorted(daily_map.keys()):
            row = {"date": d, "day_of_week": _ist_dow(d), **_summarize_runs(daily_map[d])}
            daily_breakdown.append(row)

        # ── M-o-M trend (coverage % per month, oldest → newest)
        trend = []
        m_iter = month
        # Collect oldest-first
        months_in_order = []
        for _ in range(trend_months):
            months_in_order.append(m_iter)
            m_iter = _prev_month(m_iter)
        months_in_order.reverse()

        for m in months_in_order:
            try:
                ms, me = _month_window(m)
            except ValueError:
                continue
            tq = _q(ctx, company_id)
            tq["run_date"] = {"$gte": ms, "$lt": me}
            if salesman:
                tq["salesman"] = salesman
            mruns = await db.beat_runs.find(tq, {"_id": 0, "planned": 1, "unplanned": 1}).to_list(20000)
            tsum = _summarize_runs(mruns)
            trend.append({
                "month": m,
                "coverage_pct": tsum["coverage_pct"],
                "planned": tsum["planned"],
                "visited": tsum["visited"],
                "unplanned": tsum["unplanned"],
            })

        return APIResponse(success=True, data={
            "summary": summary,
            "per_salesman": per_salesman,
            "per_customer": per_customer,
            "daily_breakdown": daily_breakdown,
            "trend": trend,
        })
    except Exception as e:
        logger.error(f"beat-run/monthly-report error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/salesman-orders/beat-run/monthly-report/export")
async def beat_run_monthly_report_export(
    request: Request,
    month: Optional[str] = None,
    salesman: Optional[str] = None,
    company_id: Optional[str] = None,
    format: str = "excel",
):
    """Excel (multi-sheet) or CSV export of the monthly beat-run report.
    format=excel → 4-sheet workbook (summary, per_salesman, per_customer, raw_runs).
    format=csv   → flat raw runs (one row per planned visit + one row per unplanned).
    """
    from fastapi.responses import StreamingResponse
    import io
    try:
        ctx = await get_tenant_context(request)
        user = await get_current_user(request, db)
        if not user:
            return APIResponse(success=False, error="Authentication required")
        if user.get("role") not in ("admin", "super_admin"):
            return APIResponse(success=False, error="Admin access required")

        month = month or _ist_today()[:7]
        try:
            start, end = _month_window(month)
        except (ValueError, IndexError):
            return APIResponse(success=False, error="month must be YYYY-MM")

        q = _q(ctx, company_id)
        q["run_date"] = {"$gte": start, "$lt": end}
        if salesman:
            q["salesman"] = salesman
        runs = await db.beat_runs.find(q, {"_id": 0}).sort("run_date", 1).to_list(20000)

        # Build flat raw rows (one per planned + one per unplanned visit)
        flat: list = []
        for r in runs:
            sm = r.get("salesman", "")
            d = r.get("run_date", "")
            dow = r.get("day_of_week") or _ist_dow(d)
            for p in (r.get("planned") or []):
                flat.append({
                    "Date": d,
                    "Day": dow,
                    "Salesman": sm,
                    "Customer": p.get("customer_name", ""),
                    "Type": "Planned",
                    "Visited": "Yes" if p.get("visited_at") else "No",
                    "Visited At (IST)": (p.get("visited_at") or "")[:19].replace("T", " "),
                    "Notes": p.get("notes", ""),
                    "Frequency": p.get("frequency", ""),
                })
            for u in (r.get("unplanned") or []):
                flat.append({
                    "Date": d,
                    "Day": dow,
                    "Salesman": sm,
                    "Customer": u.get("customer_name", ""),
                    "Type": "Unplanned",
                    "Visited": "Yes",
                    "Visited At (IST)": (u.get("added_at") or "")[:19].replace("T", " "),
                    "Notes": u.get("details", ""),
                    "Frequency": "—",
                })

        suffix = f"-{salesman}" if salesman else ""
        fname_stem = f"flowra-beat-run-{month}{suffix}"

        if format.lower() == "csv":
            import csv as _csv
            buf = io.StringIO()
            if flat:
                writer = _csv.DictWriter(buf, fieldnames=list(flat[0].keys()))
                writer.writeheader()
                writer.writerows(flat)
            else:
                buf.write("No beat runs in this month.\n")
            data_bytes = buf.getvalue().encode("utf-8")
            return StreamingResponse(
                io.BytesIO(data_bytes),
                media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename={fname_stem}.csv"},
            )

        # Excel — multi-sheet
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        wb = Workbook()
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        stripe_fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
        title_font = Font(bold=True, color="0F1B4C", size=14)

        def _write_sheet(ws, title, headers, rows):
            ws.cell(row=1, column=1, value=title).font = title_font
            ws.cell(row=2, column=1, value=f"Month: {month}{(' · ' + salesman) if salesman else ''}").font = Font(italic=True, color="64748B", size=10)
            for ci, h in enumerate(headers, start=1):
                c = ws.cell(row=4, column=ci, value=h)
                c.fill = header_fill
                c.font = header_font
                c.alignment = Alignment(horizontal="center")
            for ri, row in enumerate(rows, start=5):
                for ci, h in enumerate(headers, start=1):
                    val = row.get(h, "")
                    c = ws.cell(row=ri, column=ci, value=val)
                    if isinstance(val, (int, float)):
                        c.alignment = Alignment(horizontal="right")
                    if ri % 2 == 0:
                        c.fill = stripe_fill
            for col_cells in ws.columns:
                try:
                    length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
                    ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)
                except ValueError:
                    pass

        # Recompute the summary breakdowns inline (avoid re-querying)
        summary = _summarize_runs(runs)
        summary["salesmen_count"] = len({r.get("salesman", "") for r in runs if r.get("salesman")})

        per_sm_map: dict = {}
        for r in runs:
            per_sm_map.setdefault(r.get("salesman", "") or "—", []).append(r)
        per_salesman_rows = []
        for name, srunsx in per_sm_map.items():
            s = _summarize_runs(srunsx)
            per_salesman_rows.append({
                "Salesman": name,
                "Run Days": s["run_days"],
                "Planned": s["planned"],
                "Visited": s["visited"],
                "Unplanned": s["unplanned"],
                "Coverage %": s["coverage_pct"],
            })
        per_salesman_rows.sort(key=lambda x: (-x["Coverage %"], -x["Visited"]))

        per_cust: dict = {}
        for r in runs:
            sm = r.get("salesman", "") or "—"
            for p in (r.get("planned") or []):
                if not p.get("visited_at"):
                    continue
                name = (p.get("customer_name") or "").strip()
                if not name:
                    continue
                e = per_cust.setdefault(name, {
                    "Customer": name, "Visits": 0, "Last Visit": "",
                    "Salesmen": set(), "Includes Unplanned": "No",
                })
                e["Visits"] += 1
                e["Salesmen"].add(sm)
                vd = (p.get("visited_at") or "")[:10]
                if vd > e["Last Visit"]:
                    e["Last Visit"] = vd
            for u in (r.get("unplanned") or []):
                name = (u.get("customer_name") or "").strip()
                if not name:
                    continue
                e = per_cust.setdefault(name, {
                    "Customer": name, "Visits": 0, "Last Visit": "",
                    "Salesmen": set(), "Includes Unplanned": "No",
                })
                e["Visits"] += 1
                e["Includes Unplanned"] = "Yes"
                e["Salesmen"].add(sm)
                ad = (u.get("added_at") or "")[:10]
                if ad > e["Last Visit"]:
                    e["Last Visit"] = ad
        per_customer_rows = [
            {**v, "Salesmen": ", ".join(sorted(v["Salesmen"]))}
            for v in per_cust.values()
        ]
        per_customer_rows.sort(key=lambda x: (-x["Visits"], x["Customer"].lower()))

        # Sheet 1 — Summary
        ws1 = wb.active
        ws1.title = "Summary"
        _write_sheet(ws1, "Beat Run — Monthly Summary", ["Metric", "Value"], [
            {"Metric": "Month", "Value": month},
            {"Metric": "Salesman filter", "Value": salesman or "(All)"},
            {"Metric": "Run Days", "Value": summary["run_days"]},
            {"Metric": "Distinct Salesmen", "Value": summary["salesmen_count"]},
            {"Metric": "Planned Visits", "Value": summary["planned"]},
            {"Metric": "Visited", "Value": summary["visited"]},
            {"Metric": "Unplanned Visits", "Value": summary["unplanned"]},
            {"Metric": "Coverage %", "Value": summary["coverage_pct"]},
        ])

        ws2 = wb.create_sheet("By Salesman")
        _write_sheet(ws2, "Per-Salesman Roll-up",
                     ["Salesman", "Run Days", "Planned", "Visited", "Unplanned", "Coverage %"],
                     per_salesman_rows)

        ws3 = wb.create_sheet("By Customer")
        _write_sheet(ws3, "Per-Customer Visit Frequency",
                     ["Customer", "Visits", "Last Visit", "Salesmen", "Includes Unplanned"],
                     per_customer_rows)

        ws4 = wb.create_sheet("Raw Runs")
        if flat:
            _write_sheet(ws4, "Raw Visits", list(flat[0].keys()), flat)
        else:
            ws4.cell(row=1, column=1, value="No beat runs in this month.").font = title_font

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return StreamingResponse(
            out,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={fname_stem}.xlsx"},
        )
    except Exception as e:
        logger.error(f"beat-run/monthly-report/export error: {e}")
        return APIResponse(success=False, error=str(e))
