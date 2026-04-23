from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
from collections import defaultdict
import uuid
import logging
import io

from db import db
from models import APIResponse
from utils import safe_num, filter_vouchers_by_fy, get_current_fy, get_previous_fy, is_fy_completed, fy_to_date_range
from services.tenant_context import get_tenant_context

logger = logging.getLogger(__name__)
router = APIRouter()


def _build_query(ctx, company_id=None, extra=None):
    q = {}
    if ctx and ctx.get("tenant_id"):
        q["tenant_id"] = ctx["tenant_id"]
    cid = company_id or (ctx.get("company_id") if ctx else None)
    if cid:
        q["company_id"] = cid
    if extra:
        q.update(extra)
    return q


def _get_fy_targets(master, fy):
    """Get targets for a specific FY from the master record."""
    fy_targets = master.get("fy_targets", {})
    if fy and fy in fy_targets:
        return fy_targets[fy]
    # Fallback to legacy flat fields
    return {
        "monthly_target": master.get("monthly_target", 0),
        "quarterly_target": master.get("quarterly_target", 0),
    }


def _get_fy_customers(master, fy):
    """Get customer mapping for a specific FY.
    Falls back through: exact FY -> previous FYs (closest) -> legacy customers field.
    For completed FYs, the mapping is frozen and must not change."""
    fy_customers = master.get("fy_customers", {})
    if fy and fy in fy_customers:
        return fy_customers[fy]
    # Try previous FYs in descending order (closest first)
    if fy:
        sorted_fys = sorted(fy_customers.keys(), reverse=True)
        for prev in sorted_fys:
            if prev < fy:
                return fy_customers[prev]
    # Fallback to legacy flat customers field
    return master.get("customers", [])


def _get_quarter(month_str):
    """Given 'YYYY-MM', return quarter label like 'Q1 (Apr-Jun)'."""
    try:
        m = int(month_str.split('-')[1])
    except (IndexError, ValueError):
        return "Q?"
    if m in (4, 5, 6):
        return "Q1 (Apr-Jun)"
    elif m in (7, 8, 9):
        return "Q2 (Jul-Sep)"
    elif m in (10, 11, 12):
        return "Q3 (Oct-Dec)"
    else:
        return "Q4 (Jan-Mar)"


def _get_month_label(month_str):
    """Given 'YYYY-MM', return human-readable label like 'Apr 2025'."""
    months = {
        '01': 'Jan', '02': 'Feb', '03': 'Mar', '04': 'Apr',
        '05': 'May', '06': 'Jun', '07': 'Jul', '08': 'Aug',
        '09': 'Sep', '10': 'Oct', '11': 'Nov', '12': 'Dec'
    }
    try:
        parts = month_str.split('-')
        return f"{months.get(parts[1], parts[1])} {parts[0]}"
    except (IndexError, ValueError):
        return month_str


# ==================== MASTER CRUD ====================

@router.get("/salesman/master")
async def get_salesman_master(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        salesmen = await db.salesman_master.find(q, {"_id": 0}).to_list(100)

        current_fy = get_current_fy()
        target_fy = fy or current_fy

        # Auto-discover salesman-role users not yet in salesman_master
        user_q = {}
        if ctx and ctx.get("tenant_id"):
            user_q["tenant_id"] = ctx["tenant_id"]
        user_q["role"] = "salesman"
        salesman_users = await db.users.find(user_q, {"_id": 0, "name": 1, "username": 1}).to_list(100)
        existing_names = {m.get("salesman_name", "").lower() for m in salesmen}
        for su in salesman_users:
            name = (su.get("name") or su.get("username", "")).strip()
            if name.lower() not in existing_names:
                # Auto-create salesman_master record for this user
                doc = {
                    "salesman_id": str(uuid.uuid4()),
                    "salesman_name": name,
                    "phone": "", "email": "",
                    "monthly_target": 0, "quarterly_target": 0,
                    "customers": [],
                    "fy_targets": {}, "fy_customers": {},
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                if ctx and ctx.get("tenant_id"):
                    doc["tenant_id"] = ctx["tenant_id"]
                if ctx and ctx.get("company_id"):
                    doc["company_id"] = ctx["company_id"]
                await db.salesman_master.insert_one(doc)
                salesmen.append(doc)
                existing_names.add(name.lower())

        result = []
        for m in salesmen:
            targets = _get_fy_targets(m, target_fy)
            customers = _get_fy_customers(m, target_fy)
            fy_locked = is_fy_completed(target_fy)
            result.append({
                "salesman_name": m.get("salesman_name", ""),
                "phone": m.get("phone", ""),
                "email": m.get("email", ""),
                "monthly_target": safe_num(targets.get("monthly_target")),
                "quarterly_target": safe_num(targets.get("quarterly_target")),
                "customers": customers,
                "fy": target_fy,
                "fy_locked": fy_locked,
            })

        return APIResponse(success=True, data={
            "salesmen": sorted(result, key=lambda x: x["salesman_name"].lower()),
            "current_fy": current_fy,
            "target_fy": target_fy,
            "fy_locked": is_fy_completed(target_fy),
        })
    except Exception as e:
        logger.error(f"Error fetching salesman master: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/salesman/master")
async def create_salesman(request: Request):
    try:
        body = await request.json()
        ctx = await get_tenant_context(request)
        salesman_name = body.get("salesman_name", "").strip()
        if not salesman_name:
            return APIResponse(success=False, error="Salesman name is required")

        fy = body.get("fy") or get_current_fy()

        # Check if FY is completed - only allow target/mapping changes for current/future FYs
        if is_fy_completed(fy):
            return APIResponse(success=False, error=f"FY {fy} has ended. Cannot modify targets or mappings for completed financial years.")

        customers = body.get("customers", [])
        monthly_target = body.get("monthly_target", 0)
        quarterly_target = body.get("quarterly_target", 0)
        phone = body.get("phone", "")
        email = body.get("email", "")

        tq = _build_query(ctx)
        existing = await db.salesman_master.find_one({**tq, "salesman_name": salesman_name}, {"_id": 0})

        if existing:
            # Update: merge FY-specific targets and customers
            fy_targets = existing.get("fy_targets", {})
            fy_customers = existing.get("fy_customers", {})

            fy_targets[fy] = {
                "monthly_target": monthly_target,
                "quarterly_target": quarterly_target,
            }
            fy_customers[fy] = customers

            # Migrate legacy customers into the oldest known FY if not yet migrated
            legacy_customers = existing.get("customers", [])
            if legacy_customers and not any(v for v in fy_customers.values() if v == legacy_customers):
                # Preserve legacy mapping under the earliest FY we know about
                all_fys = sorted(fy_customers.keys())
                if all_fys:
                    earliest = all_fys[0]
                    if earliest != fy and earliest not in fy_customers:
                        fy_customers[earliest] = legacy_customers

            update_doc = {
                "fy_targets": fy_targets,
                "fy_customers": fy_customers,
                "phone": phone,
                "email": email,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            # NOTE: Do NOT overwrite legacy 'customers' or 'monthly_target' fields
            # to prevent cross-FY data corruption
            await db.salesman_master.update_one(
                {**tq, "salesman_name": salesman_name},
                {"$set": update_doc}
            )
        else:
            # New salesman
            doc = {
                "salesman_id": str(uuid.uuid4()),
                "salesman_name": salesman_name,
                "phone": phone,
                "email": email,
                "monthly_target": monthly_target,
                "quarterly_target": quarterly_target,
                "customers": customers,
                "fy_targets": {
                    fy: {
                        "monthly_target": monthly_target,
                        "quarterly_target": quarterly_target,
                    }
                },
                "fy_customers": {
                    fy: customers,
                },
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            if ctx and ctx.get("tenant_id"):
                doc["tenant_id"] = ctx["tenant_id"]
            if ctx and ctx.get("company_id"):
                doc["company_id"] = ctx["company_id"]
            await db.salesman_master.insert_one(doc)

        return APIResponse(
            success=True,
            message=f"Salesman '{salesman_name}' saved for FY {fy}",
        )
    except Exception as e:
        logger.error(f"Error creating salesman: {e}")
        return APIResponse(success=False, error=str(e))


@router.delete("/salesman/master/{salesman_name}")
async def delete_salesman(salesman_name: str, request: Request):
    try:
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        result = await db.salesman_master.delete_one({**tq, "salesman_name": salesman_name})
        return APIResponse(
            success=result.deleted_count > 0,
            message="Deleted" if result.deleted_count > 0 else "Not found"
        )
    except Exception as e:
        logger.error(f"Error deleting salesman: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== PERFORMANCE ====================

@router.get("/salesman/performance")
async def get_salesman_performance(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        target_fy = fy or get_current_fy()

        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        vouchers = filter_vouchers_by_fy(all_vouchers, target_fy)
        master_list = await db.salesman_master.find(q, {"_id": 0}).to_list(100)
        master_map = {m["salesman_name"]: m for m in master_list if m.get("salesman_name")}

        # Build customer-to-salesman mapping for the specific FY
        customer_to_salesman = {}
        for m in master_list:
            sname = m.get("salesman_name")
            if not sname:
                continue
            fy_custs = _get_fy_customers(m, target_fy)
            for cust in fy_custs:
                customer_to_salesman[cust.lower()] = sname

        salesman_map = {}
        for voucher in vouchers:
            customer = voucher.get("party_name", "")
            salesman = customer_to_salesman.get(customer.lower()) or voucher.get("salesman") or None
            if not salesman:
                continue
            amount = safe_num(voucher.get("total_amount"))

            if salesman not in salesman_map:
                salesman_map[salesman] = {"total_sales": 0, "customers": set(), "transactions": 0}
            salesman_map[salesman]["total_sales"] += amount
            salesman_map[salesman]["customers"].add(customer)
            salesman_map[salesman]["transactions"] += 1

        # Include registered salesmen with no sales
        for m in master_list:
            name = m.get("salesman_name")
            if not name or name in salesman_map:
                continue
            salesman_map[name] = {"total_sales": 0, "customers": set(_get_fy_customers(m, target_fy)), "transactions": 0}

        performance = []
        for salesman, data in salesman_map.items():
            master = master_map.get(salesman, {})
            targets = _get_fy_targets(master, target_fy)
            monthly_target = safe_num(targets.get("monthly_target"))
            annual_target = monthly_target * 12 if monthly_target else 0
            performance.append({
                "salesman_name": salesman,
                "monthly_target": monthly_target,
                "target_amount": annual_target,
                "achieved_amount": data["total_sales"],
                "achievement_percentage": (data["total_sales"] / annual_target * 100) if annual_target > 0 else 0,
                "total_customers": len(data["customers"]),
                "total_transactions": data["transactions"],
                "has_master": salesman in master_map,
            })

        performance.sort(key=lambda x: x["achieved_amount"], reverse=True)
        return APIResponse(success=True, data={"salesman": performance})

    except Exception as e:
        logger.error(f"Error fetching salesman performance: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== DETAILED PERFORMANCE (Monthly/Quarterly/Annual) ====================

@router.get("/salesman/performance-detailed")
async def get_salesman_performance_detailed(
    request: Request,
    fy: Optional[str] = None,
    duration: Optional[str] = "monthly",
    company_id: Optional[str] = None
):
    """Performance breakdown by monthly/quarterly/annual per salesman per customer."""
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        target_fy = fy or get_current_fy()

        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        vouchers = filter_vouchers_by_fy(all_vouchers, target_fy)
        master_list = await db.salesman_master.find(q, {"_id": 0}).to_list(100)
        master_map = {m["salesman_name"]: m for m in master_list if m.get("salesman_name")}

        # Build customer-to-salesman mapping for the specific FY
        customer_to_salesman = {}
        for m in master_list:
            sname = m.get("salesman_name")
            if not sname:
                continue
            fy_custs = _get_fy_customers(m, target_fy)
            for cust in fy_custs:
                customer_to_salesman[cust.lower()] = sname

        # Group vouchers by salesman -> customer -> period
        salesman_data = {}
        salesman_items = {}  # salesman -> item_name -> {qty, revenue, count}
        for voucher in vouchers:
            customer = voucher.get("party_name", "")
            salesman = customer_to_salesman.get(customer.lower()) or voucher.get("salesman") or None
            if not salesman:
                continue

            amount = safe_num(voucher.get("total_amount"))
            date_str = voucher.get("voucher_date", "")
            month = date_str[:7] if date_str else ""
            quarter = _get_quarter(month)

            if salesman not in salesman_data:
                salesman_data[salesman] = {}
            if salesman not in salesman_items:
                salesman_items[salesman] = {}
            if customer not in salesman_data[salesman]:
                salesman_data[salesman][customer] = {
                    "monthly": defaultdict(lambda: {"amount": 0, "count": 0}),
                    "quarterly": defaultdict(lambda: {"amount": 0, "count": 0}),
                    "annual": {"amount": 0, "count": 0},
                }

            cust_data = salesman_data[salesman][customer]
            cust_data["monthly"][month]["amount"] += amount
            cust_data["monthly"][month]["count"] += 1
            cust_data["quarterly"][quarter]["amount"] += amount
            cust_data["quarterly"][quarter]["count"] += 1
            cust_data["annual"]["amount"] += amount
            cust_data["annual"]["count"] += 1

            # Track item-wise sales from voucher line items
            for item in voucher.get("inventory_entries", voucher.get("items", [])):
                item_name = item.get("item", item.get("item_name", ""))
                if not item_name:
                    continue
                qty = safe_num(item.get("quantity", item.get("qty", 0)))
                item_amount = safe_num(item.get("amount", item.get("total_amount", 0)))
                if item_name not in salesman_items[salesman]:
                    salesman_items[salesman][item_name] = {"item_name": item_name, "total_quantity": 0, "total_revenue": 0, "transaction_count": 0}
                salesman_items[salesman][item_name]["total_quantity"] += qty
                salesman_items[salesman][item_name]["total_revenue"] += item_amount
                salesman_items[salesman][item_name]["transaction_count"] += 1

        # Include registered salesmen with no sales
        for m in master_list:
            name = m.get("salesman_name")
            if not name or name in salesman_data:
                continue
            salesman_data[name] = {}

        # Build all period columns for consistent table headers
        all_months = set()
        all_quarters = set()
        for sdata in salesman_data.values():
            for cdata in sdata.values():
                all_months.update(cdata.get("monthly", {}).keys())
                all_quarters.update(cdata.get("quarterly", {}).keys())

        sorted_months = sorted(all_months)
        quarter_order = ["Q1 (Apr-Jun)", "Q2 (Jul-Sep)", "Q3 (Oct-Dec)", "Q4 (Jan-Mar)"]
        sorted_quarters = [qr for qr in quarter_order if qr in all_quarters]

        # Build response per salesman
        performance = []
        for salesman, customers_data in salesman_data.items():
            master = master_map.get(salesman, {})
            targets = _get_fy_targets(master, target_fy)
            monthly_target = safe_num(targets.get("monthly_target"))
            quarterly_target = safe_num(targets.get("quarterly_target"))
            annual_target = monthly_target * 12 if monthly_target else 0

            total_achieved = sum(cd["annual"]["amount"] for cd in customers_data.values())
            total_txns = sum(cd["annual"]["count"] for cd in customers_data.values())

            # Per-customer breakdown
            customer_breakdown = []
            for cust_name, cdata in sorted(customers_data.items(), key=lambda x: x[0].lower()):
                row = {
                    "customer_name": cust_name,
                    "annual_amount": round(cdata["annual"]["amount"], 2),
                    "annual_count": cdata["annual"]["count"],
                }
                # Monthly columns
                monthly_vals = {}
                for m in sorted_months:
                    d = cdata["monthly"].get(m, {"amount": 0, "count": 0})
                    monthly_vals[m] = {"amount": round(d["amount"], 2), "count": d["count"]}
                row["monthly"] = monthly_vals

                # Quarterly columns
                quarterly_vals = {}
                for qr in sorted_quarters:
                    d = cdata["quarterly"].get(qr, {"amount": 0, "count": 0})
                    quarterly_vals[qr] = {"amount": round(d["amount"], 2), "count": d["count"]}
                row["quarterly"] = quarterly_vals

                customer_breakdown.append(row)

            # Weighted average: weighted by revenue contribution
            weighted_achievement = 0
            if annual_target > 0 and total_achieved > 0:
                weighted_achievement = round(total_achieved / annual_target * 100, 1)

            # Items sold by this salesman
            items = salesman_items.get(salesman, {})
            items_breakdown = sorted(items.values(), key=lambda x: x["total_revenue"], reverse=True)

            performance.append({
                "salesman_name": salesman,
                "phone": master.get("phone", ""),
                "email": master.get("email", ""),
                "monthly_target": monthly_target,
                "quarterly_target": quarterly_target if quarterly_target else monthly_target * 3,
                "annual_target": annual_target,
                "achieved_amount": round(total_achieved, 2),
                "achievement_percentage": weighted_achievement,
                "total_customers": len(customers_data),
                "total_transactions": total_txns,
                "mapped_customers": _get_fy_customers(master, target_fy),
                "customers": customer_breakdown,
                "items_sold": items_breakdown,
                "has_master": salesman in master_map,
            })

        performance.sort(key=lambda x: x["achieved_amount"], reverse=True)

        return APIResponse(success=True, data={
            "salesman": performance,
            "periods": {
                "months": sorted_months,
                "month_labels": {m: _get_month_label(m) for m in sorted_months},
                "quarters": sorted_quarters,
            },
            "fy": target_fy,
            "current_fy": get_current_fy(),
        })

    except Exception as e:
        logger.error(f"Error fetching detailed salesman performance: {e}")
        return APIResponse(success=False, error=str(e))


# ==================== EXCEL EXPORT ====================

@router.get("/salesman/export")
async def export_salesman_performance(
    request: Request,
    salesman_name: str,
    fy: Optional[str] = None,
    duration: Optional[str] = "monthly",
    company_id: Optional[str] = None
):
    """Export a salesman's customer-wise performance to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        target_fy = fy or get_current_fy()

        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        vouchers = filter_vouchers_by_fy(all_vouchers, target_fy)
        master_list = await db.salesman_master.find(q, {"_id": 0}).to_list(100)

        # Build customer-to-salesman mapping for the specific FY
        customer_to_salesman = {}
        for m in master_list:
            sname = m.get("salesman_name")
            if not sname:
                continue
            fy_custs = _get_fy_customers(m, target_fy)
            for cust in fy_custs:
                customer_to_salesman[cust.lower()] = sname

        # Group vouchers for this salesman
        cust_data = defaultdict(lambda: {"monthly": defaultdict(float), "quarterly": defaultdict(float), "annual": 0})
        for voucher in vouchers:
            customer = voucher.get("party_name", "")
            sm = customer_to_salesman.get(customer.lower()) or voucher.get("salesman") or None
            if sm != salesman_name:
                continue
            amount = safe_num(voucher.get("total_amount"))
            date_str = voucher.get("voucher_date", "")
            month = date_str[:7] if date_str else ""
            quarter = _get_quarter(month)
            cust_data[customer]["monthly"][month] += amount
            cust_data[customer]["quarterly"][quarter] += amount
            cust_data[customer]["annual"] += amount

        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{duration.capitalize()} - {salesman_name[:20]}"

        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Title row
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Salesman: {salesman_name} | FY: {target_fy} | View: {duration.capitalize()}"
        ws['A1'].font = Font(bold=True, size=12)
        ws.append([])

        if duration == "monthly":
            all_months = sorted(set(m for cd in cust_data.values() for m in cd["monthly"].keys()))
            headers = ["Customer"] + [_get_month_label(m) for m in all_months] + ["Total"]
        elif duration == "quarterly":
            quarter_order = ["Q1 (Apr-Jun)", "Q2 (Jul-Sep)", "Q3 (Oct-Dec)", "Q4 (Jan-Mar)"]
            all_quarters = [qr for qr in quarter_order if any(qr in cd["quarterly"] for cd in cust_data.values())]
            headers = ["Customer"] + all_quarters + ["Total"]
        else:
            headers = ["Customer", "Annual Total"]

        ws.append(headers)
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        for customer in sorted(cust_data.keys()):
            cd = cust_data[customer]
            if duration == "monthly":
                row = [customer] + [round(cd["monthly"].get(m, 0), 2) for m in all_months] + [round(cd["annual"], 2)]
            elif duration == "quarterly":
                row = [customer] + [round(cd["quarterly"].get(qr, 0), 2) for qr in all_quarters] + [round(cd["annual"], 2)]
            else:
                row = [customer, round(cd["annual"], 2)]
            ws.append(row)

        # Auto-width
        for col_cells in ws.columns:
            valid_cells = [c for c in col_cells if not isinstance(c, openpyxl.cell.cell.MergedCell)]
            if not valid_cells:
                continue
            max_len = max((len(str(cell.value or "")) for cell in valid_cells), default=10)
            ws.column_dimensions[valid_cells[0].column_letter].width = min(max_len + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"salesman_{salesman_name.replace(' ', '_')}_{target_fy}_{duration}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )

    except Exception as e:
        logger.error(f"Error exporting salesman data: {e}")
        return APIResponse(success=False, error=str(e))
