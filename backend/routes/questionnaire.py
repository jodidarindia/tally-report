"""Questionnaire routes — public submission + super-admin retrieval & Excel export."""
from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from datetime import datetime, timezone
import logging
import io

from db import db
from models import APIResponse
from services.auth_service import get_current_user

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("/questionnaire/submit")
async def submit_questionnaire(request: Request):
    """Public endpoint — anyone can submit a questionnaire."""
    try:
        body = await request.json()
        # iter-123: DPDP consent enforcement + audit fields
        consent_given = bool(body.get("consent_given"))
        if not consent_given:
            return APIResponse(success=False, error="Consent is required under the DPDP Act, 2023")

        client_ip = request.client.host if request.client else ""
        now_iso = datetime.now(timezone.utc).isoformat()

        doc = {
            "company_name": body.get("company_name", ""),
            "contact_person": body.get("contact_person", ""),
            "designation": body.get("designation", ""),
            "phone": body.get("phone", ""),
            "email": body.get("email", ""),
            "city": body.get("city", ""),
            "industry": body.get("industry", ""),
            "employees": body.get("employees", ""),
            "turnover": body.get("turnover", ""),
            "tally_version": body.get("tally_version", ""),
            "tally_companies": body.get("tally_companies", ""),
            "tally_users": body.get("tally_users", ""),
            "has_branches": body.get("has_branches", ""),
            "branch_count": body.get("branch_count", ""),
            "remote_access": body.get("remote_access", []),
            "tally_users_roles": body.get("tally_users_roles", []),
            "pain_points": body.get("pain_points", []),
            "biggest_challenge": body.get("biggest_challenge", ""),
            "feature_ratings": body.get("feature_ratings", {}),
            "decision_factors": body.get("decision_factors", []),
            "timeline": body.get("timeline", ""),
            "decision_maker": body.get("decision_maker", ""),
            "budget": body.get("budget", ""),
            "additional_features": body.get("additional_features", ""),
            "heard_from": body.get("heard_from", ""),
            "next_steps": body.get("next_steps", []),
            "callback_time": body.get("callback_time", ""),
            "notes": body.get("notes", ""),
            "submitted_by": body.get("submitted_by", "prospect"),
            "submitted_at": now_iso,
            "status": "new",
            # iter-123 DPDP audit trail
            "consent_given":   True,
            "consent_version": "dpdp-v1-2026-02",
            "consent_ts":      now_iso,
            "consent_ip":      client_ip,
        }

        if not doc["company_name"] and not doc["contact_person"] and not doc["phone"]:
            return APIResponse(success=False, error="Please fill at least company name, contact person, or phone number.")

        await db.questionnaires.insert_one(doc)

        # Fire prospect welcome email if we have an email address.
        try:
            if doc.get("email"):
                import asyncio
                from services.email_service import send_prospect_welcome_email
                asyncio.create_task(send_prospect_welcome_email({
                    "prospect_id":    "Q-" + doc["submitted_at"][:19].replace("-", "").replace(":", "").replace("T", ""),
                    "company_name":   doc["company_name"],
                    "contact_person": doc["contact_person"],
                    "email":          doc["email"],
                    "selected_plan":  doc.get("budget") or "Free Trial",
                }))
        except Exception as mail_err:
            logger.warning(f"lead welcome email failed: {mail_err}")

        return APIResponse(success=True, data={"message": "Questionnaire submitted successfully. Our team will reach out shortly."})
    except Exception as e:
        logger.error(f"Questionnaire submit error: {e}")
        return APIResponse(success=False, error="Failed to submit questionnaire.")


@router.get("/super-admin/questionnaires")
async def list_questionnaires(request: Request):
    """Super-admin: list all questionnaire submissions."""
    user = await get_current_user(request, db)
    if not user or user.get("role") != "super_admin":
        return APIResponse(success=False, error="Super admin access required")
    try:
        docs = await db.questionnaires.find({}, {"_id": 0}).sort("submitted_at", -1).to_list(1000)
        return APIResponse(success=True, data={"questionnaires": docs, "total": len(docs)})
    except Exception as e:
        logger.error(f"List questionnaires error: {e}")
        return APIResponse(success=False, error="Failed to fetch questionnaires.")


@router.get("/super-admin/questionnaires/export")
async def export_questionnaires(request: Request):
    """Super-admin: export all questionnaires to Excel."""
    user = await get_current_user(request, db)
    if not user or user.get("role") != "super_admin":
        return APIResponse(success=False, error="Super admin access required")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        docs = await db.questionnaires.find({}, {"_id": 0}).sort("submitted_at", -1).to_list(5000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Questionnaires"

        headers = [
            "Submitted At", "Status", "Submitted By",
            "Company Name", "Contact Person", "Designation", "Phone", "Email", "City",
            "Industry", "Employees", "Turnover",
            "Tally Version", "Tally Companies", "Tally Users", "Has Branches", "Branch Count",
            "Remote Access Methods", "Tally User Roles",
            "Pain Points", "Biggest Challenge",
            "Feature Ratings", "Decision Factors", "Timeline", "Decision Maker", "Budget",
            "Additional Features", "Heard From", "Next Steps",
            "Callback Time", "Notes",
        ]

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_num, doc in enumerate(docs, 2):
            ws.cell(row=row_num, column=1, value=doc.get("submitted_at", ""))
            ws.cell(row=row_num, column=2, value=doc.get("status", ""))
            ws.cell(row=row_num, column=3, value=doc.get("submitted_by", ""))
            ws.cell(row=row_num, column=4, value=doc.get("company_name", ""))
            ws.cell(row=row_num, column=5, value=doc.get("contact_person", ""))
            ws.cell(row=row_num, column=6, value=doc.get("designation", ""))
            ws.cell(row=row_num, column=7, value=doc.get("phone", ""))
            ws.cell(row=row_num, column=8, value=doc.get("email", ""))
            ws.cell(row=row_num, column=9, value=doc.get("city", ""))
            ws.cell(row=row_num, column=10, value=doc.get("industry", ""))
            ws.cell(row=row_num, column=11, value=doc.get("employees", ""))
            ws.cell(row=row_num, column=12, value=doc.get("turnover", ""))
            ws.cell(row=row_num, column=13, value=doc.get("tally_version", ""))
            ws.cell(row=row_num, column=14, value=doc.get("tally_companies", ""))
            ws.cell(row=row_num, column=15, value=doc.get("tally_users", ""))
            ws.cell(row=row_num, column=16, value=doc.get("has_branches", ""))
            ws.cell(row=row_num, column=17, value=doc.get("branch_count", ""))
            ws.cell(row=row_num, column=18, value=", ".join(doc.get("remote_access", [])))
            ws.cell(row=row_num, column=19, value=", ".join(doc.get("tally_users_roles", [])))
            ws.cell(row=row_num, column=20, value=", ".join(doc.get("pain_points", [])))
            ws.cell(row=row_num, column=21, value=doc.get("biggest_challenge", ""))
            ratings = doc.get("feature_ratings", {})
            ws.cell(row=row_num, column=22, value="; ".join(f"{k}={v}" for k, v in ratings.items()) if ratings else "")
            ws.cell(row=row_num, column=23, value=", ".join(doc.get("decision_factors", [])))
            ws.cell(row=row_num, column=24, value=doc.get("timeline", ""))
            ws.cell(row=row_num, column=25, value=doc.get("decision_maker", ""))
            ws.cell(row=row_num, column=26, value=doc.get("budget", ""))
            ws.cell(row=row_num, column=27, value=doc.get("additional_features", ""))
            ws.cell(row=row_num, column=28, value=doc.get("heard_from", ""))
            ws.cell(row=row_num, column=29, value=", ".join(doc.get("next_steps", [])))
            ws.cell(row=row_num, column=30, value=doc.get("callback_time", ""))
            ws.cell(row=row_num, column=31, value=doc.get("notes", ""))

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=FLOWRA_Questionnaires.xlsx"}
        )
    except Exception as e:
        logger.error(f"Export questionnaires error: {e}")
        return APIResponse(success=False, error="Failed to export questionnaires.")


@router.put("/super-admin/questionnaires/{idx}/status")
async def update_questionnaire_status(idx: int, request: Request):
    """Super-admin: update status of a questionnaire entry by index."""
    user = await get_current_user(request, db)
    if not user or user.get("role") != "super_admin":
        return APIResponse(success=False, error="Super admin access required")
    try:
        body = await request.json()
        new_status = body.get("status", "reviewed")

        docs = await db.questionnaires.find({}, {"_id": 1}).sort("submitted_at", -1).to_list(5000)
        if idx < 0 or idx >= len(docs):
            return APIResponse(success=False, error="Invalid index")

        doc_id = docs[idx]["_id"]
        await db.questionnaires.update_one({"_id": doc_id}, {"$set": {"status": new_status}})
        return APIResponse(success=True, data={"message": f"Status updated to {new_status}"})
    except Exception as e:
        logger.error(f"Update questionnaire status error: {e}")
        return APIResponse(success=False, error="Failed to update status.")
