from fastapi import APIRouter, Request
from typing import Optional
import logging

from db import db
from models import SalesVoucher, APIResponse
from utils import safe_num, filter_vouchers_by_fy
from services.tenant_context import get_tenant_context
from routes.branch_ledgers import get_branch_parties

logger = logging.getLogger(__name__)
router = APIRouter()


def _build_query(ctx, company_id=None, extra=None):
    q = {}
    if ctx and ctx.get("tenant_id"):
        q["tenant_id"] = ctx["tenant_id"]
    cid = company_id or (ctx.get("company_id") if ctx else None)
    if cid:
        q["company_id"] = cid
    if extra:
        q.update(extra)
    return q


async def _apply_branch_filter(q, ctx, request: Request):
    """Add $nin filter for branch party names if X-Exclude-Branches header is set."""
    exclude = request.headers.get("X-Exclude-Branches", "").lower() == "true"
    if not exclude:
        return q
    tenant_id = ctx.get("tenant_id", "") if ctx else ""
    company_id = ctx.get("company_id", "") if ctx else ""
    if not tenant_id or not company_id:
        return q
    branch_parties = await get_branch_parties(tenant_id, company_id)
    if branch_parties:
        q["party_name"] = {"$nin": branch_parties}
    return q


@router.get("/sales/vouchers")
async def get_sales_vouchers(request: Request, start_date: Optional[str] = None, end_date: Optional[str] = None, party_name: Optional[str] = None, fy: Optional[str] = None, month: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        extra = {}
        if party_name:
            extra["party_name"] = {"$regex": party_name, "$options": "i"}

        query = _build_query(ctx, company_id, extra)
        query = await _apply_branch_filter(query, ctx, request)
        vouchers = await db.sales_vouchers.find(query, {"_id": 0}).to_list(10000)

        if fy:
            vouchers = filter_vouchers_by_fy(vouchers, fy)

        if month:
            if len(month) <= 2:
                vouchers = [v for v in vouchers if v.get("voucher_date", "")[5:7] == month.zfill(2)]
            else:
                vouchers = [v for v in vouchers if v.get("voucher_date", "").startswith(month)]

        if vouchers and (start_date or end_date):
            filtered = []
            for v in vouchers:
                v_date = v.get("voucher_date", "")
                if start_date and v_date < start_date:
                    continue
                if end_date and v_date > end_date:
                    continue
                filtered.append(v)
            vouchers = filtered

        base_q = _build_query(ctx, company_id)
        all_vouchers_for_meta = await db.sales_vouchers.find(base_q, {"_id": 0, "party_name": 1, "voucher_date": 1}).to_list(10000)
        if fy:
            all_vouchers_for_meta = filter_vouchers_by_fy(all_vouchers_for_meta, fy)

        unique_parties = sorted(list(set(v.get("party_name", "") for v in all_vouchers_for_meta if v.get("party_name"))))
        unique_months = sorted(list(set(v.get("voucher_date", "")[:7] for v in all_vouchers_for_meta if v.get("voucher_date", "")[:7])))

        return APIResponse(
            success=True,
            data={
                "vouchers": vouchers,
                "count": len(vouchers),
                "unique_parties": unique_parties,
                "unique_months": unique_months
            }
        )
    except Exception as e:
        logger.error(f"Error fetching sales vouchers: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/sales/vouchers/{voucher_id:path}")
async def get_voucher_detail(voucher_id: str, request: Request):
    try:
        from urllib.parse import unquote
        decoded_id = unquote(voucher_id)
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)

        voucher = await db.sales_vouchers.find_one({**tq, "voucher_id": decoded_id}, {"_id": 0})
        if not voucher:
            voucher = await db.sales_vouchers.find_one(
                {**tq, "voucher_id": {"$regex": f"^{decoded_id}$", "$options": "i"}},
                {"_id": 0}
            )
        if not voucher:
            return APIResponse(success=False, error="Voucher not found")

        items = voucher.get("items", [])
        subtotal = sum(safe_num(item.get("amount"), safe_num(item.get("quantity")) * safe_num(item.get("rate"))) for item in items)
        total = safe_num(voucher.get("total_amount"), subtotal)

        discount_amount = 0
        gst_details = []
        dispatch_details = {}

        ledger_entries = voucher.get("ledger_entries", [])
        for entry in ledger_entries:
            if isinstance(entry, dict):
                ledger_name = str(entry.get("ledger_name", "")).lower()
                amount = entry.get("amount", 0)
                if "discount" in ledger_name:
                    discount_amount += abs(safe_num(amount))
                elif "gst" in ledger_name or "cgst" in ledger_name or "sgst" in ledger_name or "igst" in ledger_name or "tax" in ledger_name:
                    gst_details.append({
                        "tax_name": entry.get("ledger_name", ""),
                        "amount": abs(safe_num(amount))
                    })

        dispatch_details = {
            "delivery_note": voucher.get("delivery_note", voucher.get("reference_number", "")),
            "dispatch_through": voucher.get("dispatch_through", ""),
            "destination": voucher.get("destination", ""),
            "carrier_name": voucher.get("carrier_name", ""),
            "bill_of_lading": voucher.get("bill_of_lading", ""),
            "dispatch_date": voucher.get("dispatch_date", voucher.get("voucher_date", ""))
        }

        gst_total = sum(g.get("amount", 0) for g in gst_details)

        voucher["subtotal"] = round(subtotal, 2)
        voucher["discount_amount"] = round(discount_amount, 2)
        voucher["gst_details"] = gst_details
        voucher["gst_total"] = round(gst_total, 2)
        voucher["dispatch_details"] = dispatch_details
        voucher["computed_total"] = round(total, 2)
        voucher["item_count"] = len(items)

        return APIResponse(success=True, data=voucher)
    except Exception as e:
        logger.error(f"Error fetching voucher detail: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/sales/summary")
async def get_sales_summary(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        q = await _apply_branch_filter(q, ctx, request)
        vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        if fy:
            vouchers = filter_vouchers_by_fy(vouchers, fy)

        if not vouchers:
            return APIResponse(
                success=True,
                data={"total_vouchers": 0, "total_sales": 0, "top_customers": [], "recent_vouchers": []}
            )

        total_vouchers = len(vouchers)
        total_sales = sum(safe_num(v.get("total_amount")) for v in vouchers)

        customer_sales = {}
        for v in vouchers:
            party = v.get("party_name", "Unknown")
            customer_sales[party] = customer_sales.get(party, 0) + safe_num(v.get("total_amount"))

        top_customers = sorted(
            [{"name": k, "total": round(v, 2)} for k, v in customer_sales.items()],
            key=lambda x: x["total"],
            reverse=True
        )[:10]

        recent_vouchers = sorted(
            vouchers,
            key=lambda x: x.get("voucher_date", ""),
            reverse=True
        )[:10]

        return APIResponse(
            success=True,
            data={
                "total_vouchers": total_vouchers,
                "total_sales": round(total_sales, 2),
                "top_customers": top_customers,
                "recent_vouchers": recent_vouchers
            }
        )
    except Exception as e:
        logger.error(f"Error getting sales summary: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/sales/analytics")
async def get_sales_analytics(request: Request, fy: Optional[str] = None, party_name: Optional[str] = None, month: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        extra = {}
        if party_name:
            extra["party_name"] = {"$regex": party_name, "$options": "i"}
        q = _build_query(ctx, company_id, extra)
        q = await _apply_branch_filter(q, ctx, request)
        vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        vouchers = filter_vouchers_by_fy(vouchers, fy)

        if month:
            if len(month) <= 2:
                vouchers = [v for v in vouchers if v.get("voucher_date", "")[5:7] == month.zfill(2)]
            else:
                vouchers = [v for v in vouchers if v.get("voucher_date", "").startswith(month)]

        if not vouchers:
            return APIResponse(success=True, data={"daily_sales": [], "category_sales": []})

        daily_sales = {}
        for v in vouchers:
            date = v.get("voucher_date", "Unknown")
            daily_sales[date] = daily_sales.get(date, 0) + safe_num(v.get("total_amount"))

        daily_sales_data = sorted(
            [{"date": k, "amount": v} for k, v in daily_sales.items()],
            key=lambda x: x["date"]
        )

        return APIResponse(success=True, data={"daily_sales": daily_sales_data})
    except Exception as e:
        logger.error(f"Error getting sales analytics: {e}")
        return APIResponse(success=False, error=str(e))



@router.get("/sales/customer-names")
async def get_customer_names(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Get distinct customer (party) names from sales vouchers for combobox."""
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)
        q = await _apply_branch_filter(q, ctx, request)
        vouchers = await db.sales_vouchers.find(q, {"_id": 0, "party_name": 1, "voucher_date": 1}).to_list(50000)
        if fy:
            vouchers = filter_vouchers_by_fy(vouchers, fy)
        names = sorted(set(v.get("party_name", "") for v in vouchers if v.get("party_name")))
        return APIResponse(success=True, data={"customers": names, "total": len(names)})
    except Exception as e:
        logger.error(f"Error fetching customer names: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/sales/customer-item-sales")
async def get_customer_item_sales(request: Request, customer: str = "", fy: Optional[str] = None, company_id: Optional[str] = None):
    """Get item-wise sales breakdown for a specific customer in a given FY."""
    try:
        ctx = await get_tenant_context(request)
        if not customer:
            return APIResponse(success=False, error="Customer name is required")

        # If branch exclusion is on, check if this customer is a branch party
        exclude = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        if exclude:
            from routes.branch_ledgers import get_branch_parties
            bp = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
            if customer in bp:
                return APIResponse(success=True, data={"customer": customer, "financial_year": fy, "items": [], "total_items": 0, "total_quantity": 0, "total_amount": 0, "total_vouchers": 0})

        q = _build_query(ctx, company_id, {"party_name": customer})
        vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        if fy:
            vouchers = filter_vouchers_by_fy(vouchers, fy)

        # Aggregate item-wise: item_name -> {quantity, amount, voucher_count, rates}
        from collections import defaultdict
        items_agg = defaultdict(lambda: {"quantity": 0.0, "amount": 0.0, "voucher_count": 0, "rates": []})

        total_quantity = 0.0
        total_amount = 0.0

        for v in vouchers:
            seen_items_in_voucher = set()
            for item in v.get("items", []):
                item_name = item.get("item", "Unknown")
                qty = safe_num(item.get("quantity", 0))
                amt = safe_num(item.get("amount", 0))
                rate = safe_num(item.get("rate", 0))

                items_agg[item_name]["quantity"] += qty
                items_agg[item_name]["amount"] += amt
                if rate and rate not in items_agg[item_name]["rates"]:
                    items_agg[item_name]["rates"].append(rate)

                if item_name not in seen_items_in_voucher:
                    items_agg[item_name]["voucher_count"] += 1
                    seen_items_in_voucher.add(item_name)

                total_quantity += qty
                total_amount += amt

        # Build result
        result_items = []
        for item_name, agg in sorted(items_agg.items(), key=lambda x: x[1]["amount"], reverse=True):
            avg_rate = agg["amount"] / agg["quantity"] if agg["quantity"] else 0
            result_items.append({
                "item_name": item_name,
                "quantity": round(agg["quantity"], 2),
                "amount": round(agg["amount"], 2),
                "avg_rate": round(avg_rate, 2),
                "voucher_count": agg["voucher_count"],
                "rates": sorted(agg["rates"])
            })

        return APIResponse(success=True, data={
            "customer": customer,
            "financial_year": fy,
            "items": result_items,
            "total_items": len(result_items),
            "total_quantity": round(total_quantity, 2),
            "total_amount": round(total_amount, 2),
            "total_vouchers": len(vouchers)
        })
    except Exception as e:
        logger.error(f"Error fetching customer item sales: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/sales/customer-item-sales-export")
async def export_customer_item_sales(request: Request, customer: str = "", fy: Optional[str] = None, company_id: Optional[str] = None):
    """Export customer item-wise sales to Excel."""
    from fastapi.responses import StreamingResponse
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        return APIResponse(success=False, error="openpyxl not installed")

    try:
        ctx = await get_tenant_context(request)
        if not customer:
            return APIResponse(success=False, error="Customer name is required")

        q = _build_query(ctx, company_id, {"party_name": customer})
        vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        if fy:
            vouchers = filter_vouchers_by_fy(vouchers, fy)

        from collections import defaultdict
        items_agg = defaultdict(lambda: {"quantity": 0.0, "amount": 0.0, "rates": [], "voucher_count": 0})
        total_qty = 0.0
        total_amt = 0.0
        for v in vouchers:
            seen = set()
            for item in v.get("items", []):
                name = item.get("item", "Unknown")
                qty = safe_num(item.get("quantity", 0))
                amt = safe_num(item.get("amount", 0))
                rate = safe_num(item.get("rate", 0))
                items_agg[name]["quantity"] += qty
                items_agg[name]["amount"] += amt
                if rate and rate not in items_agg[name]["rates"]:
                    items_agg[name]["rates"].append(rate)
                if name not in seen:
                    items_agg[name]["voucher_count"] += 1
                    seen.add(name)
                total_qty += qty
                total_amt += amt

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Item Sales"

        # Header
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Customer Item-wise Sales — {customer}"
        ws["A1"].font = header_font
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:F2")
        ws["A2"] = f"Financial Year: {fy or 'All'} | Total Vouchers: {len(vouchers)} | Total Items: {len(items_agg)}"
        ws["A2"].font = Font(size=10, color="666666")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Column headers
        col_headers = ["#", "Item Name", "Quantity", "Avg Rate", "Amount", "Invoice Count"]
        col_font = Font(bold=True, size=10, color="FFFFFF")
        col_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, h in enumerate(col_headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = col_font
            cell.fill = col_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Data rows
        sorted_items = sorted(items_agg.items(), key=lambda x: x[1]["amount"], reverse=True)
        row = 5
        for idx, (item_name, agg) in enumerate(sorted_items, 1):
            avg_rate = agg["amount"] / agg["quantity"] if agg["quantity"] else 0
            ws.cell(row=row, column=1, value=idx).border = thin_border
            ws.cell(row=row, column=2, value=item_name).border = thin_border
            c = ws.cell(row=row, column=3, value=round(agg["quantity"], 2))
            c.border = thin_border
            c.number_format = '#,##0.00'
            c = ws.cell(row=row, column=4, value=round(avg_rate, 2))
            c.border = thin_border
            c.number_format = '#,##0.00'
            c = ws.cell(row=row, column=5, value=round(agg["amount"], 2))
            c.border = thin_border
            c.number_format = '#,##0.00'
            ws.cell(row=row, column=6, value=agg["voucher_count"]).border = thin_border
            row += 1

        # Totals row
        total_fill = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
        total_font = Font(bold=True, size=10)
        ws.cell(row=row, column=1, value="").fill = total_fill
        ws.cell(row=row, column=2, value="TOTAL").font = total_font
        ws.cell(row=row, column=2).fill = total_fill
        c = ws.cell(row=row, column=3, value=round(total_qty, 2))
        c.font = total_font
        c.fill = total_fill
        c.number_format = '#,##0.00'
        ws.cell(row=row, column=4, value="").fill = total_fill
        c = ws.cell(row=row, column=5, value=round(total_amt, 2))
        c.font = total_font
        c.fill = total_fill
        c.number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=len(vouchers)).font = total_font
        ws.cell(row=row, column=6).fill = total_fill

        # Column widths
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 45
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_name = customer.replace(" ", "_").replace(",", "").replace("/", "_")[:30]
        filename = f"customer_items_{safe_name}_{fy or 'all'}.xlsx"

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting customer item sales: {e}")
        return APIResponse(success=False, error=str(e))
