from fastapi import APIRouter, Request
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timezone
import logging

from db import db
from models import (
    CustomerFollowup, CustomerFollowupCreate, APIResponse
)
from utils import safe_num, safe_str, filter_vouchers_by_fy, fy_to_date_range, get_previous_fy, get_jv_party_amount, fuzzy_match
from services.auth_service import get_current_user
from services.export_service import ExportService
from services.tenant_context import get_tenant_context
from services.customer_metrics import (
    fy_start_iso,
    base_fy_start_iso,
    split_by_fy as _split_by_fy,
    split_receipts_and_payments,
    filter_branch_parties,
    compute_opening_balance_map,
    aggregate_party_credits,
    apply_fifo_aging,
    aging_status,
)

from services.audit_service import log_audit, get_client_ip
from routes.branch_ledgers import get_branch_parties

logger = logging.getLogger(__name__)


async def _resolve_company_name(ctx) -> str:
    """iter-121: fetch the useradmin's synced company name so every export
    (Outstanding / Targets / Ledger PDF etc.) can stamp it as the header
    banner. Falls back to '' when the tenant hasn't synced yet."""
    if not ctx or not ctx.get("tenant_id"):
        return ""
    tenant_id = ctx["tenant_id"]
    company_id = ctx.get("company_id")
    try:
        if company_id:
            from services.id_mapping_service import get_company_name
            name = (await get_company_name(tenant_id, company_id) or "").strip()
            if name:
                return name
        doc = await db.sync_status.find_one(
            {"tenant_id": tenant_id, "type": "agent_sync"},
            {"_id": 0, "company_name": 1},
            sort=[("last_sync", -1)],
        )
        if doc:
            return (doc.get("company_name") or "").strip()
    except Exception as e:
        logger.warning(f"_resolve_company_name failed: {e}")
    return ""


router = APIRouter()


def _build_query(ctx, company_id=None, extra=None):
    q = {}
    if ctx and ctx.get("tenant_id"):
        q["tenant_id"] = ctx["tenant_id"]
    cid = company_id or (ctx.get("company_id") if ctx else None)
    if cid:
        q["company_id"] = cid
    if extra:
        q.update(extra)
    return q


async def _salesman_customer_filter(ctx) -> Optional[list]:
    """If the logged-in user is a salesman, return the list of customer
    names they are mapped to for the current FY. Returns None for
    admin/super_admin/dispatch/employee (no role-based filter applied).

    SECURITY: every customer-listing endpoint accessible to salesmen MUST
    chain this filter into its query so a salesman can never enumerate
    customers they're not assigned to (iter98 security audit finding)."""
    if not ctx or not ctx.get("user"):
        return None
    user = ctx["user"]
    if user.get("role") != "salesman":
        return None
    from utils import get_current_fy as _gcf
    fy = _gcf()
    base = {}
    if ctx.get("tenant_id"):
        base["tenant_id"] = ctx["tenant_id"]
    if ctx.get("company_id"):
        base["company_id"] = ctx["company_id"]

    # Try exact-name match first, fall back to case-insensitive.
    sm_name = user.get("name", "")
    master = await db.salesman_master.find_one(
        {**base, "salesman_name": sm_name}, {"_id": 0}
    )
    if not master and sm_name:
        master = await db.salesman_master.find_one(
            {**base, "salesman_name": {"$regex": f"^{sm_name}$",
                                         "$options": "i"}}, {"_id": 0}
        )
    if not master:
        # Salesman has no mapping → expose nothing.
        return []
    fy_map = master.get("fy_customers", {}) or {}
    return list(fy_map.get(fy, master.get("customers", []) or []))


@router.get("/customers/outstanding")
async def get_customer_outstanding(
    request: Request,
    customer: Optional[str] = None,
    fy: Optional[str] = None,
    company_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 0,  # 0 = no pagination (legacy callers); set for mobile
):
    """Get outstanding payments by customer with proper aging, opening balance, credit notes, and journals."""
    try:
        from datetime import date as date_type
        today = date_type.today()
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        # SECURITY: salesman role must only see their mapped customers.
        # Apply this filter at the DB level so the salesman can never
        # observe non-mapped customers' outstanding/aging numbers.
        salesman_scope = await _salesman_customer_filter(ctx)
        if salesman_scope is not None:
            if not salesman_scope:
                # Mapped to no customers → return empty payload immediately.
                return APIResponse(success=True, data={
                    "customers": [], "total_outstanding": 0, "total_paid": 0,
                    "page": 1, "total": 0, "page_size": page_size or 0,
                })
            q["customer_name"] = {"$in": salesman_scope}

        synced_customers = await db.customers.find(q, {"_id": 0}).to_list(5000)

        # Determine branch exclusion from header
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        branch_parties = []
        if exclude_branches:
            branch_parties = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))

        # Filter branch customers from synced list
        if branch_parties:
            branch_set = set(p.lower() for p in branch_parties)
            synced_customers = [c for c in synced_customers if safe_str(c.get("customer_name")).lower() not in branch_set]

        # Fetch ALL vouchers (not FY filtered) for opening balance calculation
        all_sales = await db.sales_vouchers.find(q, {"_id": 0}).to_list(50000)
        all_receipts_raw = await db.receipt_vouchers.find(q, {"_id": 0}).to_list(50000)
        all_credit_notes = await db.credit_notes.find(q, {"_id": 0}).to_list(50000)
        all_journals = await db.journal_vouchers.find(q, {"_id": 0}).to_list(50000)

        # The receipt_vouchers collection holds receipts (CR party = reduces OS) AND
        # payment vouchers (DR party = increases OS, e.g., cheque-bounce refund).
        all_receipts, all_payments = split_receipts_and_payments(all_receipts_raw)

        # Filter branch parties from vouchers
        if branch_parties:
            bp_lower = set(p.lower().strip() for p in branch_parties)
            all_sales = filter_branch_parties(all_sales, bp_lower)
            all_receipts = filter_branch_parties(all_receipts, bp_lower)
            all_payments = filter_branch_parties(all_payments, bp_lower)
            all_credit_notes = filter_branch_parties(all_credit_notes, bp_lower)
            all_journals = filter_branch_parties(all_journals, bp_lower)

        # Compute FY boundaries
        fy_start_str = fy_start_iso(fy)

        # Split vouchers into pre-FY (opening) and current FY
        _, fy_sales = _split_by_fy(all_sales, fy_start_str)
        _, fy_receipts = _split_by_fy(all_receipts, fy_start_str)
        _, fy_payments = _split_by_fy(all_payments, fy_start_str)
        _, fy_cns = _split_by_fy(all_credit_notes, fy_start_str)
        _, fy_jvs = _split_by_fy(all_journals, fy_start_str)

        # FY-filter the current period
        if fy:
            fy_sales = filter_vouchers_by_fy(fy_sales, fy)
            fy_receipts = filter_vouchers_by_fy([{**r, 'voucher_date': r.get('voucher_date','')} for r in fy_receipts], fy)
            fy_payments = filter_vouchers_by_fy([{**p, 'voucher_date': p.get('voucher_date','')} for p in fy_payments], fy)
            fy_cns = filter_vouchers_by_fy(fy_cns, fy)
            fy_jvs = filter_vouchers_by_fy(fy_jvs, fy)

        # Compute opening balance per customer (Tally master OB anchored to today's
        # FY, replayed forward/backward for any other requested FY).
        base_fy_start = base_fy_start_iso(today)
        opening_balance = compute_opening_balance_map(
            synced_customers, all_sales, all_receipts, all_payments,
            all_credit_notes, all_journals, fy_start_str, base_fy_start,
        )

        # Compute current-FY credits per customer
        customer_receipts, customer_cn_total, customer_jv_adjustment, customer_jv_debit = \
            aggregate_party_credits(fy_receipts, fy_cns, fy_jvs)
        # Payment vouchers paid TO party → DR (increases OS). Aggregated separately
        # so the UI can show non-sales DR movements distinct from JV DR.
        customer_payments_dr = {}
        for pmt in fy_payments:
            p = safe_str(pmt.get("party_name")).strip().lower()
            if p:
                customer_payments_dr[p] = customer_payments_dr.get(p, 0) + safe_num(pmt.get("amount"))

        customer_map = {}
        customer_vouchers = {}

        # ── SOURCE OF TRUTH = customers collection (synced from Tally Sundry Debtors group)
        # Build customer_map ONLY from synced_customers. Do NOT auto-add parties
        # from sales_vouchers (those can leak creditors / non-customer ledgers).
        synced_name_lower_to_canonical = {}
        for sc in synced_customers:
            name = sc.get("customer_name")
            if not name:
                continue
            ob = round(opening_balance.get(name, 0), 2)
            customer_map[name] = {
                "customer_name": name,
                "ledger_group": sc.get("ledger_group", ""),
                "phone": sc.get("phone", "") or sc.get("mobile_number", ""),
                "contact_person": sc.get("contact_person", ""),
                "state": sc.get("state", ""),
                "opening_balance": ob,
                "tally_outstanding": safe_num(sc.get("outstanding_amount")),  # Tally closing balance
                "outstanding_amount": 0,
                "total_sales": 0,        # Total DR-side (sales + payment vouchers + JV DR)
                "sales_only": 0,         # Only sales vouchers
                "adjustment_dr": 0,      # Non-sales DR (payment vouchers + JV DR)
                "voucher_count": 0,
                "last_transaction": None,
                "aging_0_30": 0.0, "aging_30_60": 0.0,
                "aging_60_90": 0.0, "aging_90_plus": 0.0,
                "oldest_invoice_days": 0,
                # iter-140/141: enriched fields from Busy Agent v1.5.0 (and
                # equivalent Tally sync where present). Surfaced here so the
                # CRM UI can display contact / GST / salesman / address / group
                # / price-category / balance columns without needing extra
                # endpoints. Every field defaults to '' or 0 so pre-v1.5
                # customer docs still render cleanly.
                "customer_id": sc.get("customer_id", ""),
                "group_id": sc.get("group_id", ""),
                "group_name": sc.get("group_name", ""),
                "mobile_number": sc.get("mobile_number", "") or sc.get("phone", ""),
                "whatsapp_number": sc.get("whatsapp_number", ""),
                "email": sc.get("email", ""),
                "address": sc.get("address", ""),
                "address_line_1": sc.get("address_line_1", ""),
                "address_line_2": sc.get("address_line_2", ""),
                "address_line_3": sc.get("address_line_3", ""),
                "address_line_4": sc.get("address_line_4", ""),
                "city": sc.get("city", ""),
                "station": sc.get("station", ""),
                "pin_code": sc.get("pin_code", ""),
                "country": sc.get("country", ""),
                "gst_number": sc.get("gst_number", ""),
                "pan_number": sc.get("pan_number", ""),
                "salesman_id": sc.get("salesman_id", ""),
                "salesman_name": sc.get("salesman_name", ""),
                "salesman_mobile_number": sc.get("salesman_mobile_number", ""),
                "salesman_whatsapp_number": sc.get("salesman_whatsapp_number", ""),
                "price_category": sc.get("price_category", ""),
                "closing_balance": safe_num(sc.get("closing_balance", 0)),
                "balance": safe_num(sc.get("balance", sc.get("closing_balance", 0))),
            }
            customer_vouchers[name] = []
            synced_name_lower_to_canonical[name.lower().strip()] = name

        # Enrich synced customers with FY sales activity. Skip parties not in synced list.
        for voucher in fy_sales:
            party_raw = (voucher.get("party_name") or "").strip()
            if not party_raw:
                continue
            party = synced_name_lower_to_canonical.get(party_raw.lower())
            if not party:
                continue  # Not a synced customer → ignore (creditors, depots, etc.)
            amount = safe_num(voucher.get("total_amount"))
            v_date_str = voucher.get("voucher_date", "")

            customer_map[party]["total_sales"] += amount
            customer_map[party]["sales_only"] += amount
            customer_map[party]["voucher_count"] += 1
            if v_date_str and v_date_str > (customer_map[party].get("last_transaction") or ""):
                customer_map[party]["last_transaction"] = v_date_str
            try:
                parts = v_date_str.split("-")
                if len(parts) == 3:
                    v_date = date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                    days_old = (today - v_date).days
                    customer_vouchers[party].append({"amount": amount, "days_old": days_old})
                    if days_old > customer_map[party]["oldest_invoice_days"]:
                        customer_map[party]["oldest_invoice_days"] = days_old
            except (ValueError, TypeError):
                customer_vouchers[party].append({"amount": amount, "days_old": 0})

        # Enrich synced customers with FY payment-voucher activity (DR party — e.g.,
        # cheque-bounce refund, expense advance). Stored in receipt_vouchers with
        # voucher_type='payment'. Booked into total_sales (DR side) AND adjustment_dr
        # so the UI can show non-sales DR movements separately.
        for pmt in fy_payments:
            party_raw = (pmt.get("party_name") or "").strip()
            if not party_raw:
                continue
            party = synced_name_lower_to_canonical.get(party_raw.lower())
            if not party:
                continue
            amount = safe_num(pmt.get("amount"))
            v_date_str = pmt.get("voucher_date", "")
            customer_map[party]["total_sales"] += amount  # DR side
            customer_map[party]["adjustment_dr"] += amount
            customer_map[party]["voucher_count"] += 1
            if v_date_str and v_date_str > (customer_map[party].get("last_transaction") or ""):
                customer_map[party]["last_transaction"] = v_date_str
            try:
                parts = v_date_str.split("-")
                if len(parts) == 3:
                    v_date = date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                    days_old = (today - v_date).days
                    customer_vouchers[party].append({"amount": amount, "days_old": days_old})
                    if days_old > customer_map[party]["oldest_invoice_days"]:
                        customer_map[party]["oldest_invoice_days"] = days_old
            except (ValueError, TypeError):
                customer_vouchers[party].append({"amount": amount, "days_old": 0})

        customers = list(customer_map.values())

        # Source list is the customers collection (synced from Tally Sundry Debtors group)
        # No additional group filter needed — sync already handled it.

        if customer:
            customers = [c for c in customers if fuzzy_match(safe_str(c.get("customer_name")), customer)]

        # Build master-OB lookup for Tally-Verified badge logic (FY-1 comparison)
        synced_master_ob = {
            (sc.get("customer_name") or "").lower().strip(): safe_num(sc.get("opening_balance", 0))
            for sc in synced_customers if sc.get("customer_name")
        }

        # Calculate outstanding, paid, aging for each customer
        for cust in customers:
            party = cust["customer_name"]
            party_key = party.lower().strip()
            ob = cust.get("opening_balance", 0)
            total_sales = cust["total_sales"]
            receipt_paid = customer_receipts.get(party_key, 0)
            cn_credit = customer_cn_total.get(party_key, 0)
            jv_adjustment = customer_jv_adjustment.get(party_key, 0)  # Net: positive = reduces outstanding
            total_credits = receipt_paid + cn_credit + jv_adjustment

            # Outstanding = Opening + Sales - Receipts - Credit Notes - Net Journal Adjustments
            outstanding = ob + total_sales - total_credits
            cust["outstanding_amount"] = round(outstanding, 2)  # Allow negative (advance payments)
            cust["paid_amount"] = round(total_credits, 2)
            cust["receipt_count"] = len([r for r in fy_receipts if (r.get("party_name") or "").lower().strip() == party_key])
            cust["credit_note_total"] = round(cn_credit, 2)
            cust["journal_credit"] = round(jv_adjustment, 2)
            # Add JV DR-only into the visible Adjustment column (already includes payment vouchers)
            cust["adjustment_dr"] = round(cust.get("adjustment_dr", 0) + customer_jv_debit.get(party_key, 0), 2)

            # Tally-verified badge:
            # - For today's FY: compare computed OS to Tally master CB (customers.outstanding_amount)
            # - For previous FY (today_fy - 1): compare computed OS to Tally master OB
            #   (master OB = closing balance of previous FY by accounting identity)
            today_fy_year = today.year if today.month >= 4 else today.year - 1
            base_fy_start = f"{today_fy_year}-04-01"
            prev_fy_start = f"{today_fy_year - 1}-04-01"
            tally_master_ob = synced_master_ob.get(party_key, 0)
            cust["tally_master_ob"] = tally_master_ob
            os_val = cust["outstanding_amount"]
            verified = False
            if fy_start_str == base_fy_start or not fy_start_str:
                verified = abs(os_val - cust.get("tally_outstanding", 0)) < 1.0
            elif fy_start_str == prev_fy_start:
                verified = abs(os_val - tally_master_ob) < 1.0
            cust["tally_verified"] = bool(verified)

            # FIFO aging on outstanding
            voucher_list = customer_vouchers.get(party, [])
            aging, fallback_oldest = apply_fifo_aging(
                cust["outstanding_amount"], voucher_list, today, fy_start_str,
            )
            cust["aging_0_30"] = aging["aging_0_30"]
            cust["aging_30_60"] = aging["aging_30_60"]
            cust["aging_60_90"] = aging["aging_60_90"]
            cust["aging_90_plus"] = aging["aging_90_plus"]
            if fallback_oldest:
                cust["oldest_invoice_days"] = max(cust.get("oldest_invoice_days", 0), fallback_oldest)

            cust["overdue_amount"] = round(cust["aging_60_90"] + cust["aging_90_plus"], 2)

            cust["status"], cust["status_label"] = aging_status(
                cust["outstanding_amount"], cust["oldest_invoice_days"],
            )

        customers.sort(key=lambda c: c.get("customer_name", "").lower())

        all_groups = list(set(c.get("ledger_group", "") for c in customers if c.get("ledger_group")))
        all_states = list(set(c.get("state", "") for c in customers if c.get("state")))

        # Search + pagination — applied AFTER full computation so totals
        # (total_outstanding, total_paid) reflect the entire dataset, not just
        # the page. Mobile/CRM page hits this with page_size=50; web dashboard
        # totals need the un-paged sums.
        full_total_os = round(sum(c["outstanding_amount"] for c in customers), 2)
        full_total_paid = round(sum(c.get("paid_amount", 0) for c in customers), 2)
        if search and search.strip():
            # Fuzzy match — ignores spaces & separator chars in BOTH the search
            # term and the customer fields (so "abc co" finds "ABC & Co.", etc.)
            customers = [c for c in customers
                          if fuzzy_match(c.get("customer_name") or "", search)
                          or fuzzy_match(c.get("phone") or "", search)
                          or fuzzy_match(c.get("ledger_group") or "", search)
                          or fuzzy_match(c.get("contact_person") or "", search)
                          or fuzzy_match(c.get("state") or "", search)]
        total_after_search = len(customers)
        if page_size and page_size > 0:
            skip = max(0, (page - 1) * page_size)
            customers = customers[skip:skip + page_size]

        return APIResponse(
            success=True,
            data={
                "customers": customers,
                "total": total_after_search,
                "page": page if page_size else 1,
                "page_size": page_size if page_size else total_after_search,
                "total_outstanding": full_total_os,
                "total_paid": full_total_paid,
                "groups": sorted(all_groups),
                "states": sorted(all_states)
            }
        )

    except Exception as e:
        logger.error(f"Error fetching customer outstanding: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/customers/followups")
async def get_followups(request: Request, status: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        extra = {}
        if status:
            extra["status"] = status
        query = _build_query(ctx, company_id, extra)

        # SECURITY: salesman role only sees follow-ups for their mapped customers.
        salesman_scope = await _salesman_customer_filter(ctx)
        if salesman_scope is not None:
            if not salesman_scope:
                return APIResponse(success=True, data={"followups": [], "count": 0})
            query["customer_name"] = {"$in": salesman_scope}

        followups = await db.customer_followups.find(query, {"_id": 0}).sort("followup_date", -1).to_list(100)

        # Apply branch exclusion
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        if exclude_branches:
            branch_parties = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
            if branch_parties:
                branch_set = set(p.lower() for p in branch_parties)
                followups = [f for f in followups if f.get("customer_name", "").lower() not in branch_set]

        return APIResponse(success=True, data={"followups": followups, "count": len(followups)})
    except Exception as e:
        logger.error(f"Error fetching followups: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/followups")
async def create_followup(followup: CustomerFollowupCreate, request: Request):
    try:
        user = await get_current_user(request, db)
        ctx = await get_tenant_context(request)
        followup_obj = CustomerFollowup(
            customer_name=followup.customer_name,
            followup_date=datetime.fromisoformat(followup.followup_date),
            followup_type=followup.followup_type,
            status="pending",
            notes=followup.notes,
            created_by=user["username"] if user else "unknown",
            created_by_name=user.get("name", "") if user else "Unknown"
        )

        doc = followup_obj.model_dump()
        doc['followup_date'] = doc['followup_date'].isoformat()
        doc['created_at'] = doc['created_at'].isoformat()
        if ctx and ctx.get("tenant_id"):
            doc['tenant_id'] = ctx["tenant_id"]
        if ctx and ctx.get("company_id"):
            doc['company_id'] = ctx["company_id"]

        await db.customer_followups.insert_one(doc)

        return APIResponse(
            success=True,
            message="Follow-up created successfully",
            data={"id": followup_obj.id}
        )
    except Exception as e:
        logger.error(f"Error creating followup: {e}")
        return APIResponse(success=False, error=str(e))


@router.patch("/customers/followups/{followup_id}")
async def update_followup_status(followup_id: str, status: str, request: Request):
    try:
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        result = await db.customer_followups.update_one(
            {**tq, "id": followup_id},
            {"$set": {"status": status}}
        )
        return APIResponse(
            success=result.modified_count > 0,
            message="Follow-up updated" if result.modified_count > 0 else "Follow-up not found"
        )
    except Exception as e:
        logger.error(f"Error updating followup: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/customers/targets")
async def get_customer_targets(request: Request, fy: Optional[str] = None, company_id: Optional[str] = None):
    try:
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        # SECURITY: salesman role only sees targets for their mapped customers.
        salesman_scope = await _salesman_customer_filter(ctx)
        if salesman_scope is not None:
            if not salesman_scope:
                return APIResponse(success=True, data={"targets": [], "count": 0})
            q["customer_name"] = {"$in": salesman_scope}

        all_vouchers = await db.sales_vouchers.find(q, {"_id": 0}).to_list(10000)
        custom_targets = await db.customer_targets.find(q, {"_id": 0}).to_list(100)
        custom_target_map = {t["customer_name"]: t for t in custom_targets}

        # Get removed customers for this FY
        removed_docs = await db.customer_target_removals.find({**q, "fy": fy} if fy else q, {"_id": 0, "customer_name": 1}).to_list(5000)
        removed_set = {d["customer_name"] for d in removed_docs}

        # Apply branch exclusion
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        branch_parties = []
        if exclude_branches:
            branch_parties = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
        if branch_parties:
            all_vouchers = [v for v in all_vouchers if v.get("party_name") not in branch_parties]

        current_fy_vouchers = filter_vouchers_by_fy(all_vouchers, fy) if fy else all_vouchers

        prev_fy = get_previous_fy(fy) if fy else None
        prev_fy_vouchers = filter_vouchers_by_fy(all_vouchers, prev_fy) if prev_fy else []

        current_sales = {}
        current_monthly = {}
        for v in current_fy_vouchers:
            party = v.get("party_name", "Unknown")
            amount = safe_num(v.get("total_amount"))
            v_date = v.get("voucher_date", "")
            current_sales[party] = current_sales.get(party, 0) + amount
            if v_date:
                month_key = v_date[:7]
                if party not in current_monthly:
                    current_monthly[party] = {}
                current_monthly[party][month_key] = current_monthly[party].get(month_key, 0) + amount

        prev_sales = {}
        for v in prev_fy_vouchers:
            party = v.get("party_name", "Unknown")
            amount = safe_num(v.get("total_amount"))
            prev_sales[party] = prev_sales.get(party, 0) + amount

        all_customers = set(list(current_sales.keys()) + list(prev_sales.keys()))
        # Filter out removed customers + sentinel "Unknown" / blank parties
        all_customers = {
            c for c in all_customers
            if c and c.strip().lower() not in {"unknown", "n/a", "none", ""} and c not in removed_set
        }
        targets = []
        for cust in all_customers:
            ct = custom_target_map.get(cust, {})
            last_fy = prev_sales.get(cust, 0)
            achieved = current_sales.get(cust, 0)

            if cust in custom_target_map:
                target_amount = ct.get("target_amount", last_fy * 1.15)
                last_fy_override = ct.get("last_fy_sales", last_fy)
                if last_fy_override > 0:
                    last_fy = last_fy_override
            else:
                target_amount = last_fy * 1.15 if last_fy > 0 else achieved * 1.2 if achieved > 0 else 0

            monthly = current_monthly.get(cust, {})
            monthly_data = [{"month": k, "amount": v} for k, v in sorted(monthly.items())]

            targets.append({
                "customer_name": cust,
                "target_amount": round(target_amount, 2),
                "last_fy_sales": round(last_fy, 2),
                "achieved_amount": round(achieved, 2),
                "achievement_percentage": round((achieved / target_amount * 100), 1) if target_amount > 0 else 0,
                "remaining": round(max(0, target_amount - achieved), 2),
                "monthly_sales": monthly_data,
                "has_custom_target": cust in custom_target_map,
                "previous_fy": prev_fy or "",
                "current_fy": fy or ""
            })

        targets.sort(key=lambda x: x["customer_name"].lower())

        return APIResponse(
            success=True,
            data={"targets": targets, "current_fy": fy, "previous_fy": prev_fy, "removed_count": len(removed_set)}
        )

    except Exception as e:
        logger.error(f"Error fetching customer targets: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/targets/set")
async def set_customer_target(request: Request):
    try:
        from datetime import date as date_type
        body = await request.json()
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)

        customer_name = body.get("customer_name", "").strip()
        target_amount = body.get("target_amount", 0)
        last_fy_sales = body.get("last_fy_sales", 0)
        target_fy = body.get("fy", "")

        if not customer_name:
            return APIResponse(success=False, error="Customer name is required")

        if target_fy:
            fy_start, fy_end = fy_to_date_range(target_fy)
            if fy_end:
                end_parts = fy_end.split('-')
                fy_end_date = date_type(int(end_parts[0]), int(end_parts[1]), int(end_parts[2]))
                if date_type.today() > fy_end_date:
                    return APIResponse(success=False, error=f"FY {target_fy} has ended. Targets cannot be modified for completed financial years.")

        doc = {
            "customer_name": customer_name,
            "target_amount": float(target_amount),
            "last_fy_sales": float(last_fy_sales),
            "fy": target_fy,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        if ctx and ctx.get("tenant_id"):
            doc["tenant_id"] = ctx["tenant_id"]
        if ctx and ctx.get("company_id"):
            doc["company_id"] = ctx["company_id"]

        await db.customer_targets.update_one(
            {**tq, "customer_name": customer_name, "fy": target_fy},
            {"$set": doc},
            upsert=True
        )

        return APIResponse(
            success=True,
            message=f"Target set for {customer_name}",
            data=doc
        )
    except Exception as e:
        logger.error(f"Error setting customer target: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/targets/bulk-percentage")
async def bulk_set_target_percentage(request: Request):
    """Set target as a percentage of previous FY sales for multiple customers at once."""
    try:
        body = await request.json()
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        fy = body.get("fy", "")
        percentage = float(body.get("percentage", 115))  # default 115%
        customer_names = body.get("customer_names", [])  # empty = all

        if not fy:
            return APIResponse(success=False, error="FY is required")

        # Verify FY is current or future
        from datetime import date as date_type
        fy_start, fy_end = fy_to_date_range(fy)
        if fy_end:
            end_parts = fy_end.split('-')
            fy_end_date = date_type(int(end_parts[0]), int(end_parts[1]), int(end_parts[2]))
            if date_type.today() > fy_end_date:
                return APIResponse(success=False, error=f"FY {fy} has ended. Cannot modify targets.")

        # Get previous FY sales
        prev_fy = get_previous_fy(fy)
        all_vouchers = await db.sales_vouchers.find(tq, {"_id": 0}).to_list(10000)
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        if exclude_branches:
            branch_parties = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
            if branch_parties:
                all_vouchers = [v for v in all_vouchers if v.get("party_name") not in branch_parties]

        prev_vouchers = filter_vouchers_by_fy(all_vouchers, prev_fy) if prev_fy else []
        prev_sales = {}
        for v in prev_vouchers:
            party = v.get("party_name", "")
            amt = safe_num(v.get("total_amount"))
            prev_sales[party] = prev_sales.get(party, 0) + amt

        # Check for removed customers
        removed_docs = await db.customer_target_removals.find({**tq, "fy": fy}, {"_id": 0, "customer_name": 1}).to_list(5000)
        removed_set = {d["customer_name"] for d in removed_docs}

        updated = 0
        targets_to_process = customer_names if customer_names else list(prev_sales.keys())
        for cust in targets_to_process:
            if cust in removed_set:
                continue
            last_fy = prev_sales.get(cust, 0)
            if last_fy <= 0:
                continue
            target_amount = last_fy * (percentage / 100)
            doc = {
                "customer_name": cust,
                "target_amount": round(target_amount, 2),
                "last_fy_sales": round(last_fy, 2),
                "fy": fy,
                "target_percentage": percentage,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
            if ctx and ctx.get("tenant_id"):
                doc["tenant_id"] = ctx["tenant_id"]
            if ctx and ctx.get("company_id"):
                doc["company_id"] = ctx["company_id"]
            await db.customer_targets.update_one(
                {**tq, "customer_name": cust, "fy": fy},
                {"$set": doc},
                upsert=True
            )
            updated += 1

        return APIResponse(success=True, data={"updated": updated, "percentage": percentage, "message": f"Updated {updated} customer targets to {percentage}% of previous FY"})
    except Exception as e:
        logger.error(f"Bulk target error: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/targets/remove")
async def remove_customers_from_targets(request: Request):
    """Remove selected customers from target reports for current FY. Does not affect Tally data."""
    try:
        body = await request.json()
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        fy = body.get("fy", "")
        customer_names = body.get("customer_names", [])

        if not fy or not customer_names:
            return APIResponse(success=False, error="FY and customer_names required")

        count = 0
        for cust in customer_names:
            doc = {
                "customer_name": cust,
                "fy": fy,
                "removed_at": datetime.now(timezone.utc).isoformat(),
            }
            if ctx and ctx.get("tenant_id"):
                doc["tenant_id"] = ctx["tenant_id"]
            if ctx and ctx.get("company_id"):
                doc["company_id"] = ctx["company_id"]
            await db.customer_target_removals.update_one(
                {**tq, "customer_name": cust, "fy": fy},
                {"$set": doc},
                upsert=True
            )
            count += 1

        return APIResponse(success=True, data={"removed": count, "message": f"{count} customers removed from targets"})
    except Exception as e:
        logger.error(f"Remove targets error: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/targets/reactivate")
async def reactivate_customers_in_targets(request: Request):
    """Reactivate previously removed customers in target reports."""
    try:
        body = await request.json()
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        fy = body.get("fy", "")
        customer_names = body.get("customer_names", [])

        if not fy or not customer_names:
            return APIResponse(success=False, error="FY and customer_names required")

        result = await db.customer_target_removals.delete_many(
            {**tq, "fy": fy, "customer_name": {"$in": customer_names}}
        )

        return APIResponse(success=True, data={"reactivated": result.deleted_count})
    except Exception as e:
        logger.error(f"Reactivate targets error: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/customers/targets/removed")
async def get_removed_customers(request: Request, fy: Optional[str] = None):
    """Get list of customers removed from targets for a given FY."""
    try:
        ctx = await get_tenant_context(request)
        tq = _build_query(ctx)
        q = {**tq}
        if fy:
            q["fy"] = fy
        docs = await db.customer_target_removals.find(q, {"_id": 0}).to_list(5000)
        return APIResponse(success=True, data={"removed": docs})
    except Exception as e:
        logger.error(f"Get removed customers error: {e}")
        return APIResponse(success=False, error=str(e))



@router.post("/customers/ledger/export")
async def export_customer_ledger(request: Request):
    """Export complete customer ledger in Tally-style PDF format.
    Includes: Sales, Receipts, Credit Notes, and Journal entries."""
    try:
        body = await request.json()
        customer_name = body.get("customer_name", "")
        fy = body.get("fy", "")

        if not customer_name:
            return APIResponse(success=False, error="Customer name is required")

        ctx = await get_tenant_context(request)

        # SECURITY: salesman role can only export ledgers for their mapped customers.
        salesman_scope = await _salesman_customer_filter(ctx)
        if salesman_scope is not None and customer_name not in salesman_scope:
            return APIResponse(success=False, error="Access denied for this customer")

        tq = _build_query(ctx, body.get("company_id"))

        # Fetch all voucher types for this customer
        cust_q = {**tq, "party_name": customer_name}
        sales = await db.sales_vouchers.find(cust_q, {"_id": 0}).to_list(10000)
        receipts = await db.receipt_vouchers.find(cust_q, {"_id": 0}).to_list(10000)
        credit_notes = await db.credit_notes.find(cust_q, {"_id": 0}).to_list(10000)
        journals = await db.journal_vouchers.find(cust_q, {"_id": 0}).to_list(10000)

        # FY filter if provided
        if fy:
            sales = filter_vouchers_by_fy(sales, fy)
            receipts = filter_vouchers_by_fy(
                [{"voucher_date": r.get("voucher_date", ""), **r} for r in receipts], fy
            )
            credit_notes = filter_vouchers_by_fy(credit_notes, fy)
            journals = filter_vouchers_by_fy(journals, fy)

        # Build unified ledger entries (Tally format: Date | Particulars | Vch Type | Vch No. | Debit | Credit)
        entries = []
        for v in sales:
            entries.append({
                'date': v.get('voucher_date', ''),
                'particulars': f"Sales - {v.get('reference_number', v.get('voucher_id', ''))}",
                'vch_type': 'Sales',
                'vch_no': v.get('reference_number', v.get('voucher_id', '')),
                'debit': safe_num(v.get('total_amount')),
                'credit': 0.0,
                'narration': v.get('narration', '')
            })
        for r in receipts:
            entries.append({
                'date': r.get('voucher_date', ''),
                'particulars': f"Receipt - {r.get('voucher_id', '')}",
                'vch_type': r.get('voucher_type', 'Receipt').capitalize(),
                'vch_no': r.get('voucher_id', ''),
                'debit': 0.0,
                'credit': safe_num(r.get('amount')),
                'narration': r.get('narration', '')
            })
        for cn in credit_notes:
            entries.append({
                'date': cn.get('voucher_date', ''),
                'particulars': f"Credit Note - {cn.get('reference_number', cn.get('voucher_id', ''))}",
                'vch_type': 'Credit Note',
                'vch_no': cn.get('reference_number', cn.get('voucher_id', '')),
                'debit': 0.0,
                'credit': safe_num(cn.get('total_amount')),
                'narration': cn.get('narration', '')
            })
        for jv in journals:
            debit, credit = get_jv_party_amount(jv)
            entries.append({
                'date': jv.get('voucher_date', ''),
                'particulars': f"Journal - {jv.get('voucher_id', '')}",
                'vch_type': 'Journal',
                'vch_no': jv.get('voucher_id', ''),
                'debit': debit,
                'credit': credit,
                'narration': jv.get('narration', '')
            })

        if not entries and not fy:
            return APIResponse(success=False, error=f"No transactions found for {customer_name}")

        # Compute opening balance for this customer if FY is specified
        opening_balance = 0.0
        if fy:
            try:
                fy_parts = fy.split('-')
                start_year = int(fy_parts[0])
                fy_start = f"{start_year}-04-01"

                # Get Tally OB and base FY
                cust_doc = await db.customers.find_one({**tq, "customer_name": customer_name}, {"_id": 0, "opening_balance": 1})
                sync_doc = await db.sync_status.find_one({**tq, "type": "agent_sync"}, {"_id": 0, "financial_year": 1})
                base_fy = sync_doc.get("financial_year") if sync_doc else None
                base_fy_start = None
                if base_fy:
                    try:
                        base_fy_start = f"{int(base_fy.split('-')[0])}-04-01"
                    except Exception:
                        pass

                if cust_doc and base_fy_start:
                    opening_balance = safe_num(cust_doc.get("opening_balance", 0))
                    if fy_start != base_fy_start:
                        lo = min(fy_start, base_fy_start)
                        hi = max(fy_start, base_fy_start)
                        adj_sales = await db.sales_vouchers.find(
                            {**tq, "party_name": customer_name, "voucher_date": {"$gte": lo, "$lt": hi}}, {"_id": 0, "total_amount": 1}
                        ).to_list(10000)
                        adj_receipts = await db.receipt_vouchers.find(
                            {**tq, "party_name": customer_name, "voucher_date": {"$gte": lo, "$lt": hi}}, {"_id": 0, "amount": 1}
                        ).to_list(10000)
                        adj_cns = await db.credit_notes.find(
                            {**tq, "party_name": customer_name, "voucher_date": {"$gte": lo, "$lt": hi}}, {"_id": 0, "total_amount": 1}
                        ).to_list(10000)
                        adj_jvs = await db.journal_vouchers.find(
                            {**tq, "party_name": customer_name, "voucher_date": {"$gte": lo, "$lt": hi}},
                            {"_id": 0, "debit_amount": 1, "credit_amount": 1, "party_name": 1, "ledger_entries": 1}
                        ).to_list(10000)
                        if fy_start < base_fy_start:
                            opening_balance -= sum(safe_num(v.get('total_amount')) for v in adj_sales)
                            opening_balance += sum(safe_num(r.get('amount')) for r in adj_receipts)
                            opening_balance += sum(safe_num(cn.get('total_amount')) for cn in adj_cns)
                            for jv in adj_jvs:
                                d, c = get_jv_party_amount(jv)
                                opening_balance += c - d
                        else:
                            opening_balance += sum(safe_num(v.get('total_amount')) for v in adj_sales)
                            opening_balance -= sum(safe_num(r.get('amount')) for r in adj_receipts)
                            opening_balance -= sum(safe_num(cn.get('total_amount')) for cn in adj_cns)
                            for jv in adj_jvs:
                                d, c = get_jv_party_amount(jv)
                                opening_balance += d - c
            except Exception:
                pass

        # Sort by date
        entries.sort(key=lambda e: e['date'])

        # Get customer info
        cust_info = await db.customers.find_one({**tq, "customer_name": customer_name}, {"_id": 0})
        company_info = await db.sync_status.find_one({**tq, "type": "agent_sync"}, {"_id": 0})
        company_name_str = company_info.get('company_name', 'FLOWRA') if company_info else 'FLOWRA'

        from services.ledger_pdf_service import generate_tally_ledger_pdf
        output = generate_tally_ledger_pdf(
            customer_name=customer_name,
            company_name=company_name_str,
            entries=entries,
            fy=fy,
            customer_info=cust_info or {},
            opening_balance=opening_balance
        )

        filename = f"ledger_{customer_name.replace(' ', '_')}_{fy or 'all'}.pdf"
        user = await get_current_user(request, db)
        if user:
            await log_audit("data_export", user.get("username", ""), tenant_id=ctx.get("tenant_id", "") if ctx else "", company_id=ctx.get("company_id", "") if ctx else "", target=f"Ledger PDF: {customer_name}", ip_address=get_client_ip(request))
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Error exporting customer ledger: {e}")
        return APIResponse(success=False, error=str(e))


@router.get("/customers/payment-behavior")
async def get_payment_behavior(request: Request, customer: Optional[str] = None, fy: Optional[str] = None, company_id: Optional[str] = None):
    """Payment behavior analysis — filters by FY when provided. Includes opening balance."""
    try:
        from datetime import date as date_type
        today = date_type.today()
        ctx = await get_tenant_context(request)
        q = _build_query(ctx, company_id)

        # SECURITY: salesman role only sees payment behavior for their mapped customers.
        salesman_scope = await _salesman_customer_filter(ctx)
        if salesman_scope is not None:
            if not salesman_scope:
                return APIResponse(success=True, data={"customers": [], "count": 0})
            q["customer_name"] = {"$in": salesman_scope}

        # Fetch ALL vouchers first
        all_sales_raw = await db.sales_vouchers.find(q, {"_id": 0}).to_list(20000)
        all_receipts_raw_all = await db.receipt_vouchers.find(q, {"_id": 0}).to_list(20000)
        all_credit_notes_raw = await db.credit_notes.find(q, {"_id": 0}).to_list(20000)
        all_journals_raw = await db.journal_vouchers.find(q, {"_id": 0}).to_list(20000)

        # Split receipt_vouchers into actual receipts (CR party) vs payment vouchers (DR party)
        all_receipts_raw, all_payments_raw = split_receipts_and_payments(all_receipts_raw_all)

        # Apply branch exclusion
        exclude_branches = request.headers.get("X-Exclude-Branches", "").lower() == "true"
        branch_parties = []
        if exclude_branches:
            branch_parties = await get_branch_parties(ctx.get("tenant_id", ""), ctx.get("company_id", ""))
        if branch_parties:
            bp_lower = set(p.lower().strip() for p in branch_parties)
            all_sales_raw = filter_branch_parties(all_sales_raw, bp_lower)
            all_receipts_raw = filter_branch_parties(all_receipts_raw, bp_lower)
            all_payments_raw = filter_branch_parties(all_payments_raw, bp_lower)
            all_credit_notes_raw = filter_branch_parties(all_credit_notes_raw, bp_lower)
            all_journals_raw = filter_branch_parties(all_journals_raw, bp_lower)

        synced_customers = await db.customers.find(q, {"_id": 0}).to_list(5000)
        if branch_parties:
            branch_set = set(p.lower() for p in branch_parties)
            synced_customers = [c for c in synced_customers if safe_str(c.get("customer_name")).lower() not in branch_set]

        # FY boundary for opening balance calculation
        fy_start_str = fy_start_iso(fy)

        # Split vouchers into pre-FY (for opening balance) and current FY
        if fy:
            pre_sales, fy_sales_raw = _split_by_fy(all_sales_raw, fy_start_str)
            _, fy_receipts_raw = _split_by_fy(all_receipts_raw, fy_start_str)
            _, fy_payments_raw = _split_by_fy(all_payments_raw, fy_start_str)
            _, fy_cns_raw = _split_by_fy(all_credit_notes_raw, fy_start_str)
            _, fy_jvs_raw = _split_by_fy(all_journals_raw, fy_start_str)

            all_sales = filter_vouchers_by_fy(fy_sales_raw, fy)
            all_receipts = filter_vouchers_by_fy([{**r, 'voucher_date': r.get('voucher_date', '')} for r in fy_receipts_raw], fy)
            all_payments = filter_vouchers_by_fy([{**p, 'voucher_date': p.get('voucher_date', '')} for p in fy_payments_raw], fy)
            all_credit_notes = filter_vouchers_by_fy(fy_cns_raw, fy)
            all_journals = filter_vouchers_by_fy(fy_jvs_raw, fy)
        else:
            pre_sales = []
            all_sales = all_sales_raw
            all_receipts = all_receipts_raw
            all_payments = all_payments_raw
            all_credit_notes = all_credit_notes_raw
            all_journals = all_journals_raw

        # Compute opening balance per customer (anchored to TODAY's FY where Tally master OB lives)
        opening_balance_map = compute_opening_balance_map(
            synced_customers, all_sales_raw, all_receipts_raw, all_payments_raw,
            all_credit_notes_raw, all_journals_raw,
            fy_start_str, base_fy_start_iso(today),
        )

        # Build per-customer receipt totals (FY-filtered) — case-insensitive party keys
        customer_payments = {}
        customer_receipt_dates = {}
        for r in all_receipts:
            party = safe_str(r.get("party_name")).strip().lower()
            if not party or party == "unknown":
                continue
            customer_payments[party] = customer_payments.get(party, 0) + safe_num(r.get("amount"))
            customer_receipt_dates.setdefault(party, []).append(r.get("voucher_date", ""))

        # Build per-customer credit note totals (FY-filtered)
        customer_cn = {}
        for cn in all_credit_notes:
            party = safe_str(cn.get("party_name")).strip().lower()
            if party:
                customer_cn[party] = customer_cn.get(party, 0) + safe_num(cn.get("total_amount"))

        # Build per-customer journal NET adjustment (FY-filtered): positive = reduces outstanding
        customer_jv = {}
        for jv in all_journals:
            party = safe_str(jv.get("party_name")).strip().lower()
            if party:
                debit, credit = get_jv_party_amount(jv)
                net_credit = credit - debit  # positive reduces customer OS
                customer_jv[party] = customer_jv.get(party, 0) + net_credit

        # Restrict to synced customers only — prevents non-debtor parties (depots, expense
        # ledgers) from sneaking into Payment Behaviour
        synced_lower_to_canonical = {(sc.get("customer_name") or "").lower().strip(): sc.get("customer_name") for sc in synced_customers if sc.get("customer_name")}

        behavior_map = {}
        # Pre-seed behavior_map with all synced customers (so a customer with OB but no FY
        # activity still appears with their opening balance)
        for sc in synced_customers:
            name = sc.get("customer_name")
            if not name:
                continue
            ob = round(opening_balance_map.get(name, 0), 2)
            behavior_map[name] = {
                "customer_name": name,
                "ledger_group": sc.get("ledger_group", ""),
                "phone": sc.get("phone", ""),
                "state": sc.get("state", ""),
                "tally_outstanding": safe_num(sc.get("outstanding_amount")),
                "total_transactions": 0,
                "total_amount": 0,
                "opening_balance": ob,
                "average_transaction": 0,
                "outstanding_amount": 0,
                "paid_amount": 0,
                "credit_note_total": 0,
                "journal_credit": 0,
                "receipt_count": 0,
                "payment_ratio": 0,
                "payment_pattern": "regular",
                "average_payment_delay": 0,
                "credit_score": 0,
                "oldest_invoice_days": 0,
                "first_transaction": None,
                "last_transaction": None,
                "invoices": []
            }

        for voucher in all_sales:
            party_raw = voucher.get("party_name", "")
            if not party_raw or party_raw.strip().lower() in {"unknown", "n/a", "none", ""}:
                continue
            party = synced_lower_to_canonical.get(party_raw.lower().strip())
            if not party:
                continue  # Skip non-debtor parties (creditors, depots, etc.)
            amount = safe_num(voucher.get("total_amount"))
            v_date_str = voucher.get("voucher_date", "")

            behavior_map[party]["total_transactions"] += 1
            behavior_map[party]["total_amount"] += amount

            if v_date_str:
                if v_date_str < (behavior_map[party].get("first_transaction") or "9999"):
                    behavior_map[party]["first_transaction"] = v_date_str
                if v_date_str > (behavior_map[party].get("last_transaction") or ""):
                    behavior_map[party]["last_transaction"] = v_date_str

            try:
                parts = v_date_str.split("-")
                if len(parts) == 3:
                    v_date = date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                    days_old = (today - v_date).days
                    behavior_map[party]["invoices"].append({"amount": amount, "days_old": days_old, "date": v_date_str})
                    if days_old > behavior_map[party]["oldest_invoice_days"]:
                        behavior_map[party]["oldest_invoice_days"] = days_old
            except (ValueError, TypeError):
                pass

        # Add payment vouchers (DR party — e.g., cheque-bounce refunds) as DR-side activity
        # Restricted to synced customers so non-debtor payment vouchers don't leak in
        for pmt in all_payments:
            party_raw = (pmt.get("party_name") or "").strip()
            if not party_raw:
                continue
            party = synced_lower_to_canonical.get(party_raw.lower())
            if not party:
                continue
            amount = safe_num(pmt.get("amount"))
            v_date_str = pmt.get("voucher_date", "")
            behavior_map[party]["total_transactions"] += 1
            behavior_map[party]["total_amount"] += amount
            if v_date_str:
                if v_date_str < (behavior_map[party].get("first_transaction") or "9999"):
                    behavior_map[party]["first_transaction"] = v_date_str
                if v_date_str > (behavior_map[party].get("last_transaction") or ""):
                    behavior_map[party]["last_transaction"] = v_date_str

        # Item #5 fix: also include PRE-FY sales for invoice ageing (dates only, not amounts)
        # This way customers whose outstanding is entirely from previous FYs still get correct
        # oldest_invoice_days, average_payment_delay, etc.
        for voucher in pre_sales:
            party = voucher.get("party_name", "Unknown")
            if not party or party.strip().lower() in {"unknown", "n/a", "none", ""}:
                continue
            if party not in behavior_map:
                continue  # only enrich existing entries
            v_date_str = voucher.get("voucher_date", "")
            if not v_date_str:
                continue
            try:
                parts = v_date_str.split("-")
                if len(parts) == 3:
                    v_date = date_type(int(parts[0]), int(parts[1]), int(parts[2]))
                    days_old = (today - v_date).days
                    behavior_map[party]["invoices"].append({
                        "amount": safe_num(voucher.get("total_amount")),
                        "days_old": days_old, "date": v_date_str, "is_pre_fy": True,
                    })
                    if days_old > behavior_map[party]["oldest_invoice_days"]:
                        behavior_map[party]["oldest_invoice_days"] = days_old
                    if v_date_str < (behavior_map[party].get("first_transaction") or "9999"):
                        behavior_map[party]["first_transaction"] = v_date_str
            except (ValueError, TypeError):
                pass

        # All synced customers are already in behavior_map (pre-seeded earlier), so no
        # additional pass needed here.

        for party, data in behavior_map.items():
            party_key = party.lower().strip()
            total = data["total_amount"]
            ob = data.get("opening_balance", 0)
            receipt_paid = customer_payments.get(party_key, 0)
            cn_credit = customer_cn.get(party_key, 0)
            jv_credit = customer_jv.get(party_key, 0)
            total_credits = receipt_paid + cn_credit + jv_credit

            data["paid_amount"] = round(receipt_paid, 2)
            data["credit_note_total"] = round(cn_credit, 2)
            data["journal_credit"] = round(jv_credit, 2)
            data["receipt_count"] = len(customer_receipt_dates.get(party_key, []))
            # Outstanding = Opening Balance + Sales - Total Credits
            data["outstanding_amount"] = round(ob + total - total_credits, 2)

            # Total debits for ratio calculation = opening + sales
            total_debits = ob + total
            data["average_transaction"] = round(total / data["total_transactions"], 2) if data["total_transactions"] > 0 else 0
            data["payment_ratio"] = round((total_credits / total_debits * 100), 1) if total_debits > 0 else 100

            # Compute average payment delay using receipt-to-invoice matching
            if receipt_paid > 0 and data["invoices"]:
                receipt_dates = sorted([d for d in customer_receipt_dates.get(party_key, []) if d])
                invoice_dates = sorted([i["date"] for i in data["invoices"] if i.get("date")])
                # Estimate delay: avg gap between invoice dates and receipt dates
                if receipt_dates and invoice_dates:
                    delays = []
                    for rd in receipt_dates:
                        # Find closest preceding invoice
                        closest = None
                        for inv_d in invoice_dates:
                            if inv_d <= rd:
                                closest = inv_d
                        if closest:
                            try:
                                rd_parts = rd.split("-")
                                ci_parts = closest.split("-")
                                r_date = date_type(int(rd_parts[0]), int(rd_parts[1]), int(rd_parts[2]))
                                i_date = date_type(int(ci_parts[0]), int(ci_parts[1]), int(ci_parts[2]))
                                delays.append((r_date - i_date).days)
                            except:
                                pass
                    data["average_payment_delay"] = round(sum(delays) / len(delays), 0) if delays else 0
                else:
                    data["average_payment_delay"] = 0
            elif total > 0 and data["outstanding_amount"] > 0 and data["invoices"]:
                outstanding_ratio = data["outstanding_amount"] / total
                avg_invoice_age = sum(i["days_old"] for i in data["invoices"]) / len(data["invoices"])
                data["average_payment_delay"] = round(avg_invoice_age * outstanding_ratio, 0)
            else:
                data["average_payment_delay"] = 0

            # Item #5 fallback: if outstanding > 0 but no invoices found at all (pre-FY or FY),
            # use FY-start as ageing reference so the column isn't 0
            if data["outstanding_amount"] > 0 and data["oldest_invoice_days"] == 0 and fy_start_str:
                try:
                    fy_start_date = date_type(*[int(x) for x in fy_start_str.split("-")])
                    days_from_fy_start = (today - fy_start_date).days
                    data["oldest_invoice_days"] = days_from_fy_start
                except Exception:
                    pass

            # Credit score — cap payment_ratio input at 100
            capped_ratio = min(data["payment_ratio"], 100)
            payment_score = capped_ratio
            volume_score = min(20, data["total_transactions"] * 2)
            delay_penalty = min(20, data["average_payment_delay"] / 3) if data["outstanding_amount"] > 0 else 0
            data["credit_score"] = round(max(0, min(100, payment_score * 0.7 + volume_score - delay_penalty)), 1)

            # Pattern classification
            if data["payment_ratio"] >= 100:
                data["payment_pattern"] = "excellent"
            elif data["payment_ratio"] >= 90 and data["average_payment_delay"] < 15:
                data["payment_pattern"] = "excellent"
            elif data["payment_ratio"] >= 70 and data["average_payment_delay"] < 30:
                data["payment_pattern"] = "regular"
            elif data["payment_ratio"] >= 50:
                data["payment_pattern"] = "irregular"
            elif data["total_transactions"] == 0:
                data["payment_pattern"] = "no_transactions"
            else:
                data["payment_pattern"] = "risky"

            # Relationship duration
            if data.get("first_transaction") and data.get("last_transaction"):
                try:
                    ft = data["first_transaction"].split("-")
                    lt = data["last_transaction"].split("-")
                    f_date = date_type(int(ft[0]), int(ft[1]), int(ft[2]))
                    l_date = date_type(int(lt[0]), int(lt[1]), int(lt[2]))
                    data["relationship_months"] = max(1, round((l_date - f_date).days / 30))
                except:
                    data["relationship_months"] = 0
            else:
                data["relationship_months"] = 0

            # Build monthly payment timeline for the detail dropdown
            monthly_map = {}
            for inv in data.get("invoices", []):
                m = inv.get("date", "")[:7]  # YYYY-MM
                if m:
                    monthly_map.setdefault(m, {"invoiced": 0, "month": m})
                    monthly_map[m]["invoiced"] += inv["amount"]
            for rd in customer_receipt_dates.get(party_key, []):
                m = rd[:7] if rd else ""
                if m:
                    monthly_map.setdefault(m, {"invoiced": 0, "month": m})
                    # Distribute receipts proportionally — simple approach: total receipt / months
            # Add receipt amounts to monthly timeline
            if receipt_paid > 0 and monthly_map:
                receipt_months = sorted(set(rd[:7] for rd in customer_receipt_dates.get(party_key, []) if rd))
                per_month_receipt = receipt_paid / len(receipt_months) if receipt_months else 0
                for rm in receipt_months:
                    if rm in monthly_map:
                        monthly_map[rm]["received"] = monthly_map[rm].get("received", 0) + per_month_receipt
                    else:
                        monthly_map[rm] = {"invoiced": 0, "received": per_month_receipt, "month": rm}

            data["monthly_timeline"] = sorted(monthly_map.values(), key=lambda x: x["month"])[-12:]  # Last 12 months

            # Top invoices (largest, most overdue)
            top_invoices = sorted(data.get("invoices", []), key=lambda i: i.get("days_old", 0), reverse=True)[:5]
            data["top_overdue_invoices"] = top_invoices

            del data["invoices"]

        customers = list(behavior_map.values())

        # Filter: drop customers with no FY activity AND zero opening balance AND zero
        # Tally outstanding (these are dormant ledgers — no signal for payment behaviour)
        customers = [
            c for c in customers
            if c.get("total_transactions", 0) > 0
            or abs(c.get("opening_balance", 0)) > 0.5
            or abs(c.get("tally_outstanding", 0)) > 0.5
        ]

        if customer:
            customers = [c for c in customers if fuzzy_match(safe_str(c.get("customer_name")), customer)]

        customers.sort(key=lambda c: c.get("customer_name", "").lower())

        return APIResponse(success=True, data={
            "customers": customers,
            "summary": {
                "total_customers": len(customers),
                "excellent": len([c for c in customers if c["payment_pattern"] == "excellent"]),
                "regular": len([c for c in customers if c["payment_pattern"] == "regular"]),
                "irregular": len([c for c in customers if c["payment_pattern"] == "irregular"]),
                "risky": len([c for c in customers if c["payment_pattern"] == "risky"]),
                "avg_credit_score": round(sum(c["credit_score"] for c in customers) / len(customers), 1) if customers else 0
            }
        })

    except Exception as e:
        logger.error(f"Error analyzing payment behavior: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/outstanding/export")
async def export_outstanding_excel(request: Request):
    """Export outstanding data to Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    try:
        body = await request.json()
        data = body.get("data", [])
        fy = body.get("fy", "")

        ctx = await get_tenant_context(request)
        company_name = await _resolve_company_name(ctx)

        wb = Workbook()
        ws = wb.active
        ws.title = "Outstanding"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="2563EB")
        headers = ["Customer Name", "Group", "Opening Bal", "Total Sales", "Paid", "Outstanding", "0-30d", "30-60d", "60-90d", "90+d"]

        # iter-121: banner row with the useradmin's company name
        header_row = 1
        first_data_row = 2
        if company_name:
            banner = ws.cell(row=1, column=1, value=company_name)
            banner.font = Font(bold=True, color="0F1B4C", size=14)
            banner.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            subtitle = ws.cell(row=2, column=1, value=f"CRM Outstanding — FY {fy or 'All'}")
            subtitle.font = Font(italic=True, color="64748B", size=10)
            subtitle.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.row_dimensions[1].height = 22
            header_row = 3
            first_data_row = 4

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for offset, row in enumerate(data):
            i = first_data_row + offset
            ws.cell(row=i, column=1, value=row.get("customer_name", ""))
            ws.cell(row=i, column=2, value=row.get("ledger_group", ""))
            ws.cell(row=i, column=3, value=round(row.get("opening_balance", 0), 2))
            ws.cell(row=i, column=4, value=round(row.get("total_sales", 0), 2))
            ws.cell(row=i, column=5, value=round(row.get("paid_amount", 0), 2))
            ws.cell(row=i, column=6, value=round(row.get("outstanding_amount", 0), 2))
            aging = row.get("aging_buckets", {})
            ws.cell(row=i, column=7, value=round(aging.get("0_30", 0), 2))
            ws.cell(row=i, column=8, value=round(aging.get("30_60", 0), 2))
            ws.cell(row=i, column=9, value=round(aging.get("60_90", 0), 2))
            ws.cell(row=i, column=10, value=round(aging.get("90_plus", 0), 2))

        for col in range(1, 11):
            ws.column_dimensions[chr(64 + col)].width = 16
        ws.column_dimensions['A'].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"outstanding_{fy or 'all'}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        logger.error(f"Error exporting outstanding: {e}")
        return APIResponse(success=False, error=str(e))


@router.post("/customers/targets/export")
async def export_targets_excel(request: Request):
    """Export targets data to Excel."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    try:
        body = await request.json()
        data = body.get("data", [])
        fy = body.get("fy", "")

        ctx = await get_tenant_context(request)
        company_name = await _resolve_company_name(ctx)

        wb = Workbook()
        ws = wb.active
        ws.title = "Targets"

        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_fill = PatternFill("solid", fgColor="2563EB")
        headers = ["Customer Name", "Prev FY Sales", "Target", "Current FY Achieved", "Achievement %"]
        # Add monthly columns
        months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
        headers.extend(months)

        # iter-121: banner row with useradmin company name
        header_row = 1
        first_data_row = 2
        if company_name:
            banner = ws.cell(row=1, column=1, value=company_name)
            banner.font = Font(bold=True, color="0F1B4C", size=14)
            banner.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            subtitle = ws.cell(row=2, column=1, value=f"CRM Targets — FY {fy or 'All'}")
            subtitle.font = Font(italic=True, color="64748B", size=10)
            subtitle.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
            ws.row_dimensions[1].height = 22
            header_row = 3
            first_data_row = 4

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for offset, row in enumerate(data):
            i = first_data_row + offset
            ws.cell(row=i, column=1, value=row.get("customer_name", ""))
            ws.cell(row=i, column=2, value=round(row.get("previous_fy_sales", 0), 2))
            ws.cell(row=i, column=3, value=round(row.get("target", 0), 2))
            ws.cell(row=i, column=4, value=round(row.get("current_fy_sales", 0), 2))
            pct = row.get("achievement_pct", 0)
            ws.cell(row=i, column=5, value=f"{round(pct, 1)}%")
            monthly = row.get("monthly_sales", {})
            for j, m in enumerate(months):
                ws.cell(row=i, column=6 + j, value=round(monthly.get(m, 0), 2))

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = 14
        ws.column_dimensions['A'].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"targets_{fy or 'all'}.xlsx"
        return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={"Content-Disposition": f"attachment; filename={filename}"})
    except Exception as e:
        logger.error(f"Error exporting targets: {e}")
        return APIResponse(success=False, error=str(e))
