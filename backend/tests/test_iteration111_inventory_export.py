"""Iteration 111 — Inventory export bug fixes.

Pins the three regressions reported:
  Bug 1: stock_group multi-select sent only the first group.
  Bug 2: CSV export returned an empty file (TextIOWrapper GC bug).
  Bug 3: Excel button saved file with `.excel` extension (frontend);
         backend openpyxl also crashed on list-typed values ("aliases").

Also pins the user's additional ask: /reports/export must honour the same
filters as /inventory/items (category, stock_group CSV, root_stock_group,
abc, search).
"""
import inspect
import io

import pytest


# ── Bug 1 — multi-group filter ──────────────────────────────────────────
def test_inventory_items_accepts_csv_stock_group():
    """Backend must split a CSV `stock_group` value into a $in query."""
    from routes import inventory as inv
    src = inspect.getsource(inv.get_inventory_items)
    assert "stock_group.split(',')" in src or "stock_group.split(\",\")" in src
    assert '"$in": groups' in src or "'$in': groups" in src


# ── Bug 2 — CSV export not empty ────────────────────────────────────────
def test_export_to_csv_produces_non_empty_bytes():
    from services.export_service import ExportService
    data = [
        {"item_name": "Bolt M8", "qty": 25, "aliases": ["BLT-M8", "M8"]},
        {"item_name": "Bolt M10", "qty": 12, "aliases": []},
    ]
    buf = ExportService.export_to_csv(data)
    assert isinstance(buf, io.BytesIO)
    raw = buf.getvalue()
    assert len(raw) > 0, "CSV must not be empty (was the iter-111 regression)"
    text = raw.decode("utf-8-sig")
    # Header + 2 data rows
    assert "item_name,qty,aliases" in text
    assert "Bolt M8" in text and "Bolt M10" in text


def test_export_to_csv_handles_empty_data():
    from services.export_service import ExportService
    buf = ExportService.export_to_csv([])
    assert buf.getvalue() == b""


def test_export_to_csv_coerces_non_scalar_values():
    """List / dict values used to break the writer silently."""
    from services.export_service import ExportService
    data = [{"x": [1, 2, 3], "y": {"k": "v"}}]
    buf = ExportService.export_to_csv(data)
    assert b"[1, 2, 3]" in buf.getvalue() or b"['1', '2', '3']" in buf.getvalue()


# ── Bug 3 — Excel coerces list/dict values ──────────────────────────────
def test_export_to_excel_handles_list_cells():
    """openpyxl raised 'Cannot convert [] to Excel' on the aliases field —
    iter-111 coerces lists to comma-joined strings."""
    from services.export_service import ExportService
    from openpyxl import load_workbook
    data = [
        {"item_name": "Bolt M8", "qty": 25, "aliases": ["BLT-M8", "M8"]},
        {"item_name": "Bolt M10", "qty": 12, "aliases": []},
    ]
    buf = ExportService.export_to_excel(data, "Inventory")
    assert len(buf.getvalue()) > 0
    wb = load_workbook(buf)
    ws = wb.active
    # 1 header row + 2 data rows
    assert ws.max_row == 3
    # Aliases cell renders as a string, not a list literal
    aliases_col = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)].index("aliases") + 1
    assert ws.cell(row=2, column=aliases_col).value == "BLT-M8, M8"
    # openpyxl normalises empty strings to None on save/load round-trip.
    assert ws.cell(row=3, column=aliases_col).value in (None, "")


# ── Filtered export end-to-end contract ─────────────────────────────────
def test_reports_export_route_accepts_filters_block():
    from routes import ai_reports
    src = inspect.getsource(ai_reports.export_report)
    # New filters block exists
    assert "filters = body.get(\"filters\") or {}" in src
    # All five filter fields handled
    for f in ("category", "stock_group", "root_stock_group", "abc", "search"):
        assert f"filters.get(\"{f}\")" in src, f"filter {f!r} not wired into /reports/export"
    # stock_group CSV → $in
    assert '"$in": groups' in src or "'$in': groups" in src


# ── Bug 3 frontend contract (extension map) ─────────────────────────────
def test_frontend_extension_map_present():
    """Smoke-check the Inventory.js sends `format → real extension`."""
    src = open("/app/frontend/src/pages/Inventory.js").read()
    assert "extMap = { csv: 'csv', excel: 'xlsx', pdf: 'pdf' }" in src
    # And it passes the current filters along.
    assert "const filters = {" in src or "filters: {" in src
    assert "selectedGroups.join(',')" in src


def test_inventory_js_sends_csv_join_for_multi_group():
    """Multi-select group filter must be sent as a CSV string."""
    src = open("/app/frontend/src/pages/Inventory.js").read()
    # Confirm the old `selectedGroups[0]` bug is gone.
    assert "selectedGroups[0]" not in src
    assert "selectedGroups.join(',')" in src
