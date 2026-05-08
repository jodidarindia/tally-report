"""
Iteration 68 — Beat Run Monthly Report (admin/super_admin)

Tests:
- GET /api/salesman-orders/beat-run/monthly-report shape
- summary / per_salesman / per_customer / daily_breakdown / trend keys present
- Excel export returns a valid 4-sheet xlsx
- CSV export returns a valid CSV with header row
- Salesman-role users get 403 (admin-only)
- Bad month string returns clean error
"""
import os
import io
import requests
import pytest

API_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not API_URL:
    pytest.skip("REACT_APP_BACKEND_URL not set", allow_module_level=True)


def _login(username, password):
    r = requests.post(f"{API_URL}/api/auth/login", json={
        "username": username, "password": password, "captcha_token": "",
    })
    assert r.status_code == 200, r.text
    return r.json()["data"]["token"]


@pytest.fixture(scope="module")
def admin_h():
    return {"Authorization": f"Bearer {_login('admin', 'admin123')}"}


@pytest.fixture(scope="module")
def salesman_h():
    return {"Authorization": f"Bearer {_login('ravi@test.com', 'ravi1234')}"}


def test_monthly_report_shape(admin_h):
    r = requests.get(f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=2026-05",
                     headers=admin_h)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] is True
    d = body["data"]
    for key in ("summary", "per_salesman", "per_customer", "daily_breakdown", "trend"):
        assert key in d, f"missing {key}"
    s = d["summary"]
    for key in ("planned", "visited", "unplanned", "coverage_pct", "run_days", "salesmen_count", "month"):
        assert key in s
    assert s["month"] == "2026-05"


def test_monthly_report_trend_default_six(admin_h):
    r = requests.get(f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=2026-05",
                     headers=admin_h)
    trend = r.json()["data"]["trend"]
    assert len(trend) == 6
    # Oldest first → newest last; last entry must equal the requested month
    assert trend[-1]["month"] == "2026-05"
    # All entries shape-correct
    for t in trend:
        assert set(t.keys()) >= {"month", "coverage_pct", "planned", "visited", "unplanned"}


def test_monthly_report_bad_month(admin_h):
    r = requests.get(f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=garbage",
                     headers=admin_h)
    assert r.status_code == 200
    assert r.json()["success"] is False
    assert "month must be" in (r.json().get("error") or "").lower()


def test_monthly_report_salesman_blocked(salesman_h):
    """Salesman role MUST NOT see the admin monthly report."""
    r = requests.get(f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=2026-05",
                     headers=salesman_h)
    assert r.status_code == 200
    assert r.json()["success"] is False
    assert "admin" in (r.json().get("error") or "").lower()


def test_export_excel_4_sheets(admin_h):
    r = requests.get(
        f"{API_URL}/api/salesman-orders/beat-run/monthly-report/export?month=2026-05&format=excel",
        headers=admin_h,
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    assert wb.sheetnames == ["Summary", "By Salesman", "By Customer", "Raw Runs"]


def test_export_csv(admin_h):
    r = requests.get(
        f"{API_URL}/api/salesman-orders/beat-run/monthly-report/export?month=2026-05&format=csv",
        headers=admin_h,
    )
    assert r.status_code == 200
    assert "text/csv" in r.headers.get("content-type", "")
    body = r.content.decode("utf-8")
    # First line should be the header (or "No beat runs..." if empty)
    first = body.splitlines()[0]
    assert "Salesman" in first or "No beat runs" in first


def test_export_salesman_blocked(salesman_h):
    r = requests.get(
        f"{API_URL}/api/salesman-orders/beat-run/monthly-report/export?month=2026-05&format=csv",
        headers=salesman_h,
    )
    assert r.status_code == 200
    # Returns APIResponse JSON with admin-required error
    body = r.json()
    assert body["success"] is False
    assert "admin" in (body.get("error") or "").lower()


def test_monthly_report_filter_by_salesman(admin_h):
    r = requests.get(
        f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=2026-05&salesman=Ravi%20Kumar",
        headers=admin_h,
    )
    assert r.status_code == 200
    d = r.json()["data"]
    # Either 0 or 1 salesman row (only Ravi if any data exists)
    assert all(row["salesman"] == "Ravi Kumar" for row in d["per_salesman"])


def test_per_customer_aggregation_sorted_desc(admin_h):
    """Per-customer rows must be sorted by visit_count DESC."""
    r = requests.get(f"{API_URL}/api/salesman-orders/beat-run/monthly-report?month=2026-05",
                     headers=admin_h)
    pc = r.json()["data"]["per_customer"]
    counts = [c["visit_count"] for c in pc]
    assert counts == sorted(counts, reverse=True), "customers not sorted by visit_count desc"
    for c in pc:
        assert {"customer_name", "visit_count", "last_visit_date", "salesmen", "unplanned"} <= set(c.keys())


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
