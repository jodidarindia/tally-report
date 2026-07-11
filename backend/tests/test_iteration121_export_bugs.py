"""Iteration 121 — Export regression suite covering 5 reported bugs.

Bugs fixed:
  1. Sales tab PDF/Excel export → "Not Found" (frontend hit non-existent
     GET /api/export/sales; switched to POST /api/reports/export).
  2. CRM Target Excel exported a corrupt file (field-name mismatch AND
     JSON error body served as blob → Excel refused to open).
  3. CRM Target "All Customers" dropdown filter was ignored (only
     Outstanding tab honoured it).
  4. CRM Payment Behavior column alignment broken (data-table th.numeric
     had no CSS rule → left-aligned headers over right-aligned cells;
     Pay Ratio flex was also not justify-end).
  5. Inventory PDF header rendered "FLOWRA Report" / "Anonymous" instead
     of the useradmin's synced company name.

All 5 exports (Sales/Inventory/Targets/Outstanding + CRM cases) now
stamp the useradmin's company name as the banner header.
"""
import os
import sys
sys.path.insert(0, "/app/backend")

import asyncio
import io
import pytest
from openpyxl import load_workbook
from pypdf import PdfReader
import httpx

API = os.environ.get("API_URL", "https://tally-report-ai.preview.emergentagent.com").rstrip("/") + "/api"
USERNAME = "admin"
PASSWORD = "admin123"
EXPECTED_COMPANY_NAME = "ASA AUTOTECH INDIA PRIVATE LIMITED"


def _login():
    r = httpx.post(f"{API}/auth/login", json={"username": USERNAME, "password": PASSWORD}, timeout=30)
    r.raise_for_status()
    return r.json()["data"]["token"]


@pytest.fixture(scope="module")
def auth_headers():
    return {"Authorization": f"Bearer {_login()}"}


def _first_row(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    return ws.cell(row=1, column=1).value


def _pdf_first_page(pdf_bytes):
    r = PdfReader(io.BytesIO(pdf_bytes))
    return r.pages[0].extract_text()


# ── Bug #1: Sales PDF / Excel via /reports/export ─────────────────
def test_sales_pdf_export_no_longer_404(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "sales", "format": "pdf",
        "filters": {"fy": "2026-27"}, "fy": "2026-27",
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith("application/pdf")
    assert r.content.startswith(b"%PDF")


def test_sales_excel_export_via_reports(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "sales", "format": "excel",
        "filters": {"fy": "2026-27"}, "fy": "2026-27",
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"  # xlsx magic


# ── Bug #5 (all exports show company_name) ────────────────────────
def test_sales_pdf_shows_company_name_banner(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "sales", "format": "pdf",
        "filters": {"fy": "2026-27"}, "fy": "2026-27",
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    text = _pdf_first_page(r.content)
    assert EXPECTED_COMPANY_NAME in text, (
        f"Expected '{EXPECTED_COMPANY_NAME}' in Sales PDF; got: {text[:200]!r}"
    )


def test_inventory_pdf_shows_company_name_banner(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "inventory", "format": "pdf", "filters": {},
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    text = _pdf_first_page(r.content)
    assert EXPECTED_COMPANY_NAME in text, (
        f"Expected '{EXPECTED_COMPANY_NAME}' in Inventory PDF; got: {text[:200]!r}"
    )


def test_inventory_xlsx_first_row_is_company_name(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "inventory", "format": "excel", "filters": {},
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    assert _first_row(r.content) == EXPECTED_COMPANY_NAME


def test_sales_xlsx_first_row_is_company_name(auth_headers):
    r = httpx.post(f"{API}/reports/export", json={
        "report_type": "sales", "format": "excel",
        "filters": {"fy": "2026-27"}, "fy": "2026-27",
    }, headers=auth_headers, timeout=60)
    assert r.status_code == 200
    assert _first_row(r.content) == EXPECTED_COMPANY_NAME


# ── Bug #2: Targets Excel returns valid xlsx with banner ──────────
def test_targets_excel_returns_valid_xlsx_with_banner(auth_headers):
    r = httpx.post(f"{API}/customers/targets/export", json={
        "data": [{
            "customer_name": "Test A",
            "previous_fy_sales": 100000, "target": 150000,
            "current_fy_sales": 75000, "achievement_pct": 50,
        }],
        "fy": "2026-27",
    }, headers=auth_headers, timeout=30)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.content[:2] == b"PK"
    assert _first_row(r.content) == EXPECTED_COMPANY_NAME


def test_outstanding_excel_returns_valid_xlsx_with_banner(auth_headers):
    r = httpx.post(f"{API}/customers/outstanding/export", json={
        "data": [{"customer_name": "Test A", "ledger_group": "Debtors",
                  "outstanding_amount": 50000}],
        "fy": "2026-27",
    }, headers=auth_headers, timeout=30)
    assert r.status_code == 200
    assert _first_row(r.content) == EXPECTED_COMPANY_NAME


# ── Bug #3 & #4: static source assertions ─────────────────────────
CRM_PATH = "/app/frontend/src/pages/CustomerCRM.js"
CSS_PATH = "/app/frontend/src/App.css"


def test_targets_tab_applies_group_filter():
    src = open(CRM_PATH).read()
    # The IIFE-based targets block must reference selectedGroup filter
    assert "activeTab === 'targets' && (() =>" in src, (
        "Targets tab must be wrapped in IIFE so we can compute filteredTargets."
    )
    assert "filteredTargets" in src and "selectedGroup" in src, (
        "Targets tab must apply selectedGroup dropdown filter."
    )
    assert "exportTargetsExcel(sortedTargets)" in src, (
        "Export Excel button on Targets tab must pass FILTERED rows, "
        "not raw `targets`."
    )


def test_payment_behavior_pay_ratio_uses_justify_end():
    src = open(CRM_PATH).read()
    assert 'flex items-center gap-1 justify-end' in src, (
        "Pay Ratio cell flex needs justify-end so the bar + % align to "
        "the right edge of the (right-aligned) numeric column."
    )


def test_data_table_th_numeric_css_rule_exists():
    css = open(CSS_PATH).read()
    assert ".data-table th.numeric" in css, (
        "Missing CSS rule for right-aligning numeric headers — this was "
        "the root cause of the Payment Behavior header/data misalignment."
    )
    assert "text-align: right" in css.split(".data-table th.numeric")[1][:200], (
        ".data-table th.numeric must set text-align: right"
    )


if __name__ == "__main__":
    # Run standalone (no pytest)
    headers = {"Authorization": f"Bearer {_login()}"}
    print("Running iteration-121 export regression suite...")
    tests = [
        ("sales_pdf_no_404", test_sales_pdf_export_no_longer_404),
        ("sales_excel", test_sales_excel_export_via_reports),
        ("sales_pdf_banner", test_sales_pdf_shows_company_name_banner),
        ("inventory_pdf_banner", test_inventory_pdf_shows_company_name_banner),
        ("inventory_xlsx_banner", test_inventory_xlsx_first_row_is_company_name),
        ("sales_xlsx_banner", test_sales_xlsx_first_row_is_company_name),
        ("targets_excel_valid", test_targets_excel_returns_valid_xlsx_with_banner),
        ("outstanding_excel_valid", test_outstanding_excel_returns_valid_xlsx_with_banner),
    ]
    for name, fn in tests:
        try:
            fn(headers)
            print(f"  ✅ {name}")
        except Exception as e:
            print(f"  ❌ {name}: {e}")
    # Static
    for name, fn in [
        ("targets_filter_wired", test_targets_tab_applies_group_filter),
        ("pay_ratio_justify_end", test_payment_behavior_pay_ratio_uses_justify_end),
        ("th_numeric_css", test_data_table_th_numeric_css_rule_exists),
    ]:
        try:
            fn()
            print(f"  ✅ {name}")
        except Exception as e:
            print(f"  ❌ {name}: {e}")
