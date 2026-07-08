"""Iteration 116 · FLOWRA financial pitch model.

Locks in the founder-approved projection numbers so a later edit to
`generate_financial_pitch.py` can't silently shift the Seed ask, ARR
trajectory, or cap-table math.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "scripts"))

import generate_financial_pitch as m  # noqa: E402


def test_pricing_locked_to_founder_picks():
    p = m.ASSUMPTIONS["prices"]
    assert p["starter"]      == 833
    assert p["professional"] == 2083
    assert p["enterprise"]   == 3166


def test_customers_end_of_year_locked():
    c = m.ASSUMPTIONS["customers_end_of_year"]
    assert c["Y1_FY26_27"] == 100
    assert c["Y5_FY30_31"] == 2000
    assert c["Y6_FY31_32"] == 2700


def test_seed_and_series_a_ask():
    assert m.ASSUMPTIONS["seed_amount_cr"]     == 2.5
    assert m.ASSUMPTIONS["series_a_amount_cr"] == 6.0
    assert m.ASSUMPTIONS["seed_dilution_pct"]  == 18
    assert m.ASSUMPTIONS["series_a_dilution_pct"] == 18


def test_arr_hits_5cr_by_y5():
    proj = m.compute_projections()
    y5 = proj[4]  # index 4 = Y5 (FY30-31)
    assert y5["label"] == "FY30-31"
    assert 4.9 <= y5["arr_end_cr"] <= 5.2, y5["arr_end_cr"]
    assert y5["customers_end"] == 2000


def test_ebitda_flips_positive_by_y5():
    """Business milestone — Y5 must be EBITDA-positive for the pitch to hold."""
    proj = m.compute_projections()
    ebitdas = [p["ebitda_cr"] for p in proj]
    # Y1-Y4 all negative (burn), Y5 positive, Y6 more positive
    assert all(e < 0 for e in ebitdas[:4]), (
        f"Y1-Y4 should burn; got {ebitdas[:4]}"
    )
    assert ebitdas[4] > 0, f"Y5 EBITDA must be positive; got {ebitdas[4]}"
    assert ebitdas[5] > ebitdas[4], (
        f"Y6 EBITDA must exceed Y5; got {ebitdas[5]} vs {ebitdas[4]}"
    )


def test_cumulative_burn_within_raise():
    """The Seed + Series A together must cover Y1-Y5 net burn."""
    proj = m.compute_projections()
    cumulative_burn = sum(p["ebitda_cr"] for p in proj[:5])
    total_raise = (m.ASSUMPTIONS["seed_amount_cr"] +
                    m.ASSUMPTIONS["series_a_amount_cr"])
    # cumulative_burn is negative; we need |burn| ≤ total_raise
    assert abs(cumulative_burn) <= total_raise, (
        f"Cumulative burn ₹{-cumulative_burn:.2f} Cr exceeds "
        f"total raise ₹{total_raise:.2f} Cr — under-funded"
    )


def test_unit_economics_healthy():
    ue = m.compute_unit_economics()
    assert ue["ltv_to_cac"] >= 3.0, f"LTV:CAC below benchmark: {ue['ltv_to_cac']}"
    assert ue["payback_months"] <= 18, (
        f"Payback exceeds industry norm: {ue['payback_months']} mo"
    )
    assert 78 <= ue["gross_margin_pct"] <= 85, ue["gross_margin_pct"]


def test_deliverables_exist_and_nonzero():
    assert m.PDF_PATH.exists(), f"PDF missing: {m.PDF_PATH}"
    assert m.XLSX_PATH.exists(), f"XLSX missing: {m.XLSX_PATH}"
    # Sanity: PDF > 50 KB (thin PDFs indicate render failure)
    assert m.PDF_PATH.stat().st_size > 50_000
    # Sanity: XLSX > 5 KB (empty workbooks are ~2 KB)
    assert m.XLSX_PATH.stat().st_size > 5_000


def test_xlsx_has_all_seven_sheets_plus_readme():
    from openpyxl import load_workbook
    wb = load_workbook(m.XLSX_PATH)
    expected = ["0. Read Me", "1. Assumptions", "2. P&L Summary",
                "3. Revenue Build", "4. Cash Flow & Fundraise",
                "5. Unit Economics", "6. Cap Table", "7. Exit & Returns"]
    assert wb.sheetnames == expected, wb.sheetnames
