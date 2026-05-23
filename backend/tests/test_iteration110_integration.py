"""Iteration 110 — LIVE integration tests against the deployed backend.

Covers:
  • Feature 1   — Employee toggle-active + login block + email-reserved
  • Feature 2   — Salesman /copy-from validations + tenant isolation
  • Feature 3   — Beat-Run check-in mandatory order/payment + close-day + reopen
  • Feature 4   — Add-unplanned (existing customer) + /my-customers
  • Feature 5a  — Day-report export (PDF / Excel)
  • Feature 5b  — Monthly report Excel new columns (openpyxl introspection)
"""
import io
import os
import uuid
import pytest
import requests

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "").rstrip("/")
if not BASE_URL:
    # Fallback to frontend/.env value
    try:
        with open("/app/frontend/.env") as f:
            for line in f:
                if line.startswith("REACT_APP_BACKEND_URL="):
                    BASE_URL = line.split("=", 1)[1].strip().rstrip("/")
    except Exception:
        pass

ADMIN_USER = "admin"
ADMIN_PASS = "admin123"
SALESMAN_USER = "ravi@test.com"
SALESMAN_PASS = "ravi1234"


# ── helpers ──────────────────────────────────────────────────────────────
def _login(username: str, password: str) -> str:
    r = requests.post(
        f"{BASE_URL}/api/auth/login",
        json={"username": username, "password": password},
        timeout=60,
    )
    assert r.status_code == 200, f"login {username} failed: {r.status_code} {r.text}"
    body = r.json()
    if isinstance(body, dict) and body.get("success") is False:
        pytest.skip(f"login refused for {username}: {body.get('error')}")
    data = body.get("data") or {}
    tok = (body.get("access_token") or body.get("token")
           or data.get("token") or data.get("access_token"))
    assert tok, f"no token in login response: {body}"
    return tok


def _h(tok: str) -> dict:
    return {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}


@pytest.fixture(scope="module")
def admin_token():
    return _login(ADMIN_USER, ADMIN_PASS)


@pytest.fixture(scope="module")
def salesman_token():
    return _login(SALESMAN_USER, SALESMAN_PASS)


# ════════════════════════════════════════════════════════════════════════
# Feature 1 — Employee active toggle
# ════════════════════════════════════════════════════════════════════════
class TestEmployeeToggle:
    def test_admin_self_toggle_blocked(self, admin_token):
        r = requests.put(
            f"{BASE_URL}/api/auth/users/{ADMIN_USER}/toggle-active",
            headers=_h(admin_token), timeout=15,
        )
        body = r.json()
        # Either 400/403 or success=False with proper error
        assert r.status_code in (200, 400, 403)
        if r.status_code == 200:
            assert body.get("success") is False
            assert "your own" in (body.get("error") or "").lower() or "cannot toggle" in (body.get("error") or "").lower()

    def test_non_admin_blocked(self, salesman_token):
        r = requests.put(
            f"{BASE_URL}/api/auth/users/{ADMIN_USER}/toggle-active",
            headers=_h(salesman_token), timeout=15,
        )
        # Could be 401/403 OR APIResponse with success=False
        if r.status_code == 200:
            assert r.json().get("success") is False
        else:
            assert r.status_code in (401, 403)

    def test_toggle_creates_then_blocks_login(self, admin_token):
        """Create a TEST employee, toggle off, verify login blocked, then re-enable."""
        uname = f"test_emp_{uuid.uuid4().hex[:6]}@example.com"
        pwd = "TempPass#123"

        # Create employee via /api/auth/users (admin-only)
        create = requests.post(
            f"{BASE_URL}/api/auth/users",
            headers=_h(admin_token),
            json={"username": uname, "password": pwd, "name": "Test Emp", "role": "dispatch"},
            timeout=20,
        )
        if create.status_code != 200 or create.json().get("success") is False:
            # Try alternative employees endpoint
            create = requests.post(
                f"{BASE_URL}/api/employees/create",
                headers=_h(admin_token),
                json={"username": uname, "password": pwd, "name": "Test Emp", "role": "dispatch"},
                timeout=20,
            )
        assert create.status_code == 200, f"create employee failed: {create.text}"
        cb = create.json()
        assert cb.get("success") is not False, f"create employee err: {cb}"

        try:
            # Confirm login works pre-toggle
            ok_login = requests.post(
                f"{BASE_URL}/api/auth/login",
                json={"username": uname, "password": pwd}, timeout=15,
            )
            assert ok_login.status_code == 200 and (ok_login.json().get("data") or {}).get("token"), \
                f"pre-toggle login failed: {ok_login.text}"

            # Toggle OFF
            t1 = requests.put(
                f"{BASE_URL}/api/auth/users/{uname}/toggle-active",
                headers=_h(admin_token), timeout=15,
            )
            assert t1.status_code == 200, t1.text
            tb = t1.json()
            assert tb.get("success") is True, f"toggle failed: {tb}"

            # Login now should be blocked
            blocked = requests.post(
                f"{BASE_URL}/api/auth/login",
                json={"username": uname, "password": pwd}, timeout=15,
            )
            # Expect 401/403 OR 200 with success=False / detail with "deactivated"
            blocked_body = blocked.json() if blocked.headers.get("content-type", "").startswith("application/json") else {}
            txt = (blocked.text + " " + (blocked_body.get("error") or "") + " " + (blocked_body.get("message") or "")).lower()
            assert "deactivated" in txt or blocked.status_code in (401, 403) or blocked_body.get("success") is False, \
                f"deactivated login should fail: {blocked.status_code} {blocked.text}"

            # Re-creating with same email must be blocked
            recreate = requests.post(
                f"{BASE_URL}/api/auth/users",
                headers=_h(admin_token),
                json={"username": uname, "password": pwd, "name": "Test Emp", "role": "dispatch"},
                timeout=15,
            )
            rb = recreate.json() if recreate.status_code == 200 else {"error": recreate.text}
            err = (rb.get("error") or rb.get("detail") or "").lower()
            assert "already" in err or "registered" in err or "exists" in err, \
                f"recreate should be blocked: {recreate.status_code} {rb}"

            # Toggle ON
            t2 = requests.put(
                f"{BASE_URL}/api/auth/users/{uname}/toggle-active",
                headers=_h(admin_token), timeout=15,
            )
            assert t2.json().get("success") is True

            # Login should work again
            ok2 = requests.post(
                f"{BASE_URL}/api/auth/login",
                json={"username": uname, "password": pwd}, timeout=15,
            )
            assert ok2.status_code == 200 and (ok2.json().get("data") or {}).get("token"), \
                f"re-activated login failed: {ok2.text}"
        finally:
            # Cleanup — best effort delete
            for path in (f"/api/auth/users/{uname}", f"/api/employees/{uname}"):
                try:
                    requests.delete(f"{BASE_URL}{path}", headers=_h(admin_token), timeout=10)
                except Exception:
                    pass


# ════════════════════════════════════════════════════════════════════════
# Feature 2 — Salesman /copy-from
# ════════════════════════════════════════════════════════════════════════
class TestSalesmanCopy:
    def test_copy_validation_same_source_target(self, admin_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman/copy-from",
            headers=_h(admin_token),
            json={"from_salesman": "X", "to_salesman": "X", "copy_customers": True},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "differ" in (body.get("error") or "").lower()

    def test_copy_validation_no_action_selected(self, admin_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman/copy-from",
            headers=_h(admin_token),
            json={"from_salesman": "Ravi Kumar", "to_salesman": "Other",
                  "copy_customers": False, "copy_beats": False},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "select at least one" in (body.get("error") or "").lower()

    def test_copy_validation_source_not_found(self, admin_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman/copy-from",
            headers=_h(admin_token),
            json={"from_salesman": f"NotExist_{uuid.uuid4().hex[:5]}",
                  "to_salesman": "Other", "copy_customers": True},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "not found" in (body.get("error") or "").lower()

    def test_copy_blocked_for_non_admin(self, salesman_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman/copy-from",
            headers=_h(salesman_token),
            json={"from_salesman": "A", "to_salesman": "B", "copy_customers": True},
            timeout=15,
        )
        if r.status_code == 200:
            assert r.json().get("success") is False
        else:
            assert r.status_code in (401, 403)


# ════════════════════════════════════════════════════════════════════════
# Feature 3 + 4 — Beat-Run check-in/unplanned/close-day/reopen
# ════════════════════════════════════════════════════════════════════════
class TestBeatRun:
    def test_my_customers_endpoint(self, salesman_token):
        r = requests.get(
            f"{BASE_URL}/api/salesman-orders/my-customers",
            headers=_h(salesman_token), timeout=15,
        )
        assert r.status_code == 200
        body = r.json()
        assert body.get("success") is True
        data = body.get("data") or {}
        assert "customers" in data
        assert isinstance(data["customers"], list)

    def test_check_in_missing_order_collected_rejected(self, salesman_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman-orders/beat-run/check-in",
            headers=_h(salesman_token),
            json={"customer_name": "Ankit Automobiles, Indore", "visited": True},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "order_collected" in (body.get("error") or "")

    def test_check_in_missing_payment_collected_rejected(self, salesman_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman-orders/beat-run/check-in",
            headers=_h(salesman_token),
            json={"customer_name": "Ankit Automobiles, Indore",
                  "visited": True, "order_collected": True},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "payment_collected" in (body.get("error") or "")

    def test_check_in_unvisit_does_not_require_flags(self, salesman_token):
        # Unchecking a visit should NOT require Yes/No
        r = requests.post(
            f"{BASE_URL}/api/salesman-orders/beat-run/check-in",
            headers=_h(salesman_token),
            json={"customer_name": "Ankit Automobiles, Indore", "visited": False},
            timeout=15,
        )
        body = r.json()
        # success or "no run/closed" both acceptable — what we forbid is the
        # mandatory-flag error.
        assert "order_collected" not in (body.get("error") or "")
        assert "payment_collected" not in (body.get("error") or "")

    def test_add_unplanned_requires_order_payment(self, salesman_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman-orders/beat-run/add-unplanned",
            headers=_h(salesman_token),
            json={"customer_name": "Random Prospect", "details": "drive-by"},
            timeout=15,
        )
        body = r.json()
        assert body.get("success") is False
        assert "order_collected" in (body.get("error") or "") or \
               "payment_collected" in (body.get("error") or "")

    def test_reopen_day_blocked_for_salesman(self, salesman_token):
        r = requests.post(
            f"{BASE_URL}/api/salesman-orders/beat-run/reopen-day",
            headers=_h(salesman_token),
            json={"salesman": "Ravi Kumar"},
            timeout=15,
        )
        body = r.json() if r.status_code == 200 else {}
        if r.status_code == 200:
            assert body.get("success") is False
            assert "admin" in (body.get("error") or "").lower()
        else:
            assert r.status_code in (401, 403)


# ════════════════════════════════════════════════════════════════════════
# Feature 5 — Day report + Monthly report exports
# ════════════════════════════════════════════════════════════════════════
class TestExports:
    def test_day_report_pdf_export(self, salesman_token):
        r = requests.get(
            f"{BASE_URL}/api/salesman-orders/beat-run/day-report/export",
            headers={"Authorization": f"Bearer {salesman_token}"},
            params={"format": "pdf"},
            timeout=30,
        )
        # Could be 200 PDF, or 200 JSON if no run today
        ct = r.headers.get("content-type", "").lower()
        if "application/pdf" in ct:
            assert len(r.content) > 100
            assert r.content[:4] == b"%PDF"
        else:
            # JSON path — acceptable only when no run exists
            body = r.json()
            print(f"day-report PDF JSON fallback: {body}")
            # Endpoint must still respond, not 500
            assert r.status_code == 200

    def test_day_report_excel_export(self, salesman_token):
        r = requests.get(
            f"{BASE_URL}/api/salesman-orders/beat-run/day-report/export",
            headers={"Authorization": f"Bearer {salesman_token}"},
            params={"format": "excel"},
            timeout=30,
        )
        ct = r.headers.get("content-type", "").lower()
        if "spreadsheet" in ct or "officedocument" in ct or "xlsx" in ct:
            assert len(r.content) > 100
            # XLSX is a zip → starts with PK
            assert r.content[:2] == b"PK"
        else:
            body = r.json()
            print(f"day-report Excel JSON fallback: {body}")
            assert r.status_code == 200

    def test_monthly_report_excel_has_new_columns(self, admin_token):
        r = requests.get(
            f"{BASE_URL}/api/salesman-orders/beat-run/monthly-report/export",
            headers={"Authorization": f"Bearer {admin_token}"},
            params={"format": "excel"},
            timeout=60,
        )
        ct = r.headers.get("content-type", "").lower()
        if "spreadsheet" not in ct and "officedocument" not in ct:
            # Maybe JSON fallback when no data
            body = r.json() if r.status_code == 200 else {"raw": r.text}
            print(f"monthly export non-xlsx response: {body}")
            pytest.skip(f"monthly export returned {ct}, no xlsx to introspect")
        assert r.content[:2] == b"PK"

        # Introspect headers with openpyxl
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip("openpyxl not available")
        wb = load_workbook(io.BytesIO(r.content))
        sheet_names = wb.sheetnames
        print(f"Sheets: {sheet_names}")
        # Find "By Salesman" sheet
        target = None
        for s in sheet_names:
            if "salesman" in s.lower():
                target = s
                break
        assert target, f"No 'By Salesman' sheet found: {sheet_names}"
        ws = wb[target]
        # Title is often on row 1 (merged cell). Headers may be on row 2/3.
        # Scan first 5 rows for the new columns.
        all_text = []
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            all_text.extend([str(v) for v in row if v])
        joined = " | ".join(all_text)
        print(f"First 5 rows in {target}: {joined}")
        assert "Orders Collected" in joined
        assert "Payments Collected" in joined
        assert "Order Conv %" in joined
        assert "Payment Conv %" in joined
